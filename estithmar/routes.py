from __future__ import annotations

import csv
import io
import os
import secrets
from urllib.parse import urlencode
from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation

from dateutil.relativedelta import relativedelta
from flask import (
    Response,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_login import current_user, login_user, logout_user
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import func, or_
from sqlalchemy.orm import joinedload
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from estithmar import db
from estithmar.auth import admin_required, role_required
from estithmar.config import PROJECT_ROOT, resolve_static_folder
from estithmar.validators import validate_phone
from estithmar.services.member_notify import (
    notify_member_certificate_issued,
    notify_member_new_subscription,
    notify_member_payment,
    notify_member_profit_share,
)
from estithmar.services.notifications import (
    mail_configured,
    notify_member_welcome,
    notify_password_reset,
    notify_user_credentials,
    send_email,
)
from estithmar.models import (
    ELIGIBILITY_POLICIES,
    INVESTMENT_STATUSES,
    InvestmentLedger,
    MEMBER_DOCUMENT_TYPES,
    MEMBER_GENDER_CHOICES,
    MEMBER_KINDS,
    PAYMENT_PLANS,
    PROFIT_DISTRIBUTION_FREQUENCIES,
    PROJECT_CATEGORIES,
    PROJECT_STATUSES,
    USER_ROLES,
    Agent,
    AppSettings,
    AppUser,
    AuditLog,
    Contribution,
    EligibilitySnapshot,
    InstallmentPlan,
    Investment,
    InvestmentProfitLog,
    Member,
    MemberDocument,
    PaymentBank,
    PaymentBankAccount,
    PaymentMobileProvider,
    ProfitDistribution,
    ProfitDistributionBatch,
    ShareCertificate,
    ShareSubscription,
    Project,
    get_or_create_settings,
    next_agent_id,
    next_investment_code,
    next_member_id,
    next_profit_batch_no,
    next_project_code,
    next_receipt_no,
)
from estithmar.services import (
    auto_allocate_payment_to_installments,
    available_pool_for_investment,
    create_subscription,
    issue_certificate,
    max_payment_for_subscription,
    maybe_auto_issue_certificate,
    project_budget_headroom,
    recompute_installment_statuses,
    recompute_subscription_status,
    subscription_payment_running_rows,
    total_invested_across_investments,
    total_member_contributions_collected,
)
from estithmar.services.certificates import (
    certificate_share_position_detail,
    certificate_stock_of_name,
    format_certificate_share_quantity,
)
from estithmar.dashboard_geo import build_members_region_map_data
from estithmar.services.profit_distribution import (
    build_profit_distribution_preview,
    eligible_pools_for_investments,
    policy_label_for_batch,
    profit_basis_verified_only,
)
from estithmar.services.certificate_pdf import build_share_certificate_pdf
from estithmar.accounting_models import Account, JournalEntry, JournalLine
from estithmar.services.accounting_service import (
    JOURNAL_SOURCE_TYPES,
    SYSTEM_ACCOUNT_KEYS,
    accounting_enabled,
    ensure_chart_of_accounts,
    gl_posting_error_message,
    journal_entries_filtered,
    ledger_lines_for_account,
    lines_for_entry,
    post_capital_return_delta,
    post_contribution_unverified,
    post_contribution_verified,
    post_investment_deployment_delta,
    post_profit_distribution_batch,
    post_profit_recognition_delta,
    trial_balance_rows,
    void_manual_journal_entry,
)
from estithmar.share_policy import (
    MAX_SHARE_UNITS,
    MIN_SHARE_UNITS,
    SHARE_UNIT_PRICE,
    max_subscribed_amount,
    min_subscribed_amount,
)

# Member portal users: read-only, scoped data; block admin/ops/agent tools.
_MEMBER_PORTAL_BLOCKED_ENDPOINTS = frozenset(
    {
        "settings_page",
        "settings_notifications",
        "members_new",
        "members_edit",
        "subscriptions_new",
        "contributions_new",
        "projects_new",
        "projects_edit",
        "investments_new",
        "investments_edit",
        "investments_delete",
        "investments_ledger_snapshot",
        "profit_distribute",
        "profit_batch_detail",
        "audit_logs",
        "audit_export_csv",
        "certificates_issue",
        "certificates_revoke",
        "certificates_reinstate",
        "contributions_verify",
        "contributions_unverify",
        "subscriptions_set_investment",
        "subscriptions_cancel",
    }
)


def log_audit(action, entity_type=None, entity_id=None, details=None):
    db.session.add(
        AuditLog(action=action, entity_type=entity_type, entity_id=entity_id, details=details)
    )


def _parse_date(s: str | None) -> date | None:
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def _parse_decimal(s: str | None) -> Decimal | None:
    if not s:
        return None
    try:
        return Decimal(s.strip().replace(",", ""))
    except (InvalidOperation, ValueError):
        return None


def _accounts_grouped_for_journal(accounts) -> list[tuple[str, list]]:
    """Group chart accounts by type for optgroup labels in manual journal UI."""
    labels = {
        "asset": "Assets",
        "liability": "Liabilities",
        "equity": "Equity",
        "revenue": "Revenue",
        "expense": "Expenses",
    }
    order = ("asset", "liability", "equity", "revenue", "expense")
    buckets: dict[str, list] = {k: [] for k in order}
    for a in accounts:
        if a.account_type in buckets:
            buckets[a.account_type].append(a)
    return [(labels[t], buckets[t]) for t in order if buckets[t]]


def _valid_member_kind(value: str | None) -> str:
    allowed = {k for k, _ in MEMBER_KINDS}
    v = (value or "member").strip()
    return v if v in allowed else "member"


_MEMBER_GENDER_VALUES = frozenset(k for k, _ in MEMBER_GENDER_CHOICES)
_MEMBER_DOCUMENT_TYPE_KEYS = frozenset(k for k, _ in MEMBER_DOCUMENT_TYPES)
_MAX_MEMBER_DOCUMENT_BYTES = 15 * 1024 * 1024
_MEMBER_DOCUMENT_ALLOWED_EXT = frozenset({".pdf", ".png", ".jpg", ".jpeg", ".webp", ".gif"})


def _parse_member_personal_extended(form, *, require_all: bool) -> tuple[dict | None, str | None]:
    """Parse DOB, gender, occupation, next-of-kin from a form. If ``require_all``, every field is mandatory."""
    dob_s = (form.get("date_of_birth") or "").strip()
    gender = (form.get("gender") or "").strip().lower()
    occ = (form.get("occupation_employer") or "").strip()[:200]
    nok_name = (form.get("next_of_kin_name") or "").strip()[:200]
    nok_rel = (form.get("next_of_kin_relationship") or "").strip()[:100]
    nok_phone_raw = (form.get("next_of_kin_phone") or "").strip()
    nok_addr = (form.get("next_of_kin_address") or "").strip()

    if require_all:
        if not dob_s:
            return None, "Date of birth is required."
        if not gender:
            return None, "Gender is required."
        if gender not in _MEMBER_GENDER_VALUES:
            return None, "Please choose a valid gender option."
        if not occ:
            return None, "Occupation or employer is required."
        if not nok_name:
            return None, "Next of kin (emergency contact) name is required."
        if not nok_rel:
            return None, "Next of kin relationship is required."
        if not nok_phone_raw:
            return None, "Next of kin phone number is required."
        if not nok_addr:
            return None, "Next of kin address is required."
    elif gender and gender not in _MEMBER_GENDER_VALUES:
        return None, "Please choose a valid gender option."

    dob = _parse_date(dob_s) if dob_s else None
    if dob_s and dob is None:
        return None, "Invalid date of birth."
    if dob:
        if dob > date.today():
            return None, "Date of birth cannot be in the future."
        if dob.year < 1900:
            return None, "Date of birth must be year 1900 or later."

    nok_phone_norm, nok_err = validate_phone(nok_phone_raw)
    if nok_phone_raw and nok_err:
        return None, f"Next of kin phone: {nok_err}"
    if require_all and not nok_phone_norm:
        return None, "Next of kin phone number is required."

    return {
        "date_of_birth": dob,
        "gender": gender or None,
        "occupation_employer": occ or None,
        "next_of_kin_name": nok_name or None,
        "next_of_kin_relationship": nok_rel or None,
        "next_of_kin_phone": nok_phone_norm,
        "next_of_kin_address": nok_addr or None,
    }, None


def _apply_member_personal_fields(m: Member, data: dict) -> None:
    m.date_of_birth = data.get("date_of_birth")
    m.gender = data.get("gender")
    m.occupation_employer = data.get("occupation_employer")
    m.next_of_kin_name = data.get("next_of_kin_name")
    m.next_of_kin_relationship = data.get("next_of_kin_relationship")
    m.next_of_kin_phone = data.get("next_of_kin_phone")
    m.next_of_kin_address = data.get("next_of_kin_address")


def _allowed_member_document_filename(filename: str | None) -> bool:
    if not filename or "." not in filename:
        return False
    ext = os.path.splitext(filename)[1].lower()
    return ext in _MEMBER_DOCUMENT_ALLOWED_EXT


def _save_member_document_file(member_pk: int, storage) -> tuple[str, str] | None:
    if not storage or not storage.filename:
        return None
    if not _allowed_member_document_filename(storage.filename):
        return None
    safe = secure_filename(storage.filename)
    if not safe:
        return None
    rel_dir = os.path.join("assets", "documents", "members", str(member_pk))
    abs_dir = os.path.join(resolve_static_folder(), rel_dir)
    os.makedirs(abs_dir, exist_ok=True)
    _, ext = os.path.splitext(safe)
    ext = ext.lower() or ".bin"
    out_name = f"{int(datetime.utcnow().timestamp())}_{secrets.token_hex(4)}{ext}"
    dest = os.path.join(abs_dir, out_name)
    storage.save(dest)
    try:
        sz = os.path.getsize(dest)
    except OSError:
        sz = 0
    if sz > _MAX_MEMBER_DOCUMENT_BYTES:
        try:
            os.remove(dest)
        except OSError:
            pass
        return None
    rel = os.path.join(rel_dir, out_name).replace("\\", "/")
    return rel, safe


def _save_optional_new_member_documents(m: Member, req) -> None:
    """Persist up to two optional identity files from ``/members/new`` (after member row exists)."""
    any_saved = False
    for idx in (1, 2):
        f = req.files.get(f"new_member_doc_file_{idx}")
        if not f or not f.filename:
            continue
        dt = (req.form.get(f"new_member_doc_type_{idx}") or "").strip().lower()
        if dt not in _MEMBER_DOCUMENT_TYPE_KEYS:
            flash(
                f"Attachment {idx} was not saved: choose a document type when you select a file.",
                "warning",
            )
            continue
        notes = (req.form.get(f"new_member_doc_notes_{idx}") or "").strip()[:500] or None
        saved = _save_member_document_file(m.id, f)
        if not saved:
            flash(
                f"Attachment {idx} was not saved (PDF or image only, max 15 MB).",
                "warning",
            )
            continue
        rel, orig = saved
        db.session.add(
            MemberDocument(
                member_id=m.id,
                document_type=dt,
                stored_path=rel,
                original_name=orig[:255],
                notes=notes,
                uploaded_by_user_id=current_user.id if current_user.is_authenticated else None,
            )
        )
        any_saved = True
    if any_saved:
        log_audit(
            "member_document_uploaded",
            "MemberDocument",
            None,
            f"member_id={m.id} registration_attachments",
        )
        db.session.commit()


def _valid_project_status(value: str | None) -> str:
    allowed = {k for k, _ in PROJECT_STATUSES}
    v = (value or "Planned").strip()
    return v if v in allowed else "Planned"


def _valid_investment_status(value: str | None) -> str:
    allowed = {k for k, _ in INVESTMENT_STATUSES}
    v = (value or "Planned").strip()
    return v if v in allowed else "Planned"


def _pool_verified_only() -> bool:
    """When True, pooled funds use only verified contributions (Settings)."""
    return bool(get_or_create_settings().get_extra().get("pool_use_verified_contributions"))


def _funds_pool_summary():
    """For investment forms: collected vs allocated vs available."""
    v = _pool_verified_only()
    collected = total_member_contributions_collected(verified_only=v)
    invested = total_invested_across_investments()
    return {
        "collected": collected,
        "invested": invested,
        "available": collected - invested,
        "verified_only": v,
    }


def _safe_brand_static_filename(v) -> str | None:
    """Relative path under the static folder; None if unusable (avoids url_for BuildError)."""
    if not isinstance(v, str):
        return None
    s = v.strip().replace("\\", "/")
    if not s or ".." in s or "://" in s or s.startswith("/"):
        return None
    return s


def _extra_for_settings_template(ex: dict) -> dict:
    """Drop invalid logo paths so settings preview img tags do not break rendering."""
    out = dict(ex)
    for k in ("logo_light", "logo_dark"):
        if _safe_brand_static_filename(out.get(k)) is None:
            out.pop(k, None)
    return out


def _member_scoped_member_pk() -> int | None:
    if current_user.is_authenticated and current_user.role == "member" and current_user.member_id:
        return int(current_user.member_id)
    return None


def _member_investment_ids_for_portal() -> list[int]:
    mid = _member_scoped_member_pk()
    if not mid:
        return []
    sub_ids = (
        db.session.query(Investment.id)
        .join(ShareSubscription, ShareSubscription.investment_id == Investment.id)
        .filter(ShareSubscription.member_id == mid)
        .distinct()
        .all()
    )
    pr_ids = (
        db.session.query(ProfitDistribution.investment_id)
        .filter(ProfitDistribution.member_id == mid)
        .distinct()
        .all()
    )
    out = {row[0] for row in sub_ids if row[0]}
    out.update(row[0] for row in pr_ids if row[0])
    return list(out)


def _member_project_ids_for_portal() -> list[int]:
    iids = _member_investment_ids_for_portal()
    if not iids:
        return []
    rows = Investment.query.filter(Investment.id.in_(iids)).with_entities(Investment.project_id).all()
    return list({r[0] for r in rows if r[0]})


def _dashboard_scoped_member_ids():
    """None = all members; list = restrict to these member PKs (agent scope)."""
    if current_user.is_authenticated and current_user.role == "agent" and current_user.agent_id:
        return [m.id for m in Member.query.filter_by(agent_id=current_user.agent_id).all()]
    mid = _member_scoped_member_pk()
    if mid is not None:
        return [mid]
    return None


def _dashboard_monthly_totals(member_ids):
    labels = []
    values = []
    today = date.today()
    for i in range(11, -1, -1):
        d = today - relativedelta(months=i)
        start = date(d.year, d.month, 1)
        end = start + relativedelta(months=1) - relativedelta(days=1)
        q = db.session.query(func.coalesce(func.sum(Contribution.amount), 0)).filter(
            Contribution.date >= start,
            Contribution.date <= end,
        )
        if member_ids is not None:
            if not member_ids:
                val = 0
            else:
                val = q.filter(Contribution.member_id.in_(member_ids)).scalar() or 0
        else:
            val = q.scalar() or 0
        labels.append(d.strftime("%b %y"))
        values.append(float(val))
    return labels, values


def _dashboard_payment_totals(member_ids):
    """Cash, Mobile, Bank, Other — matches contribution form payment types."""

    def sum_type(ptype: str) -> float:
        q = db.session.query(func.coalesce(func.sum(Contribution.amount), 0)).filter(
            Contribution.payment_type == ptype
        )
        if member_ids is not None:
            if not member_ids:
                return 0.0
            q = q.filter(Contribution.member_id.in_(member_ids))
        return float(q.scalar() or 0)

    return (
        sum_type("Cash"),
        sum_type("Mobile"),
        sum_type("Bank"),
        sum_type("Other"),
    )


def _dashboard_recent_contributions(member_ids, limit=12):
    q = (
        Contribution.query.join(Member, Contribution.member_id == Member.id)
        .options(
            joinedload(Contribution.payment_bank_account).joinedload(PaymentBankAccount.bank),
            joinedload(Contribution.payment_mobile_provider),
        )
        .order_by(Contribution.date.desc(), Contribution.id.desc())
    )
    if member_ids is not None:
        if not member_ids:
            return []
        q = q.filter(Contribution.member_id.in_(member_ids))
    return q.limit(limit).all()


def _fmt_time_ago(dt: datetime | None) -> str:
    if not dt:
        return ""
    now = datetime.utcnow()
    if getattr(dt, "tzinfo", None) is not None:
        dt = dt.replace(tzinfo=None)
    secs = int((now - dt).total_seconds())
    if secs < 0:
        secs = 0
    if secs < 60:
        return "Just now"
    if secs < 3600:
        m = secs // 60
        return f"{m} min ago" if m != 1 else "1 min ago"
    if secs < 86400:
        h = secs // 3600
        return f"{h} h ago" if h != 1 else "1 h ago"
    if secs < 86400 * 7:
        d = secs // 86400
        return f"{d} d ago" if d != 1 else "1 day ago"
    return dt.strftime("%d %b %Y")


def _money_display(v) -> str:
    try:
        return f"{float(v):,.2f}"
    except (TypeError, ValueError):
        return str(v)


def _audit_details_kv(details: str | None) -> dict[str, str]:
    """Best-effort key=value pairs from audit log details (space / comma / semicolon separated)."""
    if not details:
        return {}
    out: dict[str, str] = {}
    for token in details.replace(";", " ").replace(",", " ").split():
        token = token.strip()
        if not token or "=" not in token:
            continue
        k, _, v = token.partition("=")
        k, v = k.strip(), v.strip()
        if k:
            out[k] = v
    return out


def _audit_href(log: AuditLog) -> str:
    """Resolve a deep link for an audit row; fall back to audit log list when unknown."""
    et, eid = log.entity_type, log.entity_id
    kv = _audit_details_kv(log.details)
    try:
        if log.action == "certificate_issued" and log.details:
            sid_raw = kv.get("subscription_id")
            if sid_raw:
                try:
                    sid = int(sid_raw)
                    return url_for("subscriptions_profile", id=sid)
                except ValueError:
                    pass
            cno = kv.get("certificate_no")
            if cno:
                cert = ShareCertificate.query.filter_by(certificate_no=cno).first()
                if cert:
                    return url_for("certificates_print", id=cert.id)
            return url_for("certificates_list")

        if log.action == "contribution_recorded" and kv.get("member_id"):
            try:
                mid = int(kv["member_id"])
                if Member.query.get(mid):
                    return url_for("members_profile", id=mid)
            except ValueError:
                pass

        if log.action == "installment_added" and kv.get("subscription_id"):
            try:
                sid = int(kv["subscription_id"])
                return url_for("subscriptions_installments", id=sid)
            except ValueError:
                pass

        if log.action == "subscription_created":
            sno = kv.get("subscription_no")
            if sno:
                sub = ShareSubscription.query.filter_by(subscription_no=sno).first()
                if sub:
                    return url_for("subscriptions_profile", id=sub.id)
            return url_for("subscriptions_list")

        if log.action == "member_created":
            mid = kv.get("member_id")
            if mid and not eid:
                m = Member.query.filter_by(member_id=mid).first()
                if m:
                    return url_for("members_profile", id=m.id)
            return url_for("members_list")

        if log.action == "investment_created":
            return url_for("investments_profile", id=eid) if eid else url_for("investments_list")
        if log.action == "project_created":
            return url_for("projects_profile", id=eid) if eid else url_for("projects_list")

        if log.action == "agent_created":
            if eid:
                return url_for("agents_profile", id=eid)
            code = (log.details or "").strip()
            if code and "=" not in code:
                ag = Agent.query.filter_by(agent_id=code).first()
                if ag:
                    return url_for("agents_profile", id=ag.id)
            return url_for("agents_list")

        if et == "AppUser" and eid:
            return url_for("users_edit", id=eid)
        if et == "Member" and eid:
            return url_for("members_profile", id=eid)
        if et == "ShareSubscription" and eid:
            return url_for("subscriptions_profile", id=eid)
        if et == "ShareCertificate" and eid:
            return url_for("certificates_print", id=eid)
        if et == "Investment" and eid:
            return url_for("investments_profile", id=eid)
        if et == "Project" and eid:
            return url_for("projects_profile", id=eid)
        if et == "Agent" and eid:
            return url_for("agents_profile", id=eid)
        if et == "ProfitDistributionBatch" and eid:
            return url_for("profit_batch_detail", batch_id=eid)
        if log.action == "settings_notifications_updated":
            return url_for("settings_notifications")
        if et == "AppSettings":
            return url_for("settings_page")
        if et == "InstallmentPlan" and eid:
            row = InstallmentPlan.query.get(eid)
            if row:
                return url_for("subscriptions_installments", id=row.subscription_id)
        if et == "InstallmentPlan" and kv.get("subscription_id"):
            try:
                sid = int(kv["subscription_id"])
                return url_for("subscriptions_installments", id=sid)
            except ValueError:
                pass
    except Exception:
        pass
    return url_for("audit_logs")


def _audit_icon_color(action: str) -> tuple[str, str]:
    if "profit" in action:
        return "ri-share-forward-line", "success"
    if "certificate" in action:
        return "ri-award-line", "warning"
    if "investment" in action:
        return "ri-line-chart-line", "primary"
    if "subscription" in action or "installment" in action:
        return "ri-stack-line", "info"
    if "member" in action or "agent" in action:
        return "ri-team-line", "secondary"
    if "settings" in action or "branding" in action:
        return "ri-settings-3-line", "primary"
    return "ri-notification-line", "primary"


def _audit_notification_item(log: AuditLog) -> dict | None:
    if log.action in (
        "contribution_recorded",
        "user_profile_updated",
        "subscription_created",
    ):
        return None
    title_map = {
        "certificate_issued": "Share certificate issued",
        "member_created": "New member registered",
        "member_updated": "Member updated",
        "member_document_uploaded": "Member document uploaded",
        "member_document_deleted": "Member document removed",
        "subscription_cancelled": "Subscription cancelled",
        "subscription_investment_updated": "Subscription investment link updated",
        "settings_branding_updated": "Branding / logos updated",
        "settings_notifications_updated": "Notification settings updated",
        "profit_distributed": "Profit distribution recorded",
        "investment_created": "Investment created",
        "investment_updated": "Investment updated",
        "investment_deleted": "Investment removed",
        "project_created": "Project created",
        "project_updated": "Project updated",
        "certificate_revoked": "Certificate revoked",
        "certificate_reinstated": "Certificate reinstated",
        "agent_created": "Agent added",
        "agent_updated": "Agent updated",
        "installment_added": "Installment scheduled",
        "installment_updated": "Installment updated",
        "installment_cancelled": "Installment cancelled",
    }
    title = title_map.get(log.action, log.action.replace("_", " ").title())
    body = (log.details or "").strip()
    if len(body) > 140:
        body = body[:137] + "…"
    icon, color = _audit_icon_color(log.action)
    when = log.created_at or datetime.utcnow()
    return {
        "title": title,
        "body": body or None,
        "when": when,
        "icon": icon,
        "color": color,
        "href": _audit_href(log),
    }


def build_header_notifications(mids: list[int] | None) -> dict:
    """Recent activity for the dashboard notifications dropdown (scoped by member list for agents)."""
    from flask_login import current_user

    sym = get_or_create_settings().currency_symbol or "$"
    raw: list[dict] = []

    cq = Contribution.query.options(
        joinedload(Contribution.member),
        joinedload(Contribution.payment_bank_account).joinedload(PaymentBankAccount.bank),
        joinedload(Contribution.payment_mobile_provider),
    )
    if mids is not None:
        if not mids:
            cq = cq.filter(Contribution.id == -1)
        else:
            cq = cq.filter(Contribution.member_id.in_(mids))
    cq = cq.order_by(Contribution.created_at.desc()).limit(12)
    for c in cq:
        m = c.member
        label = (m.full_name or m.member_id or f"Member #{m.id}").strip()
        amt = _money_display(c.amount)
        body = f"{label} · {sym}{amt} ({c.payment_display_label()})"
        if c.receipt_no:
            body += f" · {c.receipt_no}"
        when = c.created_at or datetime.combine(c.date, time.min)
        raw.append(
            {
                "title": "Contribution recorded",
                "body": body,
                "when": when,
                "icon": "ri-hand-coin-line",
                "color": "primary",
                "href": url_for("members_profile", id=c.member_id),
            }
        )

    sq = ShareSubscription.query.options(joinedload(ShareSubscription.member))
    if mids is not None:
        if not mids:
            sq = sq.filter(ShareSubscription.id == -1)
        else:
            sq = sq.filter(ShareSubscription.member_id.in_(mids))
    sq = sq.order_by(ShareSubscription.created_at.desc()).limit(10)
    for sub in sq:
        m = sub.member
        label = (m.full_name or m.member_id or f"Member #{m.id}").strip()
        raw.append(
            {
                "title": "Share subscription",
                "body": f"{sub.subscription_no} · {label} · {sub.status}",
                "when": sub.created_at or datetime.utcnow(),
                "icon": "ri-stack-line",
                "color": "info",
                "href": url_for("subscriptions_profile", id=sub.id),
            }
        )

    if current_user.is_authenticated and getattr(current_user, "role", None) == "admin":
        for log in AuditLog.query.order_by(AuditLog.created_at.desc()).limit(28):
            item = _audit_notification_item(log)
            if item:
                raw.append(item)

    raw.sort(key=lambda x: x["when"], reverse=True)
    out: list[dict] = []
    for r in raw[:14]:
        out.append(
            {
                "title": r["title"],
                "body": r.get("body"),
                "time_ago": _fmt_time_ago(r["when"]),
                "icon": r["icon"],
                "color": r.get("color", "primary"),
                "href": r["href"],
            }
        )

    if current_user.is_authenticated and getattr(current_user, "role", None) == "admin":
        view_all = url_for("audit_logs")
    else:
        view_all = url_for("contributions_list")

    return {"items": out, "view_all_url": view_all, "count": len(out)}


def _dashboard_contribution_range_total(member_ids, start: date, end: date) -> float:
    q = db.session.query(func.coalesce(func.sum(Contribution.amount), 0)).filter(
        Contribution.date >= start,
        Contribution.date <= end,
    )
    if member_ids is not None:
        if not member_ids:
            return 0.0
        q = q.filter(Contribution.member_id.in_(member_ids))
    return float(q.scalar() or 0)


def _dashboard_mtd_vs_last_month(member_ids):
    """Month-to-date total vs full previous calendar month (for KPI context)."""
    today = date.today()
    month_start = date(today.year, today.month, 1)
    mtd = _dashboard_contribution_range_total(member_ids, month_start, today)
    prev_month_end = month_start - relativedelta(days=1)
    prev_month_start = date(prev_month_end.year, prev_month_end.month, 1)
    last_month = _dashboard_contribution_range_total(member_ids, prev_month_start, prev_month_end)
    return mtd, last_month


def _dashboard_active_members_count(member_ids) -> int:
    q = Member.query.filter_by(status="Active")
    if member_ids is not None:
        if not member_ids:
            return 0
        q = q.filter(Member.id.in_(member_ids))
    return q.count()


def _dashboard_contributions_today_total(member_ids) -> Decimal:
    today = date.today()
    q = db.session.query(func.coalesce(func.sum(Contribution.amount), 0)).filter(Contribution.date == today)
    if member_ids is not None:
        if not member_ids:
            return Decimal("0")
        q = q.filter(Contribution.member_id.in_(member_ids))
    return Decimal(str(q.scalar() or 0))


def _contrib_pct_mtd_vs_last_month(mtd: Decimal, last: Decimal) -> float | None:
    last_f = float(last)
    mtd_f = float(mtd)
    if last_f > 0.005:
        return ((mtd_f - last_f) / last_f) * 100.0
    return None


def _dashboard_top_members_by_volume(member_ids, limit: int = 5):
    """Top members by contribution sum. Uses a subquery so SQL Server accepts the GROUP BY (unlike PostgreSQL functional dependency)."""
    vol_sum = func.coalesce(func.sum(Contribution.amount), 0).label("vol")
    agg = db.session.query(Contribution.member_id.label("mid"), vol_sum).group_by(Contribution.member_id)
    if member_ids is not None:
        if not member_ids:
            return []
        agg = agg.filter(Contribution.member_id.in_(member_ids))
    subq = agg.subquery()
    q = (
        db.session.query(Member, subq.c.vol)
        .join(subq, Member.id == subq.c.mid)
        .order_by(subq.c.vol.desc())
        .limit(limit)
    )
    return q.all()


def register_routes(app):
    def _allowed_profile_image(filename: str) -> bool:
        if "." not in filename:
            return False
        ext = filename.rsplit(".", 1)[1].lower()
        return ext in {"png", "jpg", "jpeg", "gif", "webp"}

    def _remove_brand_file(relative_path: str | None) -> None:
        if not relative_path:
            return
        full = os.path.join(resolve_static_folder(), relative_path.replace("/", os.sep))
        if os.path.isfile(full):
            try:
                os.remove(full)
            except OSError:
                pass

    def _save_brand_logo(file_storage, prefix: str) -> str | None:
        if not file_storage or not file_storage.filename:
            return None
        if not _allowed_profile_image(file_storage.filename):
            return None
        safe = secure_filename(file_storage.filename)
        _, ext = os.path.splitext(safe)
        ext = ext.lower() or ".png"
        out_name = f"{prefix}-{int(datetime.utcnow().timestamp())}{ext}"
        rel_dir = os.path.join("assets", "images", "branding")
        abs_dir = os.path.join(resolve_static_folder(), rel_dir)
        os.makedirs(abs_dir, exist_ok=True)
        dest = os.path.join(abs_dir, out_name)
        file_storage.save(dest)
        return os.path.join(rel_dir, out_name).replace("\\", "/")

    @app.context_processor
    def inject_branding():
        ex = get_or_create_settings().get_extra()
        lt = _safe_brand_static_filename(ex.get("logo_light"))
        dk = _safe_brand_static_filename(ex.get("logo_dark"))

        def _u(custom: str | None, default_path: str) -> str:
            path = custom or default_path
            try:
                return url_for("static", filename=path)
            except Exception:
                return url_for("static", filename=default_path)

        return {
            "brand_logo_light_sm": _u(lt, "assets/images/logo-sm-light.png"),
            "brand_logo_light_lg": _u(lt, "assets/images/logo-light.png"),
            "brand_logo_dark_sm": _u(dk, "assets/images/logo-sm-dark.png"),
            "brand_logo_dark_lg": _u(dk, "assets/images/logo-dark.png"),
        }

    def members_scope():
        q = Member.query
        if current_user.is_authenticated and current_user.role == "agent" and current_user.agent_id:
            q = q.filter(Member.agent_id == current_user.agent_id)
        mpk = _member_scoped_member_pk()
        if mpk is not None:
            q = q.filter(Member.id == mpk)
        return q

    def _members_filtered_query():
        """Apply the same GET filters as the members list (for list view + Excel export)."""
        q = request.args.get("q", "").strip()
        st = request.args.get("status", "").strip()
        kind_f = request.args.get("kind", "").strip()
        agent_f = request.args.get("agent_id", type=int)
        query = members_scope()
        if q:
            like = f"%{q}%"
            query = query.filter(
                or_(
                    Member.member_id.ilike(like),
                    Member.full_name.ilike(like),
                    Member.phone.ilike(like),
                    Member.national_id.ilike(like),
                    Member.email.ilike(like),
                )
            )
        if st in ("Active", "Inactive"):
            query = query.filter(Member.status == st)
        if kind_f in {k for k, _ in MEMBER_KINDS}:
            query = query.filter(Member.member_kind == kind_f)
        if agent_f and current_user.role != "agent":
            query = query.filter(Member.agent_id == agent_f)
        return query

    def contributions_scope():
        q = Contribution.query.join(Member, Contribution.member_id == Member.id)
        if current_user.is_authenticated and current_user.role == "agent" and current_user.agent_id:
            q = q.filter(Member.agent_id == current_user.agent_id)
        mpk = _member_scoped_member_pk()
        if mpk is not None:
            q = q.filter(Member.id == mpk)
        return q

    def _contributions_filtered_query():
        """Apply request query-string filters to contributions (same rules for list + Excel export)."""
        member_id = request.args.get("member_id", type=int)
        agent_id = request.args.get("agent_id", type=int)
        subscription_id = request.args.get("subscription_id", type=int)
        date_from = _parse_date(request.args.get("date_from"))
        date_to = _parse_date(request.args.get("date_to"))
        verified = (request.args.get("verified") or "").strip().lower()
        qtext = (request.args.get("q") or "").strip()
        payment_type = (request.args.get("payment_type") or "").strip()
        query = contributions_scope()
        if payment_type in {"Cash", "Mobile", "Bank", "Other"}:
            query = query.filter(Contribution.payment_type == payment_type)
        if member_id:
            query = query.filter(Contribution.member_id == member_id)
        if agent_id:
            query = query.filter(Member.agent_id == agent_id)
        if subscription_id:
            query = query.filter(Contribution.subscription_id == subscription_id)
        if date_from:
            query = query.filter(Contribution.date >= date_from)
        if date_to:
            query = query.filter(Contribution.date <= date_to)
        if verified == "yes":
            query = query.filter(Contribution.verified)
        elif verified == "no":
            query = query.filter(~Contribution.verified)
        if qtext:
            like = f"%{qtext}%"
            query = query.filter(
                or_(
                    Contribution.receipt_no.ilike(like),
                    Member.member_id.ilike(like),
                    Member.full_name.ilike(like),
                    Contribution.notes.ilike(like),
                )
            )
        return query

    def _projects_filtered_query():
        """Apply GET filters for project list and Excel export (same rules)."""
        category = request.args.get("category", "").strip()
        status = request.args.get("status", "").strip()
        qtext = (request.args.get("q") or "").strip()
        date_from = _parse_date(request.args.get("date_from"))
        date_to = _parse_date(request.args.get("date_to"))
        query = Project.query
        cat_keys = {k for k, _ in PROJECT_CATEGORIES}
        st_keys = {k for k, _ in PROJECT_STATUSES}
        if category in cat_keys:
            query = query.filter(Project.category == category)
        if status in st_keys:
            query = query.filter(Project.status == status)
        if qtext:
            like = f"%{qtext}%"
            query = query.filter(
                or_(
                    Project.name.ilike(like),
                    Project.project_code.ilike(like),
                    Project.project_manager.ilike(like),
                    Project.description.ilike(like),
                )
            )
        if date_from:
            query = query.filter(or_(Project.start_date.is_(None), Project.start_date >= date_from))
        if date_to:
            query = query.filter(or_(Project.start_date.is_(None), Project.start_date <= date_to))
        mpk = _member_scoped_member_pk()
        if mpk is not None:
            pids = _member_project_ids_for_portal()
            if pids:
                query = query.filter(Project.id.in_(pids))
            else:
                query = query.filter(Project.id < 0)
        return query

    def _investments_filtered_query():
        """Apply GET filters for investment list and Excel export (same rules)."""
        project_id = request.args.get("project_id", type=int)
        status = request.args.get("status", "").strip()
        qtext = (request.args.get("q") or "").strip()
        date_from = _parse_date(request.args.get("date_from"))
        date_to = _parse_date(request.args.get("date_to"))
        query = Investment.query
        if project_id:
            query = query.filter(Investment.project_id == project_id)
        st_keys = {k for k, _ in INVESTMENT_STATUSES}
        if status in st_keys:
            query = query.filter(Investment.status == status)
        if date_from:
            query = query.filter(or_(Investment.start_date.is_(None), Investment.start_date >= date_from))
        if date_to:
            query = query.filter(or_(Investment.start_date.is_(None), Investment.start_date <= date_to))
        if qtext:
            like = f"%{qtext}%"
            query = query.outerjoin(Project, Investment.project_id == Project.id).filter(
                or_(
                    Investment.name.ilike(like),
                    Investment.investment_code.ilike(like),
                    Investment.investment_type.ilike(like),
                    Investment.project_manager.ilike(like),
                    Project.name.ilike(like),
                    Project.project_code.ilike(like),
                )
            )
        mpk = _member_scoped_member_pk()
        if mpk is not None:
            iids = _member_investment_ids_for_portal()
            if iids:
                query = query.filter(Investment.id.in_(iids))
            else:
                query = query.filter(Investment.id < 0)
        return query

    def profit_rows_scope():
        q = ProfitDistribution.query.join(Member, ProfitDistribution.member_id == Member.id).join(
            Investment, ProfitDistribution.investment_id == Investment.id
        )
        if current_user.is_authenticated and current_user.role == "agent" and current_user.agent_id:
            q = q.filter(Member.agent_id == current_user.agent_id)
        mpk = _member_scoped_member_pk()
        if mpk is not None:
            q = q.filter(Member.id == mpk)
        return q

    def forbid_agent():
        if current_user.role in ("agent", "member"):
            flash("This section is not available for your role.", "warning")
            return redirect(url_for("dashboard"))
        return None

    def _audit_filtered_query():
        """Apply GET filters for audit log list and CSV export (same rules)."""
        action_f = (request.args.get("action") or "").strip()
        entity_type = (request.args.get("entity_type") or "").strip()
        date_from = _parse_date(request.args.get("date_from"))
        date_to = _parse_date(request.args.get("date_to"))
        qtext = (request.args.get("q") or "").strip()
        query = AuditLog.query
        if action_f:
            query = query.filter(AuditLog.action.ilike(f"%{action_f}%"))
        if entity_type:
            query = query.filter(AuditLog.entity_type == entity_type)
        if date_from:
            query = query.filter(AuditLog.created_at >= datetime.combine(date_from, time.min))
        if date_to:
            query = query.filter(AuditLog.created_at <= datetime.combine(date_to, time(23, 59, 59)))
        if qtext:
            like = f"%{qtext}%"
            query = query.filter(
                or_(
                    AuditLog.action.ilike(like),
                    AuditLog.entity_type.ilike(like),
                    AuditLog.details.ilike(like),
                )
            )
        return query

    @app.before_request
    def _require_login():
        if request.endpoint in (
            None,
            "login",
            "static",
            "admin_api_notifications_unread_count",
            "auth_recoverpw",
            "register",
        ):
            return
        if not current_user.is_authenticated:
            return redirect(url_for("login", next=request.path))

    @app.before_request
    def _member_portal_guard():
        if not current_user.is_authenticated or current_user.role != "member":
            return
        ep = request.endpoint
        if not ep or ep == "static":
            return
        if ep in _MEMBER_PORTAL_BLOCKED_ENDPOINTS:
            flash("This action is not available for your account.", "warning")
            return redirect(url_for("dashboard"))
        if ep.startswith(("agents_", "users_", "accounting_", "export_")):
            flash("This section is not available for your account.", "warning")
            return redirect(url_for("dashboard"))
        if ep.startswith("reports_"):
            flash("Reports are not available for your account.", "warning")
            return redirect(url_for("dashboard"))
        if ep == "subscriptions_installments" and request.method == "POST":
            flash("You cannot modify installment plans.", "warning")
            return redirect(url_for("dashboard"))

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for("dashboard"))
        if request.method == "POST":
            ident = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            nxt = request.form.get("next") or request.args.get("next") or ""
            u = None
            if ident:
                u = AppUser.query.filter(func.lower(AppUser.username) == ident.lower()).first()
                if not u and "@" in ident:
                    u = AppUser.query.filter(func.lower(AppUser.email) == ident.lower()).first()
            if u and u.is_active and check_password_hash(u.password_hash, password):
                login_user(u)
                if nxt.startswith("/"):
                    return redirect(nxt)
                return redirect(url_for("dashboard"))
            flash("Invalid email/username or password.", "danger")
        return render_template("login.html", next=request.args.get("next", ""))

    def _unique_username_from_email(email: str) -> str:
        base = email.lower().strip()[:64]
        if not AppUser.query.filter_by(username=base).first():
            return base
        for i in range(1, 1000):
            suf = f"_{i}"
            candidate = (base[: 64 - len(suf)] + suf)[:64]
            if not AppUser.query.filter_by(username=candidate).first():
                return candidate
        return (base[:48] + secrets.token_hex(4))[:64]

    def _allow_self_registration() -> bool:
        return (os.environ.get("ESTITHMAR_ALLOW_SELF_REGISTRATION") or "").strip().lower() in (
            "1",
            "true",
            "yes",
            "on",
        )

    @app.route("/register", methods=["GET", "POST"])
    def register():
        if current_user.is_authenticated:
            return redirect(url_for("dashboard"))
        if not _allow_self_registration():
            flash("Self-registration is disabled. Contact the office to open an account.", "warning")
            return redirect(url_for("login"))
        if request.method == "POST":
            email = (request.form.get("email") or "").strip().lower()
            pwd = request.form.get("password") or ""
            fn = (request.form.get("full_name") or "").strip()
            nid = (request.form.get("national_id") or "").strip()
            phone_raw = (request.form.get("phone") or "").strip()
            if not email or "@" not in email or len(email) < 5:
                flash("A valid email address is required.", "danger")
                return render_template("register.html")
            if len(pwd) < 4:
                flash("Password must be at least 4 characters.", "danger")
                return render_template("register.html")
            if not fn:
                flash("Full name is required.", "danger")
                return render_template("register.html")
            if not nid:
                flash("National ID or official document number is required.", "danger")
                return render_template("register.html")
            phone_norm, phone_err = validate_phone(phone_raw)
            if phone_err:
                flash(phone_err, "danger")
                return render_template("register.html")
            personal, perr = _parse_member_personal_extended(request.form, require_all=True)
            if perr:
                flash(perr, "danger")
                return render_template("register.html")
            if AppUser.query.filter(func.lower(AppUser.email) == email).first():
                flash("An account with this email already exists. Sign in or use Forgot password.", "warning")
                return render_template("register.html")
            s = get_or_create_settings()
            if s.get_flag("require_agent_on_member"):
                flash(
                    "Registration currently requires an assigned agent. Contact the office to complete enrollment.",
                    "warning",
                )
                return render_template("register.html")
            m = Member(
                member_id=next_member_id(),
                full_name=fn,
                phone=phone_norm,
                email=email[:120],
                national_id=nid,
                join_date=date.today(),
                status="Active",
                member_kind="member",
                agent_id=None,
            )
            _apply_member_personal_fields(m, personal)
            db.session.add(m)
            db.session.flush()
            u = AppUser(
                username=_unique_username_from_email(email),
                email=email[:120],
                password_hash=generate_password_hash(pwd),
                full_name=fn,
                phone=phone_norm,
                role="member",
                member_id=m.id,
                is_active=True,
            )
            db.session.add(u)
            db.session.commit()
            log_audit("member_self_registered", "Member", m.id, f"member_id={m.member_id}")
            notify_member_welcome(
                member_name=fn,
                member_code=m.member_id,
                email=email,
                phone=phone_norm,
            )
            login_user(u)
            flash("Welcome — your member portal is ready.", "success")
            return redirect(url_for("dashboard"))
        return render_template("register.html")

    @app.route("/auth-recoverpw.html", methods=["GET", "POST"])
    def auth_recoverpw():
        temp_password = None
        if request.method == "POST":
            email = request.form.get("email", "").strip().lower()
            if not email:
                flash("Email or username is required.", "danger")
                return render_template("auth-recoverpw.html", temp_password=temp_password)
            u = AppUser.query.filter(
                or_(func.lower(AppUser.email) == email, func.lower(AppUser.username) == email)
            ).first()
            if not u:
                flash("No account found with that email/username.", "warning")
                return render_template("auth-recoverpw.html", temp_password=temp_password)
            temp_password = secrets.token_urlsafe(6)[:10]
            u.password_hash = generate_password_hash(temp_password)
            db.session.commit()
            if u.email and mail_configured():
                notify_password_reset(to_email=u.email, temp_password=temp_password)
                flash("If your account exists, check your email for a temporary password.", "success")
                return redirect(url_for("login"))
            flash("Password reset successful.", "success")
            return render_template("auth-recoverpw.html", temp_password=temp_password)
        return render_template("auth-recoverpw.html", temp_password=temp_password)

    @app.route("/logout")
    def logout():
        logout_user()
        flash("Signed out.", "info")
        return redirect(url_for("login"))

    @app.route("/pages-profile.html", methods=["GET", "POST"])
    def pages_profile():
        u = AppUser.query.options(joinedload(AppUser.agent), joinedload(AppUser.member)).get_or_404(
            current_user.id
        )

        def _profile_ctx():
            settings = get_or_create_settings()
            return {
                "user": u,
                "settings": settings,
                "currency_code": settings.currency_code or "USD",
                "role_labels": dict(USER_ROLES),
            }

        if request.method == "POST":
            new_username = request.form.get("username", "").strip()
            new_full_name = request.form.get("full_name", "").strip()
            new_email = request.form.get("email", "").strip()
            new_phone = request.form.get("phone", "").strip()
            phone_norm, phone_err = validate_phone(new_phone)
            if phone_err:
                flash(phone_err, "danger")
                return render_template("pages-profile.html", **_profile_ctx())
            new_address = request.form.get("address", "").strip()
            pwd = request.form.get("password", "").strip()

            if not new_username:
                flash("Username is required.", "danger")
                return render_template("pages-profile.html", **_profile_ctx())
            exists = AppUser.query.filter(AppUser.username == new_username, AppUser.id != u.id).first()
            if exists:
                flash("Username already in use.", "danger")
                return render_template("pages-profile.html", **_profile_ctx())

            if new_email:
                if "@" not in new_email or len(new_email) < 5:
                    flash("Please enter a valid email address.", "danger")
                    return render_template("pages-profile.html", **_profile_ctx())
                em_ex = AppUser.query.filter(AppUser.email == new_email, AppUser.id != u.id).first()
                if em_ex:
                    flash("That email is already in use by another account.", "danger")
                    return render_template("pages-profile.html", **_profile_ctx())
                u.email = new_email[:120]
            else:
                u.email = None

            u.username = new_username
            u.full_name = new_full_name or None
            u.phone = phone_norm
            u.address = new_address or None
            if pwd:
                if len(pwd) < 4:
                    flash("Password must be at least 4 characters.", "danger")
                    return render_template("pages-profile.html", **_profile_ctx())
                u.password_hash = generate_password_hash(pwd)

            f = request.files.get("profile_image")
            if f and f.filename:
                if not _allowed_profile_image(f.filename):
                    flash("Invalid image type. Use png, jpg, jpeg, gif, or webp.", "danger")
                    return render_template("pages-profile.html", **_profile_ctx())
                safe = secure_filename(f.filename)
                _, ext = os.path.splitext(safe)
                out_name = f"user-{u.id}-{int(datetime.utcnow().timestamp())}{ext.lower()}"
                rel_dir = os.path.join("assets", "images", "users")
                # Must match Flask static_folder (see resolve_static_folder), not PROJECT_ROOT/static
                abs_dir = os.path.join(resolve_static_folder(), rel_dir)
                os.makedirs(abs_dir, exist_ok=True)
                f.save(os.path.join(abs_dir, out_name))
                u.profile_image = os.path.join(rel_dir, out_name).replace("\\", "/")

            db.session.commit()
            log_audit("user_profile_updated", "AppUser", u.id, f"username={u.username}")
            flash("Profile updated.", "success")
            return redirect(url_for("pages_profile"))

        return render_template("pages-profile.html", **_profile_ctx())

    @app.route("/")
    def index():
        return redirect(url_for("dashboard"))

    @app.route("/admin/api/notifications/unread-count")
    def admin_api_notifications_unread_count():
        """JSON for header notification badge polling (matches theme / IDE expectations)."""
        if not current_user.is_authenticated:
            return jsonify(count=0)
        mids = _dashboard_scoped_member_ids()
        hn = build_header_notifications(mids)
        return jsonify(count=hn["count"])

    @app.route("/dashboard")
    def dashboard():
        app_settings = get_or_create_settings()
        total_agents = Agent.query.count()
        total_members = Member.query.count()
        total_collected = db.session.query(func.coalesce(func.sum(Contribution.amount), 0)).scalar() or 0
        total_invested = (
            db.session.query(func.coalesce(func.sum(Investment.total_amount_invested), 0)).scalar() or 0
        )
        total_profit = (
            db.session.query(func.coalesce(func.sum(Investment.profit_generated), 0)).scalar() or 0
        )
        total_projects = Project.query.count()
        active_investments_count = Investment.query.filter_by(status="Active").count()
        completed_investments_count = Investment.query.filter_by(status="Completed").count()

        top_agents = []
        if current_user.role != "agent":
            for ag in Agent.query.filter_by(status="Active").all():
                vol = ag.contributions_managed_total()
                top_agents.append((ag, vol))
            top_agents.sort(key=lambda x: x[1], reverse=True)
            top_agents = top_agents[:8]

        if current_user.role == "agent" and current_user.agent_id:
            total_members = members_scope().count()
            total_collected = (
                db.session.query(func.coalesce(func.sum(Contribution.amount), 0))
                .join(Member)
                .filter(Member.agent_id == current_user.agent_id)
                .scalar()
                or 0
            )
            total_agents = 1
            top_agents = []

        mids = _dashboard_scoped_member_ids()
        sub_query = ShareSubscription.query
        if mids is not None:
            if mids:
                sub_query = sub_query.filter(ShareSubscription.member_id.in_(mids))
            else:
                sub_query = sub_query.filter(ShareSubscription.id == -1)
        scoped_subscriptions = sub_query.all()
        total_subscriptions = len(scoped_subscriptions)
        total_share_subscribed_value = sum(
            (s.subscribed_amount or Decimal("0") for s in scoped_subscriptions), Decimal("0")
        )
        total_confirmed_share_value = sum(
            (s.subscribed_amount or Decimal("0") for s in scoped_subscriptions if s.status == "Fully Paid"),
            Decimal("0"),
        )
        subscriptions_outstanding_total = sum(
            (s.outstanding_balance() for s in scoped_subscriptions), Decimal("0")
        )

        inst_query = InstallmentPlan.query.join(
            ShareSubscription, InstallmentPlan.subscription_id == ShareSubscription.id
        )
        if mids is not None:
            if mids:
                inst_query = inst_query.filter(ShareSubscription.member_id.in_(mids))
            else:
                inst_query = inst_query.filter(InstallmentPlan.id == -1)
        inst_rows = inst_query.all()
        installment_overdue_count = 0
        installment_due_balance = Decimal("0")
        today = date.today()
        for row in inst_rows:
            if row.status == "Cancelled":
                continue
            bal = (row.due_amount or Decimal("0")) - (row.paid_amount or Decimal("0"))
            if bal > 0:
                installment_due_balance += bal
                is_overdue = row.status == "Overdue" or (
                    row.due_date is not None
                    and row.due_date < today
                    and row.status in {"Pending", "Partially Paid"}
                )
                if is_overdue:
                    installment_overdue_count += 1

        chart_labels, chart_values = _dashboard_monthly_totals(mids)
        donut_cash, donut_mobile, donut_bank, donut_other = _dashboard_payment_totals(mids)
        recent_contributions = _dashboard_recent_contributions(mids)
        mtd_contributions, last_month_contributions = _dashboard_mtd_vs_last_month(mids)
        active_members = _dashboard_active_members_count(mids)
        total_profit_distributed = Decimal(
            str(
                profit_rows_scope()
                .with_entities(func.coalesce(func.sum(ProfitDistribution.amount), 0))
                .scalar()
                or 0
            )
        )
        cert_query = ShareCertificate.query.join(Member, ShareCertificate.member_id == Member.id)
        if mids is not None:
            if mids:
                cert_query = cert_query.filter(ShareCertificate.member_id.in_(mids))
            else:
                cert_query = cert_query.filter(ShareCertificate.id == -1)
        certificates_issued = cert_query.filter(ShareCertificate.status == "Issued").count()
        pending_certificates = sum(
            1
            for s in scoped_subscriptions
            if s.status == "Fully Paid" and (s.certificate is None or s.certificate.status != "Issued")
        )
        chart_sum = sum(chart_values)
        chart_has_data = chart_sum > 0.005

        hn = build_header_notifications(mids)

        mtd_dec = Decimal(str(mtd_contributions))
        last_dec = Decimal(str(last_month_contributions))
        contrib_pct_vs_last_mtd = _contrib_pct_mtd_vs_last_month(mtd_dec, last_dec)
        contributions_today = _dashboard_contributions_today_total(mids)
        top_members = _dashboard_top_members_by_volume(mids, limit=5)
        subscription_status_counts = dict(Counter(s.status for s in scoped_subscriptions))

        if mids is not None and not mids:
            active_investors_count = 0
        else:
            qai = (
                db.session.query(func.count(func.distinct(Member.id)))
                .select_from(Member)
                .join(ShareSubscription, ShareSubscription.member_id == Member.id)
                .filter(Member.status == "Active", ShareSubscription.status != "Cancelled")
            )
            if mids is not None:
                qai = qai.filter(Member.id.in_(mids))
            active_investors_count = int(qai.scalar() or 0)

        ta_int = int(total_agents) if total_agents else 0
        avg_members_per_agent = (
            (Decimal(str(total_members)) / Decimal(str(ta_int))).quantize(Decimal("0.01")) if ta_int else Decimal("0")
        )

        agent_member_breakdown = []
        if current_user.role != "agent":
            for ag in Agent.query.order_by(Agent.full_name).all():
                agent_member_breakdown.append((ag, Member.query.filter_by(agent_id=ag.id).count()))

        def _fnum(x):
            try:
                return float(x or 0)
            except (TypeError, ValueError):
                return 0.0

        business_kpi_bar = {
            "categories": [
                "Collected",
                "Invested",
                "Profit (gen.)",
                "Share subscribed",
                "Subs. outstanding",
                "Installment due",
            ],
            "values": [
                _fnum(total_collected),
                _fnum(total_invested),
                _fnum(total_profit),
                _fnum(total_share_subscribed_value),
                _fnum(subscriptions_outstanding_total),
                _fnum(installment_due_balance),
            ],
        }
        inv_a = int(active_investments_count or 0)
        inv_c = int(completed_investments_count or 0)
        business_kpi_inv_donut = {"labels": ["Active investments", "Completed"], "series": [inv_a, inv_c]}
        cert_p = int(pending_certificates or 0)
        cert_i = int(certificates_issued or 0)
        business_kpi_cert_donut = {"labels": ["Pending issue", "Issued"], "series": [cert_p, cert_i]}

        members_region_map = build_members_region_map_data(
            mids,
            user_role=getattr(current_user, "role", None) if current_user.is_authenticated else None,
            user_agent_id=getattr(current_user, "agent_id", None) if current_user.is_authenticated else None,
        )

        perf_month_links: list[str] = []
        for i in range(11, -1, -1):
            d = date.today() - relativedelta(months=i)
            mstart = date(d.year, d.month, 1)
            mend = mstart + relativedelta(months=1) - relativedelta(days=1)
            perf_month_links.append(
                url_for("contributions_list")
                + "?"
                + urlencode({"date_from": mstart.isoformat(), "date_to": mend.isoformat()})
            )
        dashboard_chart_links = {
            "performance_months": perf_month_links,
            "donut_payment": [
                url_for("contributions_list", payment_type="Cash"),
                url_for("contributions_list", payment_type="Mobile"),
                url_for("contributions_list", payment_type="Bank"),
                url_for("contributions_list", payment_type="Other"),
            ],
            "business_kpi_bar": [
                url_for("contributions_list"),
                url_for("investments_list"),
                url_for("investments_list"),
                url_for("subscriptions_list"),
                url_for("subscriptions_list", status="Partially Paid"),
                url_for("subscriptions_list", overdue=1),
            ],
            "investments_donut": [
                url_for("investments_list", status="Active"),
                url_for("investments_list", status="Completed"),
            ],
            "certificates_donut": [
                url_for("subscriptions_list", certificate_pending=1),
                url_for("certificates_list"),
            ],
        }

        return render_template(
            "index.html",
            today=date.today(),
            settings=app_settings,
            total_members=total_members,
            total_agents=total_agents,
            total_collected=Decimal(str(total_collected)),
            total_invested=Decimal(str(total_invested)),
            total_profit=Decimal(str(total_profit)),
            total_projects=total_projects,
            top_agents=top_agents,
            chart_labels=chart_labels,
            chart_values=chart_values,
            donut_cash=donut_cash,
            donut_mobile=donut_mobile,
            donut_bank=donut_bank,
            donut_other=donut_other,
            active_investors_count=active_investors_count,
            avg_members_per_agent=avg_members_per_agent,
            agent_member_breakdown=agent_member_breakdown,
            recent_contributions=recent_contributions,
            mtd_contributions=mtd_dec,
            last_month_contributions=last_dec,
            contrib_pct_vs_last_mtd=contrib_pct_vs_last_mtd,
            contributions_today=contributions_today,
            top_members=top_members,
            subscription_status_counts=subscription_status_counts,
            active_members=active_members,
            chart_has_data=chart_has_data,
            total_subscriptions=total_subscriptions,
            total_share_subscribed_value=total_share_subscribed_value,
            total_confirmed_share_value=total_confirmed_share_value,
            subscriptions_outstanding_total=subscriptions_outstanding_total,
            installment_overdue_count=installment_overdue_count,
            installment_due_balance=installment_due_balance,
            total_profit_distributed=total_profit_distributed,
            pending_certificates=pending_certificates,
            certificates_issued=certificates_issued,
            active_investments_count=active_investments_count,
            completed_investments_count=completed_investments_count,
            header_notifications=hn["items"],
            header_notifications_view_all=hn["view_all_url"],
            header_notifications_count=hn["count"],
            business_kpi_bar=business_kpi_bar,
            business_kpi_inv_donut=business_kpi_inv_donut,
            business_kpi_cert_donut=business_kpi_cert_donut,
            members_region_map=members_region_map,
            dashboard_chart_links=dashboard_chart_links,
        )

    # --- Agents ---
    def _agents_filtered_query():
        q = request.args.get("q", "").strip()
        st = request.args.get("status", "").strip()
        query = Agent.query
        if q:
            like = f"%{q}%"
            query = query.filter(
                or_(
                    Agent.agent_id.ilike(like),
                    Agent.full_name.ilike(like),
                    Agent.phone.ilike(like),
                    Agent.email.ilike(like),
                    Agent.region.ilike(like),
                    Agent.country.ilike(like),
                )
            )
        if st in ("Active", "Inactive"):
            query = query.filter(Agent.status == st)
        return query

    def _parse_optional_agent_email(raw: str | None) -> tuple[str | None, str | None]:
        s = (raw or "").strip()
        if not s:
            return None, None
        if "@" not in s or len(s) < 5:
            return None, "Enter a valid email or leave the field empty."
        return s[:120], None

    def _batch_agent_metrics(agent_ids: list[int]) -> dict[int, dict]:
        """One round of aggregates per metric — avoids N+1 per-row totals on the agents table."""
        if not agent_ids:
            return {}
        out: dict[int, dict] = {
            i: {"members": 0, "contrib": Decimal("0"), "share": Decimal("0")} for i in agent_ids
        }
        for aid, cnt in (
            db.session.query(Member.agent_id, func.count(Member.id))
            .filter(Member.agent_id.in_(agent_ids))
            .group_by(Member.agent_id)
            .all()
        ):
            if aid in out:
                out[aid]["members"] = int(cnt)
        for aid, v in (
            db.session.query(Member.agent_id, func.coalesce(func.sum(Contribution.amount), 0))
            .select_from(Contribution)
            .join(Member, Contribution.member_id == Member.id)
            .filter(Member.agent_id.in_(agent_ids))
            .group_by(Member.agent_id)
            .all()
        ):
            if aid in out:
                out[aid]["contrib"] = Decimal(str(v))
        for aid, v in (
            db.session.query(Member.agent_id, func.coalesce(func.sum(ShareSubscription.subscribed_amount), 0))
            .select_from(ShareSubscription)
            .join(Member, ShareSubscription.member_id == Member.id)
            .filter(Member.agent_id.in_(agent_ids))
            .filter(ShareSubscription.status != "Cancelled")
            .group_by(Member.agent_id)
            .all()
        ):
            if aid in out:
                out[aid]["share"] = Decimal(str(v))
        return out

    def _agent_form_ctx(agent: Agent | None):
        return {"agent": agent, "is_edit": agent is not None}

    @app.route("/agents")
    def agents_list():
        r = forbid_agent()
        if r:
            return r
        q = request.args.get("q", "").strip()
        st = request.args.get("status", "").strip()
        sort = request.args.get("sort", "created_desc").strip()
        page = max(1, request.args.get("page", 1, type=int) or 1)
        per_page = min(max(10, request.args.get("per_page", 50, type=int) or 50), 200)

        base_q = _agents_filtered_query()
        total = base_q.count()
        ordered = _agents_filtered_query()
        if sort == "name_asc":
            ordered = ordered.order_by(Agent.full_name.asc(), Agent.id.asc())
        elif sort == "name_desc":
            ordered = ordered.order_by(Agent.full_name.desc(), Agent.id.desc())
        elif sort == "agent_id":
            ordered = ordered.order_by(Agent.agent_id.asc())
        elif sort == "created_asc":
            ordered = ordered.order_by(Agent.created_at.asc(), Agent.id.asc())
        else:
            ordered = ordered.order_by(Agent.created_at.desc(), Agent.id.desc())
        total_pages = max(1, (total + per_page - 1) // per_page) if total else 1
        page = min(max(1, page), total_pages)
        agents = ordered.offset((page - 1) * per_page).limit(per_page).all()
        metrics = _batch_agent_metrics([a.id for a in agents])

        all_filtered_ids = [r[0] for r in base_q.with_entities(Agent.id).all()]
        rollup_all = _batch_agent_metrics(all_filtered_ids) if all_filtered_ids else {}
        filtered_totals = {
            "members": sum((v["members"] for v in rollup_all.values()), 0),
            "contrib": sum((v["contrib"] for v in rollup_all.values()), Decimal("0")),
            "share": sum((v["share"] for v in rollup_all.values()), Decimal("0")),
        }
        page_members = sum((metrics.get(a.id, {}).get("members", 0) for a in agents), 0)
        page_contrib = sum((metrics.get(a.id, {}).get("contrib", Decimal("0")) for a in agents), Decimal("0"))
        page_share = sum((metrics.get(a.id, {}).get("share", Decimal("0")) for a in agents), Decimal("0"))
        row_start = (page - 1) * per_page + 1 if total else 0
        row_end = min(page * per_page, total) if total else 0

        def _agents_list_qs(page_num: int) -> str:
            d = request.args.to_dict(flat=True)
            d["page"] = str(page_num)
            return url_for("agents_list") + "?" + urlencode(d)

        qs_no_page = request.args.to_dict(flat=True)
        qs_no_page.pop("page", None)
        export_url = url_for("export_agents_xlsx") + "?" + urlencode(qs_no_page)
        next_page_url = _agents_list_qs(page + 1) if page < total_pages else None
        prev_page_url = _agents_list_qs(page - 1) if page > 1 else None
        first_page_url = _agents_list_qs(1) if page > 1 else None
        last_page_url = _agents_list_qs(total_pages) if total_pages > 1 and page < total_pages else None

        active_count = Agent.query.filter_by(status="Active").count()
        settings = get_or_create_settings()
        return render_template(
            "agents/list.html",
            agents=agents,
            metrics=metrics,
            q=q,
            status_filter=st,
            sort=sort,
            page=page,
            per_page=per_page,
            total_count=total,
            total_pages=total_pages,
            export_url=export_url,
            next_page_url=next_page_url,
            prev_page_url=prev_page_url,
            first_page_url=first_page_url,
            last_page_url=last_page_url,
            active_count=active_count,
            filtered_totals=filtered_totals,
            page_members=page_members,
            page_contrib=page_contrib,
            page_share=page_share,
            row_start=row_start,
            row_end=row_end,
            currency_code=settings.currency_code or "USD",
        )

    @app.route("/agents/new", methods=["GET", "POST"])
    def agents_new():
        r = forbid_agent()
        if r:
            return r
        if request.method == "POST":
            phone_norm, phone_err = validate_phone(request.form.get("phone", "").strip())
            if phone_err:
                flash(phone_err, "danger")
                return render_template("agents/form.html", **_agent_form_ctx(None))
            em, em_err = _parse_optional_agent_email(request.form.get("email"))
            if em_err:
                flash(em_err, "danger")
                return render_template("agents/form.html", **_agent_form_ctx(None))
            a = Agent(
                agent_id=next_agent_id(),
                full_name=request.form.get("full_name", "").strip(),
                phone=phone_norm,
                email=em,
                region=request.form.get("region", "").strip(),
                territory=request.form.get("territory", "").strip(),
                country=request.form.get("country", "").strip(),
                status=request.form.get("status") or "Active",
            )
            if not a.full_name:
                flash("Full name is required.", "danger")
                return render_template("agents/form.html", **_agent_form_ctx(None))
            db.session.add(a)
            log_audit("agent_created", "Agent", None, a.agent_id)
            db.session.commit()
            flash(f"Agent registered: {a.agent_id}", "success")
            return redirect(url_for("agents_profile", id=a.id))
        return render_template("agents/form.html", **_agent_form_ctx(None))

    @app.route("/agents/<int:id>")
    def agents_profile(id):
        r = forbid_agent()
        if r:
            return r
        a = Agent.query.get_or_404(id)
        member_rows = a.members.order_by(Member.join_date.desc()).limit(50).all()
        member_count = a.members_count()
        active_members_count = Member.query.filter_by(agent_id=a.id, status="Active").count()
        active_investors_count = (
            db.session.query(Member.id)
            .join(ShareSubscription, ShareSubscription.member_id == Member.id)
            .filter(
                Member.agent_id == a.id,
                Member.status == "Active",
                ShareSubscription.status != "Cancelled",
            )
            .distinct()
            .count()
        )
        contrib_total = a.contributions_managed_total()
        share_total = a.total_subscribed_share_value()
        share_units = a.total_share_units_subscribed()
        settings = get_or_create_settings()
        cur = settings.currency_code or "USD"
        return render_template(
            "agents/profile.html",
            agent=a,
            members_preview=member_rows,
            member_count=member_count,
            active_members_count=active_members_count,
            active_investors_count=active_investors_count,
            contrib_total=contrib_total,
            share_total=share_total,
            share_units=share_units,
            settings=settings,
            currency_code=cur,
            member_kinds=MEMBER_KINDS,
        )

    @app.route("/agents/<int:id>/edit", methods=["GET", "POST"])
    def agents_edit(id):
        r = forbid_agent()
        if r:
            return r
        a = Agent.query.get_or_404(id)
        if request.method == "POST":
            phone_norm, phone_err = validate_phone(request.form.get("phone", "").strip())
            if phone_err:
                flash(phone_err, "danger")
                return render_template("agents/form.html", **_agent_form_ctx(a))
            em, em_err = _parse_optional_agent_email(request.form.get("email"))
            if em_err:
                flash(em_err, "danger")
                return render_template("agents/form.html", **_agent_form_ctx(a))
            old_status = a.status
            a.full_name = request.form.get("full_name", "").strip()
            a.phone = phone_norm
            a.email = em
            a.region = request.form.get("region", "").strip()
            a.territory = request.form.get("territory", "").strip()
            a.country = request.form.get("country", "").strip()
            a.status = request.form.get("status") or "Active"
            if not a.full_name:
                db.session.rollback()
                a = Agent.query.get_or_404(id)
                flash("Full name is required.", "danger")
                return render_template("agents/form.html", **_agent_form_ctx(a))
            log_audit("agent_updated", "Agent", a.id)
            db.session.commit()
            flash("Agent updated.", "success")
            if old_status == "Active" and a.status == "Inactive":
                n_m = Member.query.filter_by(agent_id=a.id).count()
                if n_m:
                    flash(
                        f"This agent is inactive but {n_m} member(s) still reference them. "
                        "Update member assignments when roles change.",
                        "info",
                    )
            return redirect(url_for("agents_profile", id=a.id))
        return render_template("agents/form.html", **_agent_form_ctx(a))

    # --- Members ---
    @app.route("/members")
    def members_list():
        q = request.args.get("q", "").strip()
        st = request.args.get("status", "").strip()
        kind_f = request.args.get("kind", "").strip()
        agent_f = request.args.get("agent_id", type=int)
        sort = request.args.get("sort", "join_desc").strip()
        page = max(1, request.args.get("page", 1, type=int) or 1)
        per_page = min(max(10, request.args.get("per_page", 50, type=int) or 50), 200)

        base_q = _members_filtered_query()
        total = base_q.count()
        ordered = base_q.options(joinedload(Member.agent))
        if sort == "name_asc":
            ordered = ordered.order_by(Member.full_name.asc(), Member.id.asc())
        elif sort == "name_desc":
            ordered = ordered.order_by(Member.full_name.desc(), Member.id.desc())
        elif sort == "member_id":
            ordered = ordered.order_by(Member.member_id.asc())
        else:
            ordered = ordered.order_by(Member.join_date.desc(), Member.id.desc())
        total_pages = max(1, (total + per_page - 1) // per_page) if total else 1
        members = ordered.offset((page - 1) * per_page).limit(per_page).all()

        agents = (
            Agent.query.filter_by(status="Active").order_by(Agent.full_name).all()
            if current_user.role != "agent"
            else []
        )

        def _members_list_qs(page_num: int) -> str:
            d = request.args.to_dict(flat=True)
            d["page"] = str(page_num)
            return url_for("members_list") + "?" + urlencode(d)

        qs_no_page = request.args.to_dict(flat=True)
        qs_no_page.pop("page", None)
        export_url = url_for("export_members_xlsx") + "?" + urlencode(qs_no_page)
        next_page_url = _members_list_qs(page + 1) if page < total_pages else None
        prev_page_url = _members_list_qs(page - 1) if page > 1 else None

        return render_template(
            "members/list.html",
            members=members,
            q=q,
            status_filter=st,
            kind_filter=kind_f,
            agent_filter=agent_f if current_user.role != "agent" else None,
            agents=agents,
            member_kinds=MEMBER_KINDS,
            sort=sort,
            page=page,
            per_page=per_page,
            total_count=total,
            total_pages=total_pages,
            export_url=export_url,
            next_page_url=next_page_url,
            prev_page_url=prev_page_url,
        )

    def _agents_for_member_form(member: Member | None):
        """Active agents for new members; for edits, include current assignee even if inactive."""
        if member and member.agent_id:
            return (
                Agent.query.filter(or_(Agent.status == "Active", Agent.id == member.agent_id))
                .order_by(Agent.full_name)
                .all()
            )
        return Agent.query.filter_by(status="Active").order_by(Agent.full_name).all()

    def _member_form_ctx(member: Member | None, agents: list):
        s = get_or_create_settings()
        return {
            "member": member,
            "agents": agents,
            "member_kinds": MEMBER_KINDS,
            "member_genders": MEMBER_GENDER_CHOICES,
            "member_document_types": MEMBER_DOCUMENT_TYPES,
            "require_agent": s.get_flag("require_agent_on_member"),
            "is_edit": member is not None,
            "default_join_date": date.today().isoformat(),
        }

    @app.route("/members/new", methods=["GET", "POST"])
    def members_new():
        agents = _agents_for_member_form(None)
        if current_user.role == "agent" and not current_user.agent_id:
            flash("Your account is not linked to an agent profile. Contact an administrator.", "danger")
            return redirect(url_for("dashboard"))
        if request.method == "POST":
            aid = request.form.get("agent_id", type=int) or None
            if current_user.role == "agent" and current_user.agent_id:
                aid = current_user.agent_id
            phone_norm, phone_err = validate_phone(request.form.get("phone", "").strip())
            if phone_err:
                flash(phone_err, "danger")
                return render_template("members/form.html", **_member_form_ctx(None, agents))
            em_raw = (request.form.get("email") or "").strip()
            if em_raw and ("@" not in em_raw or len(em_raw) < 5):
                flash("Please enter a valid email or leave it empty.", "danger")
                return render_template("members/form.html", **_member_form_ctx(None, agents))
            personal, perr = _parse_member_personal_extended(request.form, require_all=True)
            if perr:
                flash(perr, "danger")
                return render_template("members/form.html", **_member_form_ctx(None, agents))
            m = Member(
                member_id=next_member_id(),
                full_name=request.form.get("full_name", "").strip(),
                phone=phone_norm,
                email=em_raw[:120] if em_raw else None,
                address=request.form.get("address", "").strip(),
                national_id=request.form.get("national_id", "").strip(),
                join_date=_parse_date(request.form.get("join_date")) or date.today(),
                status=request.form.get("status") or "Active",
                member_kind=_valid_member_kind(request.form.get("member_kind")),
                agent_id=aid,
            )
            _apply_member_personal_fields(m, personal)
            if not m.full_name:
                flash("Full name is required.", "danger")
                return render_template("members/form.html", **_member_form_ctx(None, agents))
            if not m.national_id:
                flash("National ID or official document number is required.", "danger")
                return render_template("members/form.html", **_member_form_ctx(None, agents))
            if get_or_create_settings().get_flag("require_agent_on_member") and not m.agent_id:
                flash("Assign an agent to this member (required by system settings).", "danger")
                return render_template("members/form.html", **_member_form_ctx(None, agents))
            db.session.add(m)
            log_audit("member_created", "Member", None, f"member_id={m.member_id}")
            db.session.commit()
            _save_optional_new_member_documents(m, request)
            notify_member_welcome(
                member_name=m.full_name,
                member_code=m.member_id,
                email=m.email,
                phone=m.phone,
            )
            flash(f"Registered ({m.member_kind}): {m.member_id}", "success")
            return redirect(url_for("members_profile", id=m.id))
        return render_template("members/form.html", **_member_form_ctx(None, agents))

    @app.route("/members/<int:id>/edit", methods=["GET", "POST"])
    def members_edit(id):
        m = members_scope().options(joinedload(Member.agent)).filter_by(id=id).first_or_404()
        agents = _agents_for_member_form(m)
        if request.method == "POST":
            phone_norm, phone_err = validate_phone(request.form.get("phone", "").strip())
            if phone_err:
                flash(phone_err, "danger")
                agents = _agents_for_member_form(m)
                return render_template("members/form.html", **_member_form_ctx(m, agents))
            m.full_name = request.form.get("full_name", "").strip()
            m.phone = phone_norm
            em_raw = (request.form.get("email") or "").strip()
            if em_raw and ("@" not in em_raw or len(em_raw) < 5):
                flash("Please enter a valid email or leave it empty.", "danger")
                agents = _agents_for_member_form(m)
                return render_template("members/form.html", **_member_form_ctx(m, agents))
            m.email = em_raw[:120] if em_raw else None
            m.address = request.form.get("address", "").strip()
            m.national_id = request.form.get("national_id", "").strip()
            personal, perr = _parse_member_personal_extended(request.form, require_all=False)
            if perr:
                flash(perr, "danger")
                agents = _agents_for_member_form(m)
                return render_template("members/form.html", **_member_form_ctx(m, agents))
            _apply_member_personal_fields(m, personal)
            jd = _parse_date(request.form.get("join_date"))
            if jd:
                m.join_date = jd
            m.status = request.form.get("status") or "Active"
            m.member_kind = _valid_member_kind(request.form.get("member_kind"))
            if current_user.role == "agent" and current_user.agent_id:
                pass
            else:
                m.agent_id = request.form.get("agent_id", type=int) or None
            if not m.full_name:
                db.session.rollback()
                m = members_scope().options(joinedload(Member.agent)).filter_by(id=id).first_or_404()
                agents = _agents_for_member_form(m)
                flash("Full name is required.", "danger")
                return render_template("members/form.html", **_member_form_ctx(m, agents))
            if not m.national_id:
                db.session.rollback()
                m = members_scope().options(joinedload(Member.agent)).filter_by(id=id).first_or_404()
                agents = _agents_for_member_form(m)
                flash("National ID or official document number is required.", "danger")
                return render_template("members/form.html", **_member_form_ctx(m, agents))
            if (
                get_or_create_settings().get_flag("require_agent_on_member")
                and not m.agent_id
                and not (current_user.role == "agent" and current_user.agent_id)
            ):
                db.session.rollback()
                m = members_scope().options(joinedload(Member.agent)).filter_by(id=id).first_or_404()
                agents = _agents_for_member_form(m)
                flash("Assign an agent to this member (required by system settings).", "danger")
                return render_template("members/form.html", **_member_form_ctx(m, agents))
            log_audit("member_updated", "Member", m.id)
            db.session.commit()
            flash("Member updated.", "success")
            return redirect(url_for("members_profile", id=m.id))
        return render_template("members/form.html", **_member_form_ctx(m, agents))

    @app.route("/members/<int:id>")
    def members_profile(id):
        m = members_scope().options(joinedload(Member.agent)).filter_by(id=id).first_or_404()
        contrib_limit = 30
        contribution_count = m.contributions.count()
        contribs = (
            m.contributions.order_by(Contribution.date.desc(), Contribution.id.desc())
            .options(
                joinedload(Contribution.subscription),
                joinedload(Contribution.payment_bank_account).joinedload(PaymentBankAccount.bank),
                joinedload(Contribution.payment_mobile_provider),
            )
            .limit(contrib_limit)
            .all()
        )
        subs = m.subscriptions.filter(ShareSubscription.status != "Cancelled").all()
        total_subscribed = sum((s.subscribed_amount or Decimal("0") for s in subs), Decimal("0"))
        total_paid = sum((s.paid_total() for s in subs), Decimal("0"))
        outstanding = sum((s.outstanding_balance() for s in subs), Decimal("0"))
        confirmed_owned = sum(
            (s.subscribed_amount or Decimal("0") for s in subs if s.status == "Fully Paid"),
            Decimal("0"),
        )
        confirmed_share_units = sum(
            (s.share_units_subscribed or Decimal("0") for s in subs if s.status == "Fully Paid"),
            Decimal("0"),
        )
        subscriptions_ordered = (
            m.subscriptions.order_by(ShareSubscription.subscription_date.desc(), ShareSubscription.id.desc()).all()
        )
        profit_rows = (
            m.profit_rows.options(joinedload(ProfitDistribution.investment))
            .order_by(ProfitDistribution.distribution_date.desc(), ProfitDistribution.id.desc())
            .limit(25)
            .all()
        )
        cert_count = m.certificates.count()
        recent_certs = (
            m.certificates.options(joinedload(ShareCertificate.subscription))
            .order_by(ShareCertificate.issued_date.desc(), ShareCertificate.id.desc())
            .limit(5)
            .all()
        )
        settings = get_or_create_settings()
        member_documents = (
            m.documents.options(joinedload(MemberDocument.uploaded_by))
            .order_by(MemberDocument.uploaded_at.desc(), MemberDocument.id.desc())
            .all()
        )
        allow_doc_upload = current_user.role in ("admin", "operator", "agent") or (
            current_user.role == "member" and current_user.member_id == m.id
        )
        allow_doc_delete = current_user.role in ("admin", "operator", "agent")
        return render_template(
            "members/profile.html",
            member=m,
            contributions=contribs,
            contribution_count=contribution_count,
            contrib_limit=contrib_limit,
            contribution_total=m.contribution_total(),
            profit_total=m.lifetime_profit_received(),
            total_subscribed=total_subscribed,
            total_paid=total_paid,
            outstanding=outstanding,
            confirmed_owned=confirmed_owned,
            confirmed_share_units=confirmed_share_units,
            subscriptions=subscriptions_ordered,
            profit_rows=profit_rows,
            member_kinds=MEMBER_KINDS,
            member_genders=MEMBER_GENDER_CHOICES,
            member_document_types=MEMBER_DOCUMENT_TYPES,
            member_documents=member_documents,
            allow_member_document_upload=allow_doc_upload,
            allow_member_document_delete=allow_doc_delete,
            cert_count=cert_count,
            recent_certs=recent_certs,
            settings=settings,
        )

    @app.route("/members/<int:member_id>/documents", methods=["POST"])
    def members_document_upload(member_id):
        m = members_scope().filter_by(id=member_id).first_or_404()
        allowed = current_user.role in ("admin", "operator", "agent") or (
            current_user.role == "member" and current_user.member_id == m.id
        )
        if not allowed:
            flash("You cannot upload documents for this member.", "danger")
            return redirect(url_for("dashboard"))
        clen = request.content_length
        if clen is not None and clen > _MAX_MEMBER_DOCUMENT_BYTES:
            flash("File is too large (maximum 15 MB).", "danger")
            return redirect(url_for("members_profile", id=member_id))
        doc_type = (request.form.get("document_type") or "").strip().lower()
        if doc_type not in _MEMBER_DOCUMENT_TYPE_KEYS:
            flash("Choose a document type.", "danger")
            return redirect(url_for("members_profile", id=member_id))
        notes = (request.form.get("notes") or "").strip()[:500] or None
        f = request.files.get("file")
        saved = _save_member_document_file(m.id, f)
        if not saved:
            flash("Upload a valid file (PDF or image: PNG, JPG, WebP, GIF).", "danger")
            return redirect(url_for("members_profile", id=member_id))
        rel, orig = saved
        row = MemberDocument(
            member_id=m.id,
            document_type=doc_type,
            stored_path=rel,
            original_name=orig[:255],
            notes=notes,
            uploaded_by_user_id=current_user.id if current_user.is_authenticated else None,
        )
        db.session.add(row)
        log_audit("member_document_uploaded", "MemberDocument", None, f"member_id={m.id} type={doc_type}")
        db.session.commit()
        flash("Document uploaded.", "success")
        return redirect(url_for("members_profile", id=member_id))

    @app.route("/members/<int:member_id>/documents/<int:doc_id>/download")
    def members_document_download(member_id, doc_id):
        members_scope().filter_by(id=member_id).first_or_404()
        doc = MemberDocument.query.filter_by(id=doc_id, member_id=member_id).first_or_404()
        full = os.path.join(resolve_static_folder(), doc.stored_path.replace("/", os.sep))
        if not os.path.isfile(full):
            flash("File is missing on the server.", "danger")
            return redirect(url_for("members_profile", id=member_id))
        return send_file(
            full,
            as_attachment=True,
            download_name=doc.original_name,
            max_age=0,
        )

    @app.route("/members/<int:member_id>/documents/<int:doc_id>/delete", methods=["POST"])
    def members_document_delete(member_id, doc_id):
        if current_user.role not in ("admin", "operator", "agent"):
            flash("Only staff can remove stored documents.", "warning")
            return redirect(url_for("members_profile", id=member_id))
        m = members_scope().filter_by(id=member_id).first_or_404()
        doc = MemberDocument.query.filter_by(id=doc_id, member_id=m.id).first_or_404()
        full = os.path.join(resolve_static_folder(), doc.stored_path.replace("/", os.sep))
        db.session.delete(doc)
        log_audit("member_document_deleted", "MemberDocument", doc_id, f"member_id={m.id}")
        db.session.commit()
        if os.path.isfile(full):
            try:
                os.remove(full)
            except OSError:
                pass
        flash("Document removed.", "info")
        return redirect(url_for("members_profile", id=member_id))

    # --- Share subscriptions ---
    @app.route("/subscriptions")
    def subscriptions_list():
        member_id = request.args.get("member_id", type=int)
        filter_agent_id = request.args.get("agent_id", type=int)
        investment_id = request.args.get("investment_id", type=int)
        uninvested_only = request.args.get("uninvested", "").strip().lower() in {"1", "true", "yes"}
        status = request.args.get("status", "").strip()
        overdue_only = request.args.get("overdue", "").strip().lower() in {"1", "true", "yes"}
        certificate_pending = request.args.get("certificate_pending", "").strip().lower() in {
            "1",
            "true",
            "yes",
        }
        date_from = _parse_date(request.args.get("date_from"))
        date_to = _parse_date(request.args.get("date_to"))
        qtext = (request.args.get("q") or "").strip()
        page = max(1, request.args.get("page", 1, type=int) or 1)
        per_page = min(max(10, request.args.get("per_page", 50, type=int) or 50), 200)

        query = ShareSubscription.query.join(Member, ShareSubscription.member_id == Member.id)
        if current_user.role == "agent" and current_user.agent_id:
            query = query.filter(Member.agent_id == current_user.agent_id)
        if current_user.role == "member" and current_user.member_id:
            query = query.filter(ShareSubscription.member_id == current_user.member_id)
        if filter_agent_id and current_user.role != "agent":
            query = query.filter(Member.agent_id == filter_agent_id)
        if member_id:
            query = query.filter(ShareSubscription.member_id == member_id)
        project_id_filter = request.args.get("project_id", type=int)
        if uninvested_only:
            query = query.filter(ShareSubscription.investment_id.is_(None))
        elif investment_id:
            query = query.filter(ShareSubscription.investment_id == investment_id)
        elif project_id_filter:
            inv_ids_for_proj = [
                row[0]
                for row in Investment.query.filter(Investment.project_id == project_id_filter)
                .with_entities(Investment.id)
                .all()
            ]
            if inv_ids_for_proj:
                query = query.filter(ShareSubscription.investment_id.in_(inv_ids_for_proj))
            else:
                query = query.filter(ShareSubscription.id < 0)
        if certificate_pending:
            query = query.filter(ShareSubscription.status == "Fully Paid")
            issued_sub_ids = [
                row[0]
                for row in db.session.query(ShareCertificate.subscription_id)
                .filter(ShareCertificate.status == "Issued")
                .all()
            ]
            if issued_sub_ids:
                query = query.filter(~ShareSubscription.id.in_(issued_sub_ids))
        elif status in {"Pending", "Partially Paid", "Fully Paid", "Cancelled"}:
            query = query.filter(ShareSubscription.status == status)
        if date_from:
            query = query.filter(ShareSubscription.subscription_date >= date_from)
        if date_to:
            query = query.filter(ShareSubscription.subscription_date <= date_to)
        if qtext:
            like = f"%{qtext}%"
            query = query.filter(
                or_(
                    ShareSubscription.subscription_no.ilike(like),
                    Member.member_id.ilike(like),
                    Member.full_name.ilike(like),
                )
            )

        ordered_ids = [
            row[0]
            for row in query.with_entities(ShareSubscription.id)
            .order_by(ShareSubscription.subscription_date.desc(), ShareSubscription.id.desc())
            .all()
        ]

        sub_stats: dict[int, dict] = {}
        for sid in ordered_ids:
            sub_stats[sid] = {"rows": 0, "overdue": 0, "balance": Decimal("0")}
        inst_rows = (
            InstallmentPlan.query.filter(InstallmentPlan.subscription_id.in_(ordered_ids)).all()
            if ordered_ids
            else []
        )
        today = date.today()
        for row in inst_rows:
            st = sub_stats.get(row.subscription_id)
            if st is None:
                continue
            st["rows"] += 1
            if row.status != "Cancelled":
                bal = (row.due_amount or Decimal("0")) - (row.paid_amount or Decimal("0"))
                if bal > 0:
                    st["balance"] += bal
                    is_overdue = row.status == "Overdue" or (
                        row.due_date is not None
                        and row.due_date < today
                        and row.status in {"Pending", "Partially Paid"}
                    )
                    if is_overdue:
                        st["overdue"] += 1

        filtered_ids = list(ordered_ids)
        if overdue_only:
            filtered_ids = [sid for sid in ordered_ids if sub_stats.get(sid, {}).get("overdue", 0) > 0]

        total_count = len(filtered_ids)
        total_pages = max(1, (total_count + per_page - 1) // per_page) if total_count else 1
        page = min(max(1, page), total_pages)

        sum_subscribed = Decimal("0")
        sum_paid = Decimal("0")
        sum_outstanding = Decimal("0")
        cid_map: dict[int, list] = defaultdict(list)
        sub_by_id: dict[int, ShareSubscription] = {}
        if filtered_ids:
            sum_scalar = (
                db.session.query(func.coalesce(func.sum(ShareSubscription.subscribed_amount), 0))
                .filter(ShareSubscription.id.in_(filtered_ids))
                .scalar()
            )
            sum_subscribed = Decimal(str(sum_scalar or 0))
            subs_batch = ShareSubscription.query.filter(ShareSubscription.id.in_(filtered_ids)).all()
            sub_by_id = {s.id: s for s in subs_batch}
            for c in Contribution.query.filter(Contribution.subscription_id.in_(filtered_ids)).all():
                cid_map[c.subscription_id].append(c)

            def _paid_amt(sid: int) -> Decimal:
                return sum((x.amount for x in cid_map.get(sid, [])), Decimal("0"))

            def _out_amt(sid: int) -> Decimal:
                s = sub_by_id[sid]
                paid = _paid_amt(sid)
                bal = (s.subscribed_amount or Decimal("0")) - paid
                return bal if bal > 0 else Decimal("0")

            sum_paid = sum((_paid_amt(sid) for sid in filtered_ids), Decimal("0"))
            sum_outstanding = sum((_out_amt(sid) for sid in filtered_ids), Decimal("0"))

        start = (page - 1) * per_page
        page_ids = filtered_ids[start : start + per_page]
        rows = []
        if page_ids:
            rows = (
                ShareSubscription.query.filter(ShareSubscription.id.in_(page_ids))
                .options(
                    joinedload(ShareSubscription.member).joinedload(Member.agent),
                    joinedload(ShareSubscription.investment),
                    joinedload(ShareSubscription.certificate),
                )
                .order_by(ShareSubscription.subscription_date.desc(), ShareSubscription.id.desc())
                .all()
            )

        subscription_row_metrics: dict[int, dict] = {}
        for s in rows:
            paid = sum((x.amount for x in cid_map.get(s.id, [])), Decimal("0"))
            sub_amt = s.subscribed_amount or Decimal("0")
            out = max(sub_amt - paid, Decimal("0"))
            if sub_amt > 0:
                pct = min((paid / sub_amt) * Decimal("100"), Decimal("100")).quantize(Decimal("0.01"))
            else:
                pct = Decimal("0")
            subscription_row_metrics[s.id] = {"paid": paid, "outstanding": out, "pct": pct}

        def _subs_list_qs(page_num: int) -> str:
            d = request.args.to_dict(flat=True)
            d["page"] = str(page_num)
            return url_for("subscriptions_list") + "?" + urlencode(d)

        qs_no_page = request.args.to_dict(flat=True)
        qs_no_page.pop("page", None)
        contrib_q: dict[str, str] = {}
        if member_id:
            contrib_q["member_id"] = str(member_id)
        if filter_agent_id and current_user.role != "agent":
            contrib_q["agent_id"] = str(filter_agent_id)
        if qtext:
            contrib_q["q"] = qtext
        contributions_filter_url = url_for("contributions_list") + (
            "?" + urlencode(contrib_q) if contrib_q else ""
        )
        next_page_url = _subs_list_qs(page + 1) if page < total_pages else None
        prev_page_url = _subs_list_qs(page - 1) if page > 1 else None
        first_page_url = _subs_list_qs(1) if page > 1 else None
        last_page_url = _subs_list_qs(total_pages) if total_pages > 1 and page < total_pages else None
        row_start = (page - 1) * per_page + 1 if total_count else 0
        row_end = min(page * per_page, total_count) if total_count else 0

        members = members_scope().filter_by(status="Active").order_by(Member.full_name).all()
        agents = []
        if current_user.role != "agent":
            agents = Agent.query.filter_by(status="Active").order_by(Agent.full_name).all()
        investments = Investment.query.order_by(Investment.name).all()
        filter_agent = Agent.query.get(filter_agent_id) if filter_agent_id else None
        filter_project_sub = Project.query.get(project_id_filter) if project_id_filter else None

        settings = get_or_create_settings()

        return render_template(
            "subscriptions/list.html",
            rows=rows,
            member_id=member_id,
            filter_agent_id=filter_agent_id,
            filter_agent=filter_agent,
            investment_id=investment_id,
            uninvested_only=uninvested_only,
            status=status,
            sub_stats=sub_stats,
            overdue_only=overdue_only,
            certificate_pending=certificate_pending,
            date_from=date_from,
            date_to=date_to,
            q=qtext,
            page=page,
            per_page=per_page,
            total_count=total_count,
            total_pages=total_pages,
            next_page_url=next_page_url,
            prev_page_url=prev_page_url,
            first_page_url=first_page_url,
            last_page_url=last_page_url,
            row_start=row_start,
            row_end=row_end,
            sum_subscribed=sum_subscribed,
            sum_paid=sum_paid,
            sum_outstanding=sum_outstanding,
            members=members,
            agents=agents,
            investments=investments,
            settings=settings,
            currency_code=settings.currency_code or "USD",
            contributions_filter_url=contributions_filter_url,
            subscription_row_metrics=subscription_row_metrics,
            filter_project_sub=filter_project_sub,
            project_id_filter=project_id_filter,
        )

    @app.route("/subscriptions/new", methods=["GET", "POST"])
    def subscriptions_new():
        members = members_scope().filter_by(status="Active").order_by(Member.full_name).all()
        investments = Investment.query.order_by(Investment.name).all()
        preselect = request.args.get("member_id", type=int)
        preselect_investment_id = request.args.get("investment_id", type=int)
        settings_sn = get_or_create_settings()

        def _sub_form_ctx():
            mn = min_subscribed_amount()
            mx = max_subscribed_amount()
            sym = settings_sn.currency_symbol or "$"
            return {
                "share_unit_price": SHARE_UNIT_PRICE,
                "share_unit_price_js": float(SHARE_UNIT_PRICE),
                "share_unit_price_str": f"{SHARE_UNIT_PRICE:,.2f}",
                "min_share_units": MIN_SHARE_UNITS,
                "max_share_units": MAX_SHARE_UNITS,
                "min_subscribed_amount": mn,
                "max_subscribed_amount": mx,
                "min_subscribed_amount_str": f"{mn:,.2f}",
                "max_subscribed_amount_str": f"{mx:,.2f}",
                "currency_code": settings_sn.currency_code or "USD",
                "currency_symbol": sym,
            }

        if request.method == "POST":
            member_id = request.form.get("member_id", type=int)
            preselect_investment_id = request.form.get("investment_id", type=int) or preselect_investment_id
            share_units = request.form.get("share_units", type=int)
            payment_plan = request.form.get("payment_plan") or "full"
            eligibility_policy = request.form.get("eligibility_policy") or "paid_proportional"
            subscription_date = _parse_date(request.form.get("subscription_date")) or date.today()
            investment_id = request.form.get("investment_id", type=int) or None
            if not member_id or share_units is None:
                flash("Select member and enter a valid number of shares.", "danger")
                return render_template(
                    "subscriptions/form.html",
                    members=members,
                    investments=investments,
                    preselect=preselect,
                    preselect_investment_id=preselect_investment_id,
                    **_sub_form_ctx(),
                )
            mem = Member.query.get(member_id)
            if not mem or (current_user.role == "agent" and current_user.agent_id and mem.agent_id != current_user.agent_id):
                flash("Invalid member.", "danger")
                return render_template(
                    "subscriptions/form.html",
                    members=members,
                    investments=investments,
                    preselect=preselect,
                    preselect_investment_id=preselect_investment_id,
                    **_sub_form_ctx(),
                )
            if mem.status != "Active":
                flash("Share subscriptions can only be created for Active members.", "danger")
                return render_template(
                    "subscriptions/form.html",
                    members=members,
                    investments=investments,
                    preselect=preselect,
                    preselect_investment_id=preselect_investment_id,
                    **_sub_form_ctx(),
                )
            if investment_id and not Investment.query.get(investment_id):
                flash("Invalid investment selected.", "danger")
                return render_template(
                    "subscriptions/form.html",
                    members=members,
                    investments=investments,
                    preselect=preselect,
                    preselect_investment_id=preselect_investment_id,
                    **_sub_form_ctx(),
                )
            try:
                sub = create_subscription(
                    member_id=member_id,
                    share_units=share_units,
                    payment_plan=payment_plan,
                    eligibility_policy=eligibility_policy,
                    subscription_date=subscription_date,
                    agent_id=mem.agent_id,
                    investment_id=investment_id,
                    commit=False,
                )
            except ValueError as e:
                flash(str(e), "danger")
                return render_template(
                    "subscriptions/form.html",
                    members=members,
                    investments=investments,
                    preselect=preselect,
                    preselect_investment_id=preselect_investment_id,
                    **_sub_form_ctx(),
                )
            log_audit("subscription_created", "ShareSubscription", None, f"subscription_no={sub.subscription_no}")
            db.session.commit()
            try:
                notify_member_new_subscription(sub, mem)
            except Exception:
                pass
            flash(f"Subscription created: {sub.subscription_no}", "success")
            return redirect(url_for("subscriptions_profile", id=sub.id))
        return render_template(
            "subscriptions/form.html",
            members=members,
            investments=investments,
            preselect=preselect,
            preselect_investment_id=preselect_investment_id,
            **_sub_form_ctx(),
        )

    @app.route("/subscriptions/<int:id>")
    def subscriptions_profile(id):
        q = ShareSubscription.query.join(Member, ShareSubscription.member_id == Member.id).options(
            joinedload(ShareSubscription.member).joinedload(Member.agent),
            joinedload(ShareSubscription.investment),
            joinedload(ShareSubscription.certificate),
        )
        if current_user.role == "agent" and current_user.agent_id:
            q = q.filter(Member.agent_id == current_user.agent_id)
        if current_user.role == "member" and current_user.member_id:
            q = q.filter(ShareSubscription.member_id == current_user.member_id)
        sub = q.filter(ShareSubscription.id == id).first_or_404()
        installments = (
            sub.installments.order_by(InstallmentPlan.due_date.asc(), InstallmentPlan.sequence_no.asc()).all()
        )
        contribs = (
            Contribution.query.filter(Contribution.subscription_id == sub.id)
            .options(
                joinedload(Contribution.verified_by),
                joinedload(Contribution.payment_bank_account).joinedload(PaymentBankAccount.bank),
                joinedload(Contribution.payment_mobile_provider),
            )
            .order_by(Contribution.date.desc(), Contribution.id.desc())
            .all()
        )
        installment_rows_total = len(installments)
        installment_overdue_count = 0
        installment_due_balance = Decimal("0")
        today = date.today()
        for row in installments:
            if row.status == "Cancelled":
                continue
            bal = (row.due_amount or Decimal("0")) - (row.paid_amount or Decimal("0"))
            if bal > 0:
                installment_due_balance += bal
                is_overdue = row.status == "Overdue" or (
                    row.due_date is not None
                    and row.due_date < today
                    and row.status in {"Pending", "Partially Paid"}
                )
                if is_overdue:
                    installment_overdue_count += 1
        investments = Investment.query.order_by(Investment.name).all()
        settings = get_or_create_settings()
        paid_total = sub.paid_total()
        outstanding = sub.outstanding_balance()
        completion_pct = sub.completion_percent()
        currency_code = settings.currency_code or "USD"
        return render_template(
            "subscriptions/profile.html",
            sub=sub,
            investments=investments,
            installments=installments,
            contribs=contribs,
            installment_rows_total=installment_rows_total,
            installment_overdue_count=installment_overdue_count,
            installment_due_balance=installment_due_balance,
            settings=settings,
            currency_code=currency_code,
            paid_total=paid_total,
            outstanding=outstanding,
            completion_pct=completion_pct,
            today=today,
        )

    @app.route("/subscriptions/<int:id>/investment", methods=["POST"])
    def subscriptions_set_investment(id):
        if current_user.role == "agent":
            flash("Updating the investment link is not available for your role.", "warning")
            return redirect(url_for("subscriptions_profile", id=id))
        q = ShareSubscription.query.join(Member, ShareSubscription.member_id == Member.id)
        sub = q.filter(ShareSubscription.id == id).first_or_404()
        iid = request.form.get("investment_id", type=int) or None
        if iid and not Investment.query.get(iid):
            flash("Invalid investment.", "danger")
            return redirect(url_for("subscriptions_profile", id=sub.id))
        sub.investment_id = iid
        log_audit(
            "subscription_investment_updated",
            "ShareSubscription",
            sub.id,
            f"investment_id={iid}",
        )
        db.session.commit()
        flash("Investment link updated for this share subscription.", "success")
        return redirect(url_for("subscriptions_profile", id=sub.id))

    @app.route("/subscriptions/<int:id>/cancel", methods=["POST"])
    @role_required("admin", "operator")
    def subscriptions_cancel(id):
        q = ShareSubscription.query.join(Member, ShareSubscription.member_id == Member.id)
        sub = q.filter(ShareSubscription.id == id).first_or_404()
        if sub.status == "Cancelled":
            flash("Subscription is already cancelled.", "info")
            return redirect(url_for("subscriptions_profile", id=sub.id))
        if sub.status == "Fully Paid":
            flash("Cannot cancel a fully paid subscription. Use certificate revocation if needed.", "danger")
            return redirect(url_for("subscriptions_profile", id=sub.id))
        sub.status = "Cancelled"
        sub.confirmed_at = None
        log_audit("subscription_cancelled", "ShareSubscription", sub.id, f"subscription_no={sub.subscription_no}")
        db.session.commit()
        flash("Share subscription cancelled.", "warning")
        return redirect(url_for("subscriptions_profile", id=sub.id))

    @app.route("/subscriptions/<int:id>/installments", methods=["GET", "POST"])
    def subscriptions_installments(id):
        q = ShareSubscription.query.join(Member, ShareSubscription.member_id == Member.id)
        if current_user.role == "agent" and current_user.agent_id:
            q = q.filter(Member.agent_id == current_user.agent_id)
        if current_user.role == "member" and current_user.member_id:
            q = q.filter(ShareSubscription.member_id == current_user.member_id)
        sub = q.filter(ShareSubscription.id == id).first_or_404()

        if request.method == "POST":
            action = request.form.get("action", "").strip()
            if action == "add":
                due_date = _parse_date(request.form.get("due_date"))
                due_amount = _parse_decimal(request.form.get("due_amount"))
                sequence_no = request.form.get("sequence_no", type=int) or 1
                if due_date is None or due_amount is None or due_amount <= 0:
                    flash("Enter valid due date and due amount.", "danger")
                    return redirect(url_for("subscriptions_installments", id=sub.id))
                row = InstallmentPlan(
                    subscription_id=sub.id,
                    due_date=due_date,
                    due_amount=due_amount,
                    paid_amount=Decimal("0"),
                    status="Pending",
                    sequence_no=sequence_no,
                )
                db.session.add(row)
                recompute_installment_statuses(sub.id, commit=False)
                log_audit(
                    "installment_added",
                    "InstallmentPlan",
                    None,
                    f"subscription_id={sub.id} due={due_amount} date={due_date}",
                )
                db.session.commit()
                flash("Installment row added.", "success")
                return redirect(url_for("subscriptions_installments", id=sub.id))

            if action == "update":
                row_id = request.form.get("row_id", type=int)
                row = InstallmentPlan.query.get(row_id) if row_id else None
                if not row or row.subscription_id != sub.id:
                    flash("Invalid installment row.", "danger")
                    return redirect(url_for("subscriptions_installments", id=sub.id))
                due_date = _parse_date(request.form.get("due_date"))
                due_amount = _parse_decimal(request.form.get("due_amount"))
                paid_amount = _parse_decimal(request.form.get("paid_amount"))
                sequence_no = request.form.get("sequence_no", type=int) or row.sequence_no
                if due_date is None or due_amount is None or due_amount <= 0:
                    flash("Enter valid due date and due amount.", "danger")
                    return redirect(url_for("subscriptions_installments", id=sub.id))
                if paid_amount is None or paid_amount < 0:
                    paid_amount = row.paid_amount or Decimal("0")
                row.due_date = due_date
                row.due_amount = due_amount
                row.paid_amount = paid_amount
                row.sequence_no = sequence_no
                recompute_installment_statuses(sub.id, commit=False)
                log_audit(
                    "installment_updated",
                    "InstallmentPlan",
                    row.id,
                    f"subscription_id={sub.id} due={due_amount} paid={paid_amount}",
                )
                db.session.commit()
                flash("Installment row updated.", "success")
                return redirect(url_for("subscriptions_installments", id=sub.id))

            if action == "cancel":
                row_id = request.form.get("row_id", type=int)
                row = InstallmentPlan.query.get(row_id) if row_id else None
                if not row or row.subscription_id != sub.id:
                    flash("Invalid installment row.", "danger")
                    return redirect(url_for("subscriptions_installments", id=sub.id))
                row.status = "Cancelled"
                log_audit(
                    "installment_cancelled",
                    "InstallmentPlan",
                    row.id,
                    f"subscription_id={sub.id}",
                )
                db.session.commit()
                flash("Installment row cancelled.", "warning")
                return redirect(url_for("subscriptions_installments", id=sub.id))

            flash("Unsupported action.", "danger")
            return redirect(url_for("subscriptions_installments", id=sub.id))

        rows = sub.installments.order_by(InstallmentPlan.sequence_no.asc(), InstallmentPlan.due_date.asc()).all()
        return render_template("subscriptions/installments.html", sub=sub, rows=rows)

    # --- Certificates ---
    @app.route("/certificates")
    def certificates_list():
        query = (
            ShareCertificate.query.join(Member, ShareCertificate.member_id == Member.id)
            .options(
                joinedload(ShareCertificate.member),
                joinedload(ShareCertificate.subscription).joinedload(ShareSubscription.investment),
                joinedload(ShareCertificate.agent),
            )
        )
        if current_user.role == "agent" and current_user.agent_id:
            query = query.filter(Member.agent_id == current_user.agent_id)
        if current_user.role == "member" and current_user.member_id:
            query = query.filter(Member.id == current_user.member_id)
        rows = query.order_by(ShareCertificate.issued_date.desc(), ShareCertificate.id.desc()).all()
        settings = get_or_create_settings()
        n_issued = sum(1 for c in rows if c.status == "Issued")
        n_revoked = sum(1 for c in rows if c.status == "Revoked")
        cur_code = settings.currency_code or "USD"
        share_qty_by_cert = {c.id: format_certificate_share_quantity(c.subscription, cur_code) for c in rows}
        return render_template(
            "certificates/list.html",
            rows=rows,
            settings=settings,
            currency_code=cur_code,
            share_qty_by_cert=share_qty_by_cert,
            totals={
                "total": len(rows),
                "issued": n_issued,
                "revoked": n_revoked,
            },
        )

    @app.route("/certificates/issue/<int:subscription_id>", methods=["POST"])
    def certificates_issue(subscription_id):
        if current_user.role == "agent":
            flash("Certificate issuance is not available for your role.", "warning")
            return redirect(url_for("subscriptions_profile", id=subscription_id))
        try:
            cert = issue_certificate(
                subscription_id,
                issued_by_user_id=current_user.id if current_user.is_authenticated else None,
                notes=request.form.get("notes", ""),
                commit=False,
            )
            log_audit(
                "certificate_issued",
                "ShareCertificate",
                None,
                f"certificate_no={cert.certificate_no} subscription_id={subscription_id}",
            )
            db.session.commit()
            try:
                cm = Member.query.get(cert.member_id)
                if cm:
                    notify_member_certificate_issued(cert, cm)
            except Exception:
                pass
            flash(f"Certificate issued: {cert.certificate_no}", "success")
            return redirect(url_for("certificates_print", id=cert.id))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
            return redirect(url_for("subscriptions_profile", id=subscription_id))

    @app.route("/certificates/<int:id>/print")
    def certificates_print(id):
        query = ShareCertificate.query.join(Member, ShareCertificate.member_id == Member.id).options(
            joinedload(ShareCertificate.member),
            joinedload(ShareCertificate.agent),
            joinedload(ShareCertificate.subscription)
            .joinedload(ShareSubscription.investment)
            .joinedload(Investment.project),
            joinedload(ShareCertificate.issued_by),
        )
        if current_user.role == "agent" and current_user.agent_id:
            query = query.filter(Member.agent_id == current_user.agent_id)
        if current_user.role == "member" and current_user.member_id:
            query = query.filter(Member.id == current_user.member_id)
        cert = query.filter(ShareCertificate.id == id).first_or_404()
        s = get_or_create_settings()
        ex = s.get_extra()
        sub = cert.subscription
        cur = s.currency_code or "USD"
        company_nm = (ex.get("company_name") or "Estithmar Investment Management").strip()
        cert_share_qty = format_certificate_share_quantity(sub, cur)
        cert_stock_of = certificate_stock_of_name(sub, company_nm)
        sym = s.currency_symbol or "$"
        cert_share_detail = certificate_share_position_detail(sub, sym, cur)
        idate = cert.issued_date or date.today()
        dnum = idate.day
        if 11 <= (dnum % 100) <= 13:
            day_ord = f"{dnum}th"
        else:
            day_ord = f"{dnum}{ {1: 'st', 2: 'nd', 3: 'rd'}.get(dnum % 10, 'th') }"
        cert_issued_by_label = None
        if cert.issued_by:
            cert_issued_by_label = (cert.issued_by.full_name or cert.issued_by.username or "").strip() or None
        return render_template(
            "certificates/print.html",
            cert=cert,
            settings=s,
            extra=ex,
            currency_code=cur,
            cert_share_qty=cert_share_qty,
            cert_stock_of=cert_stock_of,
            cert_share_detail=cert_share_detail,
            cert_issued_day=idate.day,
            cert_issued_day_ordinal=day_ord,
            cert_issued_month=idate.strftime("%B"),
            cert_issued_year=idate.year,
            company_display_name=company_nm,
            cert_issued_by_label=cert_issued_by_label,
        )

    @app.route("/certificates/<int:id>/pdf")
    def certificates_pdf(id):
        """Generate certificate as PDF download (same visibility rules as print view)."""
        query = (
            ShareCertificate.query.join(Member, ShareCertificate.member_id == Member.id)
            .options(
                joinedload(ShareCertificate.member),
                joinedload(ShareCertificate.agent),
                joinedload(ShareCertificate.subscription)
                .joinedload(ShareSubscription.investment)
                .joinedload(Investment.project),
                joinedload(ShareCertificate.issued_by),
            )
        )
        if current_user.role == "agent" and current_user.agent_id:
            query = query.filter(Member.agent_id == current_user.agent_id)
        if current_user.role == "member" and current_user.member_id:
            query = query.filter(Member.id == current_user.member_id)
        cert = query.filter(ShareCertificate.id == id).first_or_404()
        if cert.status != "Issued":
            flash("PDF is only available for issued certificates.", "warning")
            return redirect(url_for("certificates_print", id=id))
        s = get_or_create_settings()
        pdf_extra = dict(s.get_extra())
        pdf_extra["currency_code"] = s.currency_code or "USD"
        pdf_extra["currency_symbol"] = s.currency_symbol or "$"
        pdf_bytes = build_share_certificate_pdf(cert, extra=pdf_extra)
        safe_name = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in (cert.certificate_no or str(id)))
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="certificate-{safe_name}.pdf"'},
        )

    @app.route("/certificates/<int:id>/revoke", methods=["POST"])
    @role_required("admin", "operator")
    def certificates_revoke(id):
        cert = ShareCertificate.query.get_or_404(id)
        reason = (request.form.get("reason", "") or "").strip()[:500]
        if cert.status == "Revoked":
            flash("Certificate is already revoked.", "info")
            return redirect(url_for("certificates_list"))
        if cert.status != "Issued":
            flash("Only issued certificates can be revoked.", "warning")
            return redirect(url_for("certificates_list"))
        cert.status = "Revoked"
        note = f"Revoked by user_id={current_user.id if current_user.is_authenticated else 'unknown'}"
        if reason:
            note = f"{note}; reason={reason}"
        log_audit("certificate_revoked", "ShareCertificate", cert.id, note)
        db.session.commit()
        flash("Certificate revoked.", "warning")
        return redirect(url_for("certificates_list"))

    @app.route("/certificates/<int:id>/reinstate", methods=["POST"])
    @role_required("admin", "operator")
    def certificates_reinstate(id):
        cert = ShareCertificate.query.get_or_404(id)
        if cert.status != "Revoked":
            flash("Only revoked certificates can be reinstated to Issued.", "info")
            return redirect(url_for("certificates_list"))
        cert.status = "Issued"
        note = f"Reinstated to Issued by user_id={current_user.id if current_user.is_authenticated else 'unknown'}"
        log_audit("certificate_reinstated", "ShareCertificate", cert.id, note)
        db.session.commit()
        flash("Certificate reinstated — status is Issued again.", "success")
        return redirect(url_for("certificates_list"))

    # --- Contributions ---
    @app.route("/contributions")
    def contributions_list():
        member_id = request.args.get("member_id", type=int)
        agent_id = request.args.get("agent_id", type=int)
        subscription_id = request.args.get("subscription_id", type=int)
        date_from = _parse_date(request.args.get("date_from"))
        date_to = _parse_date(request.args.get("date_to"))
        verified = (request.args.get("verified") or "").strip().lower()
        qtext = (request.args.get("q") or "").strip()
        payment_type_filter = (request.args.get("payment_type") or "").strip()
        page = max(1, request.args.get("page", 1, type=int) or 1)
        per_page = min(max(10, request.args.get("per_page", 50, type=int) or 50), 200)

        total = _contributions_filtered_query().count()
        sum_scalar = (
            _contributions_filtered_query()
            .with_entities(func.coalesce(func.sum(Contribution.amount), 0))
            .scalar()
        )
        tot_dec = Decimal(str(sum_scalar or 0))
        rows_q = _contributions_filtered_query()
        rows = (
            rows_q.options(
                joinedload(Contribution.member).joinedload(Member.agent),
                joinedload(Contribution.subscription),
                joinedload(Contribution.payment_bank_account).joinedload(PaymentBankAccount.bank),
                joinedload(Contribution.payment_mobile_provider),
            )
            .order_by(Contribution.date.desc(), Contribution.id.desc())
            .offset((page - 1) * per_page)
            .limit(per_page)
            .all()
        )
        total_pages = max(1, (total + per_page - 1) // per_page) if total else 1

        members = members_scope().filter_by(status="Active").order_by(Member.full_name).all()
        agents = []
        if current_user.role != "agent":
            agents = Agent.query.filter_by(status="Active").order_by(Agent.full_name).all()
        sub_filter_q = ShareSubscription.query.join(Member, ShareSubscription.member_id == Member.id)
        if current_user.role == "agent" and current_user.agent_id:
            sub_filter_q = sub_filter_q.filter(Member.agent_id == current_user.agent_id)
        if member_id:
            sub_filter_q = sub_filter_q.filter(ShareSubscription.member_id == member_id)
        filter_subscriptions = sub_filter_q.order_by(ShareSubscription.subscription_date.desc()).limit(400).all()

        def _contrib_list_qs(page_num: int) -> str:
            d = request.args.to_dict(flat=True)
            d["page"] = str(page_num)
            return url_for("contributions_list") + "?" + urlencode(d)

        qs_no_page = request.args.to_dict(flat=True)
        qs_no_page.pop("page", None)
        export_url = url_for("export_contributions_xlsx") + "?" + urlencode(qs_no_page)
        next_page_url = _contrib_list_qs(page + 1) if page < total_pages else None
        prev_page_url = _contrib_list_qs(page - 1) if page > 1 else None

        return render_template(
            "contributions/list.html",
            rows=rows,
            total_collected=tot_dec,
            total_count=total,
            members=members,
            member_id=member_id,
            agent_id=agent_id,
            subscription_id=subscription_id,
            filter_subscriptions=filter_subscriptions,
            agents=agents,
            date_from=date_from,
            date_to=date_to,
            verified_filter=verified,
            payment_type_filter=payment_type_filter,
            q=qtext,
            page=page,
            per_page=per_page,
            total_pages=total_pages,
            export_url=export_url,
            next_page_url=next_page_url,
            prev_page_url=prev_page_url,
        )

    @app.route("/contributions/new", methods=["GET", "POST"])
    def contributions_new():
        app_settings = get_or_create_settings()
        cur_lbl = app_settings.currency_code or "USD"

        def _contrib_form_ctx(members, preselect, preselect_subscription, subscriptions):
            ba, mp = _contribution_payment_form_choices()
            return {
                "members": members,
                "preselect": preselect,
                "preselect_subscription": preselect_subscription,
                "subscriptions": subscriptions,
                "settings": app_settings,
                "today": date.today(),
                "bank_accounts": ba,
                "mobile_providers": mp,
                "has_bank_accounts": len(ba) > 0,
                "has_mobile_providers": len(mp) > 0,
            }

        members = members_scope().filter_by(status="Active").order_by(Member.full_name).all()
        preselect = request.args.get("preselect", type=int)
        preselect_subscription = request.args.get("subscription_id", type=int)
        if request.method == "POST":
            mid = request.form.get("member_id", type=int)
            subscription_id = request.form.get("subscription_id", type=int) or None
            amount = _parse_decimal(request.form.get("amount"))
            d = _parse_date(request.form.get("date")) or date.today()
            ptype = request.form.get("payment_type") or "Cash"
            if ptype not in ("Cash", "Mobile", "Bank", "Other"):
                ptype = "Cash"
            bank_acc_id = request.form.get("payment_bank_account_id", type=int) or None
            mobile_prov_id = request.form.get("payment_mobile_provider_id", type=int) or None
            ba_choices, mp_choices = _contribution_payment_form_choices()
            if ptype == "Bank" and ba_choices and not bank_acc_id:
                flash("Select which company bank account received this transfer.", "danger")
                subs = (
                    ShareSubscription.query.filter_by(member_id=mid).order_by(ShareSubscription.subscription_date.desc()).all()
                    if mid
                    else []
                )
                return render_template(
                    "contributions/form.html",
                    **_contrib_form_ctx(members, mid or preselect, subscription_id or preselect_subscription, subs),
                )
            if ptype == "Mobile" and mp_choices and not mobile_prov_id:
                flash("Select the mobile payment service used (e.g. EVC, eDahab).", "danger")
                subs = (
                    ShareSubscription.query.filter_by(member_id=mid).order_by(ShareSubscription.subscription_date.desc()).all()
                    if mid
                    else []
                )
                return render_template(
                    "contributions/form.html",
                    **_contrib_form_ctx(members, mid or preselect, subscription_id or preselect_subscription, subs),
                )
            if ptype != "Bank":
                bank_acc_id = None
            if ptype != "Mobile":
                mobile_prov_id = None
            if bank_acc_id:
                acc = PaymentBankAccount.query.filter_by(id=bank_acc_id, is_active=True).first()
                if not acc or not acc.bank or not acc.bank.is_active:
                    flash("Invalid or inactive bank account.", "danger")
                    subs = (
                        ShareSubscription.query.filter_by(member_id=mid).order_by(ShareSubscription.subscription_date.desc()).all()
                        if mid
                        else []
                    )
                    return render_template(
                        "contributions/form.html",
                        **_contrib_form_ctx(members, mid or preselect, subscription_id or preselect_subscription, subs),
                    )
            if mobile_prov_id:
                mp = PaymentMobileProvider.query.filter_by(id=mobile_prov_id, is_active=True).first()
                if not mp:
                    flash("Invalid mobile payment option.", "danger")
                    subs = (
                        ShareSubscription.query.filter_by(member_id=mid).order_by(ShareSubscription.subscription_date.desc()).all()
                        if mid
                        else []
                    )
                    return render_template(
                        "contributions/form.html",
                        **_contrib_form_ctx(members, mid or preselect, subscription_id or preselect_subscription, subs),
                    )
            method_ref = request.form.get("method_ref", "").strip() or None
            notes = request.form.get("notes", "").strip()
            if not mid or amount is None or amount <= 0:
                flash("Select a member and enter a valid amount.", "danger")
                return render_template(
                    "contributions/form.html",
                    **_contrib_form_ctx(
                        members,
                        mid or preselect,
                        subscription_id or preselect_subscription,
                        [],
                    ),
                )
            mem = Member.query.get(mid)
            if not mem or (current_user.role == "agent" and current_user.agent_id and mem.agent_id != current_user.agent_id):
                flash("Invalid member.", "danger")
                return render_template(
                    "contributions/form.html",
                    **_contrib_form_ctx(members, preselect, preselect_subscription, []),
                )
            if mem.status != "Active":
                flash("Contributions can only be recorded for Active members.", "danger")
                subs = (
                    ShareSubscription.query.filter_by(member_id=mem.id)
                    .order_by(ShareSubscription.subscription_date.desc())
                    .all()
                )
                return render_template(
                    "contributions/form.html",
                    **_contrib_form_ctx(members, mem.id, subscription_id or preselect_subscription, subs),
                )
            sub = None
            if subscription_id:
                sub = ShareSubscription.query.get(subscription_id)
                if not sub or sub.member_id != mem.id:
                    flash("Selected subscription does not belong to this member.", "danger")
                    subs = (
                        ShareSubscription.query.filter_by(member_id=mem.id)
                        .order_by(ShareSubscription.subscription_date.desc())
                        .all()
                    )
                    return render_template(
                        "contributions/form.html",
                        **_contrib_form_ctx(members, mem.id, preselect_subscription, subs),
                    )
                if current_user.role == "agent" and current_user.agent_id and sub.agent_id != current_user.agent_id:
                    flash("Access denied for this subscription.", "danger")
                    return redirect(url_for("contributions_list"))
                if sub.status == "Cancelled":
                    flash("Cannot record payments against a cancelled subscription.", "danger")
                    subs = (
                        ShareSubscription.query.filter_by(member_id=mem.id)
                        .order_by(ShareSubscription.subscription_date.desc())
                        .all()
                    )
                    return render_template(
                        "contributions/form.html",
                        **_contrib_form_ctx(members, mem.id, subscription_id, subs),
                    )
                max_pay = max_payment_for_subscription(sub)
                if amount > max_pay:
                    flash(
                        f"Payment exceeds remaining balance ({max_pay} {cur_lbl}). "
                        "No overpayment beyond the subscribed amount is allowed.",
                        "danger",
                    )
                    subs = (
                        ShareSubscription.query.filter_by(member_id=mem.id)
                        .order_by(ShareSubscription.subscription_date.desc())
                        .all()
                    )
                    return render_template(
                        "contributions/form.html",
                        **_contrib_form_ctx(members, mem.id, subscription_id, subs),
                    )
            c = Contribution(
                member_id=mid,
                agent_id=mem.agent_id,
                subscription_id=subscription_id,
                amount=amount,
                date=d,
                payment_type=ptype,
                payment_bank_account_id=bank_acc_id,
                payment_mobile_provider_id=mobile_prov_id,
                method_ref=method_ref,
                receipt_no=next_receipt_no(),
                notes=notes or None,
            )
            db.session.add(c)
            db.session.flush()
            issued_auto = False
            if sub is not None:
                auto_allocate_payment_to_installments(sub.id, amount, payment_date=d, commit=False)
                recompute_subscription_status(sub.id, commit=False)
                issued_auto = maybe_auto_issue_certificate(
                    sub.id, user_id=current_user.id if current_user.is_authenticated else None
                )
            log_audit("contribution_recorded", "Contribution", None, f"member_id={mid} amount={amount}")
            db.session.commit()
            try:
                notify_member_payment(c, mem)
            except Exception:
                pass
            if issued_auto:
                flash("Share certificate issued automatically.", "success")
                sub_reload = ShareSubscription.query.get(sub.id) if sub else None
                if sub_reload and sub_reload.certificate:
                    try:
                        notify_member_certificate_issued(sub_reload.certificate, mem)
                    except Exception:
                        pass
            elif sub is not None:
                sub_chk = ShareSubscription.query.get(sub.id)
                if sub_chk and sub_chk.status == "Fully Paid":
                    if sub_chk.confirmed_at:
                        flash(
                            "Share subscription confirmed: paid equals subscribed amount. Certificate eligibility applies (issue manually or via auto-issue in Settings).",
                            "success",
                        )
            flash("Contribution recorded.", "success")
            return redirect(url_for("contributions_receipt", id=c.id))
        subscriptions = []
        if preselect:
            subscriptions = (
                ShareSubscription.query.filter_by(member_id=preselect)
                .order_by(ShareSubscription.subscription_date.desc())
                .all()
            )
        return render_template(
            "contributions/form.html",
            **_contrib_form_ctx(members, preselect, preselect_subscription, subscriptions),
        )

    @app.route("/contributions/<int:id>/receipt")
    def contributions_receipt(id):
        c = (
            Contribution.query.options(
                joinedload(Contribution.member).joinedload(Member.agent),
                joinedload(Contribution.subscription),
                joinedload(Contribution.verified_by),
                joinedload(Contribution.payment_bank_account).joinedload(PaymentBankAccount.bank),
                joinedload(Contribution.payment_mobile_provider),
            )
            .get_or_404(id)
        )
        if current_user.role == "agent" and current_user.agent_id and c.member.agent_id != current_user.agent_id:
            flash("Access denied.", "danger")
            return redirect(url_for("contributions_list"))
        if current_user.role == "member" and current_user.member_id and c.member_id != current_user.member_id:
            flash("Access denied.", "danger")
            return redirect(url_for("contributions_list"))
        receipt_schedule = None
        sub = c.subscription if c.subscription_id else None
        if sub:
            receipt_schedule = subscription_payment_running_rows(sub)
        settings = get_or_create_settings()
        extra = settings.get_extra()
        member_sub_outstanding = Decimal("0")
        if sub is None and c.member:
            for srow in c.member.subscriptions.filter(ShareSubscription.status != "Cancelled").all():
                member_sub_outstanding += srow.outstanding_balance()
        return render_template(
            "contributions/receipt.html",
            c=c,
            receipt_schedule=receipt_schedule,
            subscription=sub,
            settings=settings,
            extra=extra,
            receipt_url=request.url,
            member_sub_outstanding=member_sub_outstanding,
        )

    @app.route("/contributions/<int:id>/verify", methods=["POST"])
    @role_required("admin", "operator")
    def contributions_verify(id):
        c = Contribution.query.get_or_404(id)
        c.verified = True
        c.verified_at = datetime.utcnow()
        c.verified_by_user_id = current_user.id
        db.session.commit()
        try:
            post_contribution_verified(id, user_id=current_user.id)
            db.session.commit()
        except Exception:
            db.session.rollback()
        flash("Payment marked as verified.", "success")
        nxt = (request.form.get("next") or "").strip()
        if nxt.startswith("/") and not nxt.startswith("//"):
            return redirect(nxt)
        return redirect(request.referrer or url_for("contributions_list"))

    @app.route("/contributions/<int:id>/unverify", methods=["POST"])
    @role_required("admin", "operator")
    def contributions_unverify(id):
        c = Contribution.query.get_or_404(id)
        c.verified = False
        c.verified_at = None
        c.verified_by_user_id = None
        db.session.commit()
        try:
            post_contribution_unverified(id)
            db.session.commit()
        except Exception:
            db.session.rollback()
        flash("Verification cleared.", "info")
        nxt = (request.form.get("next") or "").strip()
        if nxt.startswith("/") and not nxt.startswith("//"):
            return redirect(nxt)
        return redirect(request.referrer or url_for("contributions_list"))

    # --- Projects ---
    @app.route("/projects")
    def projects_list():
        page = max(1, request.args.get("page", 1, type=int) or 1)
        per_page = min(max(10, request.args.get("per_page", 50, type=int) or 50), 200)
        total_count = _projects_filtered_query().count()
        sum_budget_scalar = _projects_filtered_query().with_entities(
            func.coalesce(func.sum(Project.total_budget), 0)
        ).scalar()
        sum_budget = Decimal(str(sum_budget_scalar or 0))
        pids = [row[0] for row in _projects_filtered_query().with_entities(Project.id).all()]
        if pids:
            inv_scalar = (
                db.session.query(func.coalesce(func.sum(Investment.total_amount_invested), 0))
                .filter(Investment.project_id.in_(pids))
                .scalar()
            )
            sum_invested = Decimal(str(inv_scalar or 0))
        else:
            sum_invested = Decimal("0")
        total_pages = max(1, (total_count + per_page - 1) // per_page) if total_count else 1
        page = min(max(1, page), total_pages)
        projects = (
            _projects_filtered_query()
            .order_by(Project.id.desc())
            .offset((page - 1) * per_page)
            .limit(per_page)
            .all()
        )

        def _projects_list_qs(page_num: int) -> str:
            d = request.args.to_dict(flat=True)
            d["page"] = str(page_num)
            return url_for("projects_list") + "?" + urlencode(d)

        qs_no_page = request.args.to_dict(flat=True)
        qs_no_page.pop("page", None)
        export_url = url_for("export_projects_xlsx") + "?" + urlencode(qs_no_page)
        next_page_url = _projects_list_qs(page + 1) if page < total_pages else None
        prev_page_url = _projects_list_qs(page - 1) if page > 1 else None
        first_page_url = _projects_list_qs(1) if page > 1 else None
        last_page_url = _projects_list_qs(total_pages) if total_pages > 1 and page < total_pages else None
        row_start = (page - 1) * per_page + 1 if total_count else 0
        row_end = min(page * per_page, total_count) if total_count else 0

        settings = get_or_create_settings()
        category = request.args.get("category", "").strip()
        status = request.args.get("status", "").strip()
        date_from = _parse_date(request.args.get("date_from"))
        date_to = _parse_date(request.args.get("date_to"))
        qtext = (request.args.get("q") or "").strip()
        cat_map = dict(PROJECT_CATEGORIES)
        return render_template(
            "projects/list.html",
            projects=projects,
            cat_map=cat_map,
            project_statuses=PROJECT_STATUSES,
            project_categories=PROJECT_CATEGORIES,
            total_count=total_count,
            sum_budget=sum_budget,
            sum_invested=sum_invested,
            page=page,
            per_page=per_page,
            total_pages=total_pages,
            next_page_url=next_page_url,
            prev_page_url=prev_page_url,
            first_page_url=first_page_url,
            last_page_url=last_page_url,
            row_start=row_start,
            row_end=row_end,
            export_url=export_url,
            settings=settings,
            currency_code=settings.currency_code or "USD",
            category=category,
            status=status,
            date_from=date_from,
            date_to=date_to,
            q=qtext,
        )

    @app.route("/projects/new", methods=["GET", "POST"])
    def projects_new():
        settings = get_or_create_settings()
        form_ctx = {
            "project": None,
            "categories": PROJECT_CATEGORIES,
            "project_statuses": PROJECT_STATUSES,
            "settings": settings,
            "currency_code": settings.currency_code or "USD",
        }
        if request.method == "POST":
            p = Project(
                project_code=next_project_code(),
                name=request.form.get("name", "").strip(),
                category=request.form.get("category") or "real_estate",
                description=request.form.get("description", "").strip(),
                start_date=_parse_date(request.form.get("start_date")),
                end_date=_parse_date(request.form.get("end_date")),
                project_manager=request.form.get("project_manager", "").strip(),
                status=_valid_project_status(request.form.get("status")),
                total_budget=_parse_decimal(request.form.get("total_budget")),
            )
            if not p.name:
                flash("Project name is required.", "danger")
                return render_template("projects/form.html", **form_ctx)
            db.session.add(p)
            log_audit("project_created", "Project", None, p.name)
            db.session.commit()
            flash("Project created.", "success")
            return redirect(url_for("projects_profile", id=p.id))
        return render_template("projects/form.html", **form_ctx)

    @app.route("/projects/<int:id>")
    def projects_profile(id):
        p = Project.query.get_or_404(id)
        invs = p.investments.order_by(Investment.start_date.desc(), Investment.id.desc()).all()
        inv_ids = [inv.id for inv in invs]
        dist_totals: dict[int, Decimal] = {}
        if inv_ids:
            sums = (
                db.session.query(
                    ProfitDistribution.investment_id,
                    func.coalesce(func.sum(ProfitDistribution.amount), 0),
                )
                .filter(ProfitDistribution.investment_id.in_(inv_ids))
                .group_by(ProfitDistribution.investment_id)
                .all()
            )
            dist_totals = {iid: Decimal(str(s)) for iid, s in sums}
        profit_distributed_total = sum(dist_totals.values(), Decimal("0")) if dist_totals else Decimal("0")
        subs_count = (
            ShareSubscription.query.filter(ShareSubscription.investment_id.in_(inv_ids)).count()
            if inv_ids
            else 0
        )
        budget = p.total_budget
        received = p.total_investment_received()
        headroom = project_budget_headroom(p)
        settings = get_or_create_settings()
        pool = _funds_pool_summary()
        inv_profit_sum = sum((inv.profit_generated or Decimal("0") for inv in invs), Decimal("0"))
        inv_cap_ret_sum = sum((inv.capital_returned or Decimal("0") for inv in invs), Decimal("0"))
        budget_util_pct = None
        if budget is not None and (budget or Decimal("0")) > 0:
            budget_util_pct = min(
                (received / (budget or Decimal("1"))) * Decimal("100"),
                Decimal("100"),
            ).quantize(Decimal("0.01"))
        profit_undistributed_sum = sum(
            (
                max(
                    (inv.profit_generated or Decimal("0")) - dist_totals.get(inv.id, Decimal("0")),
                    Decimal("0"),
                )
                for inv in invs
            ),
            Decimal("0"),
        )
        return render_template(
            "projects/profile.html",
            project=p,
            investments=invs,
            cat_map=dict(PROJECT_CATEGORIES),
            budget=budget,
            invested_total=received,
            budget_headroom=headroom,
            pool=pool,
            settings=settings,
            currency_code=settings.currency_code or "USD",
            inv_profit_sum=inv_profit_sum,
            inv_cap_ret_sum=inv_cap_ret_sum,
            dist_totals=dist_totals,
            profit_distributed_total=profit_distributed_total,
            profit_undistributed_sum=profit_undistributed_sum,
            subs_count=subs_count,
            budget_util_pct=budget_util_pct,
        )

    @app.route("/projects/<int:id>/edit", methods=["GET", "POST"])
    def projects_edit(id):
        p = Project.query.get_or_404(id)
        settings = get_or_create_settings()
        form_ctx = {
            "project": p,
            "categories": PROJECT_CATEGORIES,
            "project_statuses": PROJECT_STATUSES,
            "settings": settings,
            "currency_code": settings.currency_code or "USD",
            "invested_total": p.total_investment_received(),
            "budget_headroom": project_budget_headroom(p),
        }
        if request.method == "POST":
            p.name = request.form.get("name", "").strip()
            p.category = request.form.get("category") or p.category
            p.description = request.form.get("description", "").strip()
            p.start_date = _parse_date(request.form.get("start_date"))
            p.end_date = _parse_date(request.form.get("end_date"))
            p.project_manager = request.form.get("project_manager", "").strip()
            p.status = _valid_project_status(request.form.get("status"))
            p.total_budget = _parse_decimal(request.form.get("total_budget"))
            if not p.name:
                flash("Project name is required.", "danger")
                return render_template("projects/form.html", **form_ctx)
            log_audit("project_updated", "Project", p.id)
            db.session.commit()
            flash("Project updated.", "success")
            return redirect(url_for("projects_profile", id=p.id))
        return render_template("projects/form.html", **form_ctx)

    # --- Investments ---
    @app.route("/investments")
    def investments_list():
        page = max(1, request.args.get("page", 1, type=int) or 1)
        per_page = min(max(10, request.args.get("per_page", 50, type=int) or 50), 200)
        project_filter = request.args.get("project_id", type=int)
        status = request.args.get("status", "").strip()
        date_from = _parse_date(request.args.get("date_from"))
        date_to = _parse_date(request.args.get("date_to"))
        qtext = (request.args.get("q") or "").strip()

        total_count = _investments_filtered_query().count()
        sum_inv_scalar = _investments_filtered_query().with_entities(
            func.coalesce(func.sum(Investment.total_amount_invested), 0)
        ).scalar()
        sum_profit_scalar = _investments_filtered_query().with_entities(
            func.coalesce(func.sum(Investment.profit_generated), 0)
        ).scalar()
        sum_cap_scalar = _investments_filtered_query().with_entities(
            func.coalesce(func.sum(Investment.capital_returned), 0)
        ).scalar()
        sum_invested = Decimal(str(sum_inv_scalar or 0))
        sum_profit = Decimal(str(sum_profit_scalar or 0))
        sum_capital = Decimal(str(sum_cap_scalar or 0))
        all_ids = [row[0] for row in _investments_filtered_query().with_entities(Investment.id).all()]
        if all_ids:
            dist_sum_scalar = (
                db.session.query(func.coalesce(func.sum(ProfitDistribution.amount), 0))
                .filter(ProfitDistribution.investment_id.in_(all_ids))
                .scalar()
            )
            sum_distributed = Decimal(str(dist_sum_scalar or 0))
        else:
            sum_distributed = Decimal("0")

        total_pages = max(1, (total_count + per_page - 1) // per_page) if total_count else 1
        page = min(max(1, page), total_pages)
        rows = (
            _investments_filtered_query()
            .options(joinedload(Investment.project))
            .order_by(Investment.id.desc())
            .offset((page - 1) * per_page)
            .limit(per_page)
            .all()
        )
        page_ids = [r.id for r in rows]
        dist_totals = {}
        if page_ids:
            sums = (
                db.session.query(
                    ProfitDistribution.investment_id,
                    func.coalesce(func.sum(ProfitDistribution.amount), 0),
                )
                .filter(ProfitDistribution.investment_id.in_(page_ids))
                .group_by(ProfitDistribution.investment_id)
                .all()
            )
            dist_totals = {iid: Decimal(str(s)) for iid, s in sums}

        def _inv_list_qs(page_num: int) -> str:
            d = request.args.to_dict(flat=True)
            d["page"] = str(page_num)
            return url_for("investments_list") + "?" + urlencode(d)

        qs_no_page = request.args.to_dict(flat=True)
        qs_no_page.pop("page", None)
        export_url = url_for("export_investments_xlsx") + "?" + urlencode(qs_no_page)
        next_page_url = _inv_list_qs(page + 1) if page < total_pages else None
        prev_page_url = _inv_list_qs(page - 1) if page > 1 else None
        first_page_url = _inv_list_qs(1) if page > 1 else None
        last_page_url = _inv_list_qs(total_pages) if total_pages > 1 and page < total_pages else None
        row_start = (page - 1) * per_page + 1 if total_count else 0
        row_end = min(page * per_page, total_count) if total_count else 0

        filter_project = Project.query.get(project_filter) if project_filter else None
        projects = Project.query.order_by(Project.name).all()
        settings = get_or_create_settings()
        pool = _funds_pool_summary()
        sum_undistributed = max(sum_profit - sum_distributed, Decimal("0"))

        return render_template(
            "investments/list.html",
            rows=rows,
            dist_totals=dist_totals,
            filter_project_id=project_filter,
            filter_project=filter_project,
            projects=projects,
            investment_statuses=INVESTMENT_STATUSES,
            total_count=total_count,
            sum_invested=sum_invested,
            sum_profit=sum_profit,
            sum_capital=sum_capital,
            sum_distributed=sum_distributed,
            sum_undistributed=sum_undistributed,
            page=page,
            per_page=per_page,
            total_pages=total_pages,
            next_page_url=next_page_url,
            prev_page_url=prev_page_url,
            first_page_url=first_page_url,
            last_page_url=last_page_url,
            row_start=row_start,
            row_end=row_end,
            export_url=export_url,
            settings=settings,
            currency_code=settings.currency_code or "USD",
            pool=pool,
            status=status,
            date_from=date_from,
            date_to=date_to,
            q=qtext,
        )

    @app.route("/investments/new", methods=["GET", "POST"])
    def investments_new():
        projects = Project.query.order_by(Project.name).all()
        preselect_project_id = request.args.get("project_id", type=int)
        if request.method == "POST":
            preselect_project_id = request.form.get("project_id", type=int) or preselect_project_id
        settings = get_or_create_settings()
        banner_project = Project.query.get(preselect_project_id) if preselect_project_id else None
        ctx = {
            "investment": None,
            "projects": projects,
            "pool": _funds_pool_summary(),
            "investment_statuses": INVESTMENT_STATUSES,
            "categories": PROJECT_CATEGORIES,
            "profit_distribution_frequencies": PROFIT_DISTRIBUTION_FREQUENCIES,
            "locked_completed": False,
            "is_admin_view": current_user.role == "admin",
            "preselect_project_id": preselect_project_id,
            "settings": settings,
            "currency_code": settings.currency_code or "USD",
            "banner_project": banner_project,
        }
        if request.method == "POST":
            pid = request.form.get("project_id", type=int) or None
            amt = _parse_decimal(request.form.get("total_amount_invested"))
            prof = _parse_decimal(request.form.get("profit_generated")) or Decimal("0")
            cap_ret = _parse_decimal(request.form.get("capital_returned")) or Decimal("0")
            proj = None
            if pid:
                proj = Project.query.get(pid)
                if not proj:
                    flash("Invalid project.", "danger")
                    return render_template("investments/form.html", **ctx)
                if proj.status == "Closed":
                    flash("Cannot link an investment to a closed project.", "danger")
                    return render_template("investments/form.html", **ctx)
            amt = amt or Decimal("0")
            if amt <= 0:
                flash("Enter a positive total amount invested.", "danger")
                return render_template("investments/form.html", **ctx)
            avail = available_pool_for_investment(
                exclude_investment_id=None, verified_only=_pool_verified_only()
            )
            if amt > avail:
                flash("Investment exceeds available funds. Please reduce amount.", "danger")
                return render_template("investments/form.html", **ctx)
            if proj is not None:
                hr = project_budget_headroom(proj)
                if hr is not None and amt > hr:
                    flash(
                        f"Amount exceeds this project's remaining budget ({hr} USD under the project budget cap).",
                        "danger",
                    )
                    return render_template("investments/form.html", **ctx)
            new_status = _valid_investment_status(request.form.get("status"))
            if prof < 0 and new_status not in ("Planned", "Active", "Suspended"):
                flash("Negative profit is only allowed while the investment is Planned, Active, or Suspended.", "danger")
                return render_template("investments/form.html", **ctx)
            freq = (request.form.get("profit_distribution_frequency") or "").strip()
            valid_freq = {k for k, _ in PROFIT_DISTRIBUTION_FREQUENCIES}
            inv = Investment(
                investment_code=next_investment_code(),
                name=request.form.get("name", "").strip(),
                investment_type=request.form.get("investment_type", "").strip(),
                total_amount_invested=amt,
                capital_returned=cap_ret,
                profit_generated=prof,
                profit_entry_date=_parse_date(request.form.get("profit_entry_date")),
                profit_notes=(request.form.get("profit_notes") or "").strip() or None,
                profit_distribution_frequency=freq if freq in valid_freq else None,
                next_distribution_review_date=_parse_date(request.form.get("next_distribution_review_date")),
                start_date=_parse_date(request.form.get("start_date")),
                end_date=_parse_date(request.form.get("end_date")),
                project_manager=request.form.get("project_manager", "").strip(),
                status=new_status,
                project_id=pid,
                created_by_user_id=current_user.id if current_user.is_authenticated else None,
            )
            if not inv.name:
                flash("Investment name is required.", "danger")
                return render_template("investments/form.html", **ctx)
            db.session.add(inv)
            db.session.flush()
            if prof != 0:
                db.session.add(
                    InvestmentProfitLog(
                        investment_id=inv.id,
                        amount=prof,
                        entry_date=inv.profit_entry_date or date.today(),
                        notes=inv.profit_notes,
                        created_by_user_id=current_user.id if current_user.is_authenticated else None,
                    )
                )
            log_audit("investment_created", "Investment", None, inv.name)
            db.session.commit()
            uid = current_user.id if current_user.is_authenticated else None
            try:
                post_investment_deployment_delta(
                    inv.id,
                    Decimal("0"),
                    inv.total_amount_invested or Decimal("0"),
                    user_id=uid,
                )
                post_profit_recognition_delta(inv.id, Decimal("0"), prof, user_id=uid)
                post_capital_return_delta(inv.id, Decimal("0"), cap_ret, user_id=uid)
                db.session.commit()
            except Exception as exc:
                db.session.rollback()
                flash(gl_posting_error_message(exc), "warning")
            flash("Investment saved.", "success")
            return redirect(url_for("investments_profile", id=inv.id))
        return render_template("investments/form.html", **ctx)

    @app.route("/investments/<int:id>/edit", methods=["GET", "POST"])
    def investments_edit(id):
        inv = Investment.query.options(joinedload(Investment.project)).get_or_404(id)
        projects = Project.query.order_by(Project.name).all()
        settings = get_or_create_settings()
        ctx = {
            "investment": inv,
            "projects": projects,
            "pool": _funds_pool_summary(),
            "investment_statuses": INVESTMENT_STATUSES,
            "categories": PROJECT_CATEGORIES,
            "profit_distribution_frequencies": PROFIT_DISTRIBUTION_FREQUENCIES,
            "locked_completed": inv.status == "Completed",
            "is_admin_view": current_user.role == "admin",
            "preselect_project_id": inv.project_id,
            "settings": settings,
            "currency_code": settings.currency_code or "USD",
            "banner_project": inv.project,
        }
        if request.method == "POST":
            old_deployed = inv.total_amount_invested or Decimal("0")
            old_cap = inv.capital_returned or Decimal("0")
            ctx["preselect_project_id"] = request.form.get("project_id", type=int) or inv.project_id
            ctx["banner_project"] = (
                Project.query.get(ctx["preselect_project_id"]) if ctx["preselect_project_id"] else None
            )
            if inv.status == "Completed":
                if current_user.role != "admin":
                    flash("Completed investments are locked. Only an administrator can edit them.", "danger")
                    return render_template("investments/form.html", **ctx)
                if request.form.get("admin_override_completed") != "1":
                    flash(
                        "Completed investments are locked. Check admin override to save changes.",
                        "danger",
                    )
                    return render_template("investments/form.html", **ctx)
            pid = request.form.get("project_id", type=int) or None
            amt = _parse_decimal(request.form.get("total_amount_invested")) or Decimal("0")
            old_profit = inv.profit_generated or Decimal("0")
            new_profit = _parse_decimal(request.form.get("profit_generated"))
            if new_profit is None:
                new_profit = Decimal("0")
            name = request.form.get("name", "").strip()
            investment_type = request.form.get("investment_type", "").strip()
            cap_ret = _parse_decimal(request.form.get("capital_returned")) or Decimal("0")
            profit_entry_date = _parse_date(request.form.get("profit_entry_date"))
            profit_notes = (request.form.get("profit_notes") or "").strip() or None
            start_date = _parse_date(request.form.get("start_date"))
            end_date = _parse_date(request.form.get("end_date"))
            project_manager = request.form.get("project_manager", "").strip()
            new_status = _valid_investment_status(request.form.get("status"))
            proj = None
            if pid:
                proj = Project.query.get(pid)
                if not proj:
                    flash("Invalid project.", "danger")
                    return render_template("investments/form.html", **ctx)
                if proj.status == "Closed":
                    flash("Cannot link an investment to a closed project.", "danger")
                    return render_template("investments/form.html", **ctx)
            if amt <= 0:
                flash("Enter a positive total amount invested.", "danger")
                return render_template("investments/form.html", **ctx)
            if not name:
                flash("Investment name is required.", "danger")
                return render_template("investments/form.html", **ctx)
            avail = available_pool_for_investment(
                exclude_investment_id=inv.id, verified_only=_pool_verified_only()
            )
            if amt > avail:
                flash("Investment exceeds available funds. Please reduce amount.", "danger")
                return render_template("investments/form.html", **ctx)
            if proj is not None:
                hr = project_budget_headroom(proj, exclude_investment_id=inv.id)
                if hr is not None and amt > hr:
                    flash(
                        f"Amount exceeds this project's remaining budget ({hr} USD under the project budget cap).",
                        "danger",
                    )
                    return render_template("investments/form.html", **ctx)
            if new_profit < 0 and new_status not in ("Planned", "Active", "Suspended"):
                flash("Negative profit is only allowed while the investment is Planned, Active, or Suspended.", "danger")
                return render_template("investments/form.html", **ctx)
            inv.name = name
            inv.investment_type = investment_type
            inv.capital_returned = cap_ret
            inv.profit_generated = new_profit
            inv.profit_entry_date = profit_entry_date
            inv.profit_notes = profit_notes
            inv.start_date = start_date
            inv.end_date = end_date
            inv.project_manager = project_manager
            inv.status = new_status
            inv.total_amount_invested = amt
            inv.project_id = pid
            freq = (request.form.get("profit_distribution_frequency") or "").strip()
            valid_freq = {k for k, _ in PROFIT_DISTRIBUTION_FREQUENCIES}
            inv.profit_distribution_frequency = freq if freq in valid_freq else None
            inv.next_distribution_review_date = _parse_date(request.form.get("next_distribution_review_date"))
            if new_profit != old_profit:
                db.session.add(
                    InvestmentProfitLog(
                        investment_id=inv.id,
                        amount=new_profit,
                        entry_date=profit_entry_date or date.today(),
                        notes=profit_notes,
                        created_by_user_id=current_user.id if current_user.is_authenticated else None,
                    )
                )
            undist = inv.profit_undistributed_balance()
            log_audit("investment_updated", "Investment", inv.id)
            db.session.commit()
            uid = current_user.id if current_user.is_authenticated else None
            try:
                post_investment_deployment_delta(
                    inv.id,
                    old_deployed,
                    inv.total_amount_invested or Decimal("0"),
                    user_id=uid,
                )
                post_profit_recognition_delta(inv.id, old_profit, new_profit, user_id=uid)
                post_capital_return_delta(inv.id, old_cap, cap_ret, user_id=uid)
                db.session.commit()
            except Exception as exc:
                db.session.rollback()
                flash(gl_posting_error_message(exc), "warning")
            if inv.status == "Closed" and undist > 0:
                flash(
                    "Investment updated. Warning: recorded profit still exceeds amounts distributed to members.",
                    "warning",
                )
            else:
                flash("Investment updated.", "success")
            return redirect(url_for("investments_profile", id=inv.id))
        return render_template("investments/form.html", **ctx)

    @app.route("/investments/<int:id>")
    def investments_profile(id):
        inv = (
            Investment.query.options(joinedload(Investment.project))
            .filter_by(id=id)
            .first_or_404()
        )

        def _scoped_subscriptions_for_investment(inv_id: int):
            q = ShareSubscription.query.filter(
                ShareSubscription.investment_id == inv_id,
                ShareSubscription.status != "Cancelled",
            )
            if current_user.role == "agent" and current_user.agent_id:
                q = q.join(Member, ShareSubscription.member_id == Member.id).filter(
                    Member.agent_id == current_user.agent_id
                )
            return q

        dist_total = inv.profit_distributed_to_members()
        undist = inv.profit_undistributed_balance()
        members_benefit_count = (
            db.session.query(func.count(func.distinct(ProfitDistribution.member_id)))
            .filter(ProfitDistribution.investment_id == inv.id)
            .scalar()
            or 0
        )
        subscription_count = _scoped_subscriptions_for_investment(inv.id).count()
        sum_subscribed_scalar = (
            _scoped_subscriptions_for_investment(inv.id)
            .with_entities(func.coalesce(func.sum(ShareSubscription.subscribed_amount), 0))
            .scalar()
        )
        sum_subscribed = Decimal(str(sum_subscribed_scalar or 0))

        recent_subscriptions = (
            _scoped_subscriptions_for_investment(inv.id)
            .options(joinedload(ShareSubscription.member))
            .order_by(ShareSubscription.subscription_date.desc(), ShareSubscription.id.desc())
            .limit(8)
            .all()
        )
        recent_profit_batches = (
            ProfitDistributionBatch.query.filter_by(investment_id=inv.id)
            .order_by(
                ProfitDistributionBatch.distribution_date.desc(),
                ProfitDistributionBatch.id.desc(),
            )
            .limit(8)
            .all()
        )
        profit_batch_total = inv.profit_batches.count()

        profit_logs = (
            InvestmentProfitLog.query.filter_by(investment_id=inv.id)
            .order_by(InvestmentProfitLog.created_at.desc())
            .limit(50)
            .all()
        )
        ledger_entries = (
            InvestmentLedger.query.filter_by(investment_id=inv.id)
            .order_by(InvestmentLedger.created_at.desc())
            .limit(50)
            .all()
        )
        can_delete = (
            (inv.profit_generated or Decimal("0")) == Decimal("0")
            and dist_total == Decimal("0")
            and inv.profit_batches.count() == 0
            and InvestmentProfitLog.query.filter_by(investment_id=inv.id).count() == 0
        )
        settings = get_or_create_settings()
        pool = _funds_pool_summary()
        currency_code = settings.currency_code or "USD"
        return render_template(
            "investments/profile.html",
            investment=inv,
            dist_total=dist_total,
            undist=undist,
            members_benefit_count=members_benefit_count,
            profit_logs=profit_logs,
            can_delete=can_delete,
            settings=settings,
            currency_code=currency_code,
            pool=pool,
            subscription_count=subscription_count,
            sum_subscribed=sum_subscribed,
            recent_subscriptions=recent_subscriptions,
            recent_profit_batches=recent_profit_batches,
            profit_batch_total=profit_batch_total,
            ledger_entries=ledger_entries,
            profit_distribution_labels=dict(PROFIT_DISTRIBUTION_FREQUENCIES),
        )

    @app.route("/investments/<int:id>/ledger-snapshot", methods=["POST"])
    @role_required("admin", "operator")
    def investments_ledger_snapshot(id):
        inv = Investment.query.get_or_404(id)
        dist = inv.profit_distributed_to_members()
        und = inv.profit_undistributed_balance()
        row = InvestmentLedger(
            investment_id=inv.id,
            capital_invested=inv.total_amount_invested or Decimal("0"),
            capital_returned=inv.capital_returned or Decimal("0"),
            profit_generated=inv.profit_generated or Decimal("0"),
            profit_distributed=dist,
            profit_undistributed=und,
            notes=(request.form.get("ledger_notes") or "").strip()[:500] or None,
        )
        db.session.add(row)
        log_audit("investment_ledger_snapshot", "InvestmentLedger", None, f"investment_id={inv.id}")
        db.session.commit()
        flash("Financial snapshot recorded on the investment ledger.", "success")
        return redirect(url_for("investments_profile", id=id))

    @app.route("/investments/<int:id>/delete", methods=["POST"])
    @admin_required
    def investments_delete(id):
        inv = Investment.query.get_or_404(id)
        dist_total = inv.profit_distributed_to_members()
        if (inv.profit_generated or Decimal("0")) != Decimal("0"):
            flash("Cannot delete: profit has been recorded on this investment.", "danger")
            return redirect(url_for("investments_profile", id=id))
        if dist_total > 0:
            flash("Cannot delete: profit has already been distributed to members.", "danger")
            return redirect(url_for("investments_profile", id=id))
        if inv.profit_batches.count() > 0:
            flash("Cannot delete: distribution batches exist for this investment.", "danger")
            return redirect(url_for("investments_profile", id=id))
        if InvestmentProfitLog.query.filter_by(investment_id=inv.id).count() > 0:
            flash("Cannot delete: profit change history exists for this investment.", "danger")
            return redirect(url_for("investments_profile", id=id))
        name = inv.name
        db.session.delete(inv)
        log_audit("investment_deleted", "Investment", None, name)
        db.session.commit()
        flash("Investment deleted.", "success")
        return redirect(url_for("investments_list"))

    # --- Profit distribution ---
    def _open_investments_for_profit():
        invs = (
            Investment.query.filter(Investment.status.in_(["Active", "Completed"]))
            .order_by(Investment.name)
            .all()
        )
        return [i for i in invs if i.profit_undistributed_balance() > Decimal("0")]

    @app.route("/profit", methods=["GET", "POST"])
    @role_required("admin", "operator")
    def profit_distribute():
        def _profit_form_ctx():
            settings = get_or_create_settings()
            invs = _open_investments_for_profit()
            inv_avail_map = {i.id: i.profit_undistributed_balance() for i in invs}
            inv_eligible_pool = eligible_pools_for_investments(invs)
            inv_avail_json = {str(k): float(v) for k, v in inv_avail_map.items()}
            total_avail = sum(inv_avail_map.values(), Decimal("0"))
            preselect = request.args.get("investment_id", type=int)
            if request.method == "POST":
                preselect = request.form.get("investment_id", type=int) or preselect
            recent_batches = (
                ProfitDistributionBatch.query.options(joinedload(ProfitDistributionBatch.investment))
                .order_by(ProfitDistributionBatch.created_at.desc())
                .limit(12)
                .all()
            )
            review_until = date.today() + timedelta(days=90)
            upcoming_reviews = (
                Investment.query.filter(
                    Investment.next_distribution_review_date.isnot(None),
                    Investment.next_distribution_review_date <= review_until,
                    Investment.status.in_(["Active", "Completed"]),
                )
                .order_by(Investment.next_distribution_review_date.asc())
                .limit(20)
                .all()
            )
            return {
                "investments": invs,
                "inv_avail": inv_avail_map,
                "inv_eligible_pool": inv_eligible_pool,
                "inv_avail_json": inv_avail_json,
                "profit_basis_verified": profit_basis_verified_only(),
                "profit_investment_scoped": settings.get_flag("profit_use_investment_scope", default=False),
                "profit_global_fully_paid": (
                    get_or_create_settings().get_extra().get("profit_global_fully_paid_only") is True
                ),
                "settings": settings,
                "currency_code": settings.currency_code or "USD",
                "preselect_investment_id": preselect,
                "total_undistributed_available": total_avail,
                "open_investment_count": len(invs),
                "recent_batches": recent_batches,
                "upcoming_reviews": upcoming_reviews,
                "profit_review_labels": dict(PROFIT_DISTRIBUTION_FREQUENCIES),
                "today_iso": date.today().isoformat(),
            }

        if request.method == "POST":
            action = (request.form.get("action") or "preview").strip().lower()
            inv_id = request.form.get("investment_id", type=int)
            profit_amt = _parse_decimal(request.form.get("profit_amount"))
            dist_date = _parse_date(request.form.get("distribution_date")) or date.today()
            cur_code = get_or_create_settings().currency_code or "USD"
            if not inv_id or profit_amt is None or profit_amt <= 0:
                flash("Choose an investment and enter profit amount to distribute.", "danger")
                return render_template("profit/form.html", **_profit_form_ctx())
            inv = Investment.query.get_or_404(inv_id)
            if inv.status not in ("Active", "Completed"):
                flash("Profit distribution is only allowed for Active or Completed investments.", "danger")
                return render_template("profit/form.html", **_profit_form_ctx())
            avail_profit = inv.profit_undistributed_balance()
            if avail_profit <= 0:
                flash("No available profit to distribute for this investment.", "danger")
                return render_template("profit/form.html", **_profit_form_ctx())
            if profit_amt > avail_profit:
                flash(
                    f"Distribution exceeds available profit ({avail_profit} {cur_code}). "
                    "Record profit on the investment first, or reduce the amount.",
                    "danger",
                )
                return render_template("profit/form.html", **_profit_form_ctx())
            preview_rows, total_eligible = build_profit_distribution_preview(inv, profit_amt)
            policy = policy_label_for_batch()

            if total_eligible <= 0 or not preview_rows:
                flash(
                    "No eligible paid balances found. Profit share uses only real money received (paid contributions). "
                    "Check: (1) active members with payments on share subscriptions; (2) unlinked contributions only "
                    "count in global mode; (3) if investment-scoped profit is enabled, link subscriptions to this investment.",
                    "warning",
                )
                return render_template("profit/form.html", **_profit_form_ctx())

            if action != "confirm":
                _s = get_or_create_settings()
                shares_sum = sum((r["share"] for r in preview_rows), Decimal("0"))
                batch_notes = (request.form.get("batch_notes") or "").strip()[:500] or ""
                session["profit_pending_batch_notes"] = batch_notes
                session["profit_pending_inv"] = inv_id
                return render_template(
                    "profit/preview.html",
                    investment=inv,
                    profit_amount=profit_amt,
                    distribution_date=dist_date,
                    policy_used=policy,
                    total_eligible=total_eligible.quantize(Decimal("0.01")),
                    preview_rows=preview_rows,
                    preview_shares_total=shares_sum.quantize(Decimal("0.01")),
                    profit_available_before=avail_profit,
                    profit_basis_verified=profit_basis_verified_only(),
                    batch_notes=batch_notes,
                    settings=_s,
                    currency_code=_s.currency_code or "USD",
                )

            raw_notes = (request.form.get("batch_notes") or "").strip()
            if raw_notes:
                batch_notes = raw_notes[:500]
                session.pop("profit_pending_batch_notes", None)
                session.pop("profit_pending_inv", None)
            elif session.get("profit_pending_inv") == inv_id:
                batch_notes = (session.pop("profit_pending_batch_notes", "") or "").strip()[:500] or None
                session.pop("profit_pending_inv", None)
            else:
                batch_notes = None
            batch = ProfitDistributionBatch(
                batch_no=next_profit_batch_no(),
                investment_id=inv.id,
                distribution_date=dist_date,
                total_profit_input=profit_amt,
                total_profit_distributed=sum((r["share"] for r in preview_rows), Decimal("0")),
                policy_used=policy,
                notes=batch_notes,
                created_by_user_id=current_user.id if current_user.is_authenticated else None,
            )
            db.session.add(batch)
            db.session.flush()
            for r in preview_rows:
                m = r["member"]
                db.session.add(
                    EligibilitySnapshot(
                        batch_id=batch.id,
                        member_id=m.id,
                        eligible_amount=r["eligible_amount"],
                        ownership_pct=r["pct"],
                        reason="paid_real_money_basis",
                    )
                )
                db.session.add(
                    ProfitDistribution(
                        batch_id=batch.id,
                        investment_id=inv.id,
                        member_id=m.id,
                        amount=r["share"],
                        eligible_amount_basis=r["eligible_amount"],
                        share_percentage=r["pct"],
                        distribution_date=dist_date,
                        profit_pool_amount=profit_amt,
                        created_by_user_id=current_user.id if current_user.is_authenticated else None,
                    )
                )
            for r in preview_rows:
                mem = db.session.get(Member, r["member"].id)
                if mem is not None:
                    share_amt = r["share"]
                    mem.total_profit_received = (mem.total_profit_received or Decimal("0")) + share_amt
                    if mem.last_profit_distribution_date is None or dist_date > mem.last_profit_distribution_date:
                        mem.last_profit_distribution_date = dist_date
            log_audit(
                "profit_distributed",
                "ProfitDistributionBatch",
                batch.id,
                f"batch_no={batch.batch_no} amount={profit_amt} members={len(preview_rows)} policy={policy}",
            )
            db.session.commit()
            inv_nm = (inv.name or inv.investment_code or "Investment").strip()
            for pr in preview_rows:
                pm = pr["member"]
                try:
                    notify_member_profit_share(
                        pm,
                        amount=pr["share"],
                        investment_name=inv_nm,
                        batch_no=batch.batch_no,
                        distribution_date=dist_date,
                    )
                except Exception:
                    pass
            try:
                post_ok = post_profit_distribution_batch(
                    batch.id,
                    user_id=current_user.id if current_user.is_authenticated else None,
                )
                if post_ok:
                    db.session.commit()
                elif accounting_enabled():
                    flash(
                        "Profit batch was saved, but the cash distribution journal was not posted. "
                        "Check that system accounts exist and are active in the chart of accounts.",
                        "warning",
                    )
            except Exception as exc:
                db.session.rollback()
                flash(gl_posting_error_message(exc), "warning")
            flash(f"Profit batch {batch.batch_no} recorded successfully.", "success")
            return redirect(url_for("profit_batch_detail", batch_id=batch.id))

        return render_template("profit/form.html", **_profit_form_ctx())

    @app.route("/profit/batch/<int:batch_id>")
    @role_required("admin", "operator")
    def profit_batch_detail(batch_id):
        batch = (
            ProfitDistributionBatch.query.options(joinedload(ProfitDistributionBatch.investment))
            .filter_by(id=batch_id)
            .first_or_404()
        )
        rows = (
            ProfitDistribution.query.filter(ProfitDistribution.batch_id == batch_id)
            .options(joinedload(ProfitDistribution.member))
            .order_by(ProfitDistribution.amount.desc())
            .all()
        )
        total_out = batch.total_profit_distributed or Decimal("0")
        highest = rows[0] if rows else None
        settings = get_or_create_settings()
        return render_template(
            "profit/batch_complete.html",
            batch=batch,
            rows=rows,
            total_out=total_out,
            highest=highest,
            member_count=len(rows),
            settings=settings,
            currency_code=settings.currency_code or "USD",
            policy_used=batch.policy_used or "",
        )

    @app.route("/profit/history")
    def profit_history():
        rows = (
            profit_rows_scope()
            .options(
                joinedload(ProfitDistribution.batch),
                joinedload(ProfitDistribution.member),
                joinedload(ProfitDistribution.investment),
            )
            .order_by(ProfitDistribution.distribution_date.desc(), ProfitDistribution.id.desc())
            .all()
        )
        settings = get_or_create_settings()
        cur = settings.currency_code or "USD"
        total_amt = sum((r.amount for r in rows), Decimal("0"))
        batch_ids = {r.batch_id for r in rows if r.batch_id}
        return render_template(
            "profit/history.html",
            rows=rows,
            settings=settings,
            currency_code=cur,
            history_totals={
                "lines": len(rows),
                "amount": total_amt,
                "batches": len(batch_ids),
            },
        )

    @app.route("/profit/statement/<int:member_id>")
    def profit_statement(member_id):
        m = members_scope().filter_by(id=member_id).first_or_404()
        rows = m.profit_rows.order_by(ProfitDistribution.distribution_date.desc()).all()
        return render_template("profit/statement.html", member=m, rows=rows)

    # --- Settings ---
    @app.route("/settings", methods=["GET", "POST"])
    @admin_required
    def settings_page():
        s = get_or_create_settings()
        if request.method == "POST":
            s.currency_code = request.form.get("currency_code", "USD").strip()[:10]
            s.currency_symbol = request.form.get("currency_symbol", "$").strip()[:8]
            s.contribution_rules = request.form.get("contribution_rules", "")
            s.profit_rules = request.form.get("profit_rules", "")
            ex = s.get_extra()
            ex["require_agent_on_member"] = request.form.get("require_agent_on_member") == "1"
            ex["auto_issue_certificate"] = request.form.get("auto_issue_certificate") == "1"
            ex["require_verification_for_certificate"] = (
                request.form.get("require_verification_for_certificate") == "1"
            )
            ex["profit_use_investment_scope"] = request.form.get("profit_use_investment_scope") == "1"
            ex["profit_basis_verified_only"] = request.form.get("profit_basis_verified_only") == "1"
            ex["profit_global_fully_paid_only"] = request.form.get("profit_global_fully_paid_only") == "1"
            ex["accounting_enabled"] = request.form.get("accounting_enabled") == "1"
            ex["pool_use_verified_contributions"] = request.form.get("pool_use_verified_contributions") == "1"
            ex["company_name"] = request.form.get("company_name", "").strip()
            ex["company_address"] = request.form.get("company_address", "").strip()
            ex["company_registration"] = request.form.get("company_registration", "").strip()
            ex["authorized_signatory"] = request.form.get("authorized_signatory", "").strip()
            ex["signatory_title"] = request.form.get("signatory_title", "").strip()
            ex["second_signatory"] = request.form.get("second_signatory", "").strip()
            ex["second_signatory_title"] = request.form.get("second_signatory_title", "").strip()

            branding_note = []
            if request.form.get("clear_logo_light") == "1":
                _remove_brand_file(ex.get("logo_light"))
                ex.pop("logo_light", None)
                branding_note.append("cleared light logo")
            else:
                fl = request.files.get("logo_light")
                if fl and fl.filename:
                    new_lt = _save_brand_logo(fl, "logo-light")
                    if new_lt:
                        _remove_brand_file(ex.get("logo_light"))
                        ex["logo_light"] = new_lt
                        branding_note.append("updated light logo")
                    else:
                        flash(
                            "Light logo was not saved — use PNG, JPG, JPEG, GIF, or WebP.",
                            "warning",
                        )
            if request.form.get("clear_logo_dark") == "1":
                _remove_brand_file(ex.get("logo_dark"))
                ex.pop("logo_dark", None)
                branding_note.append("cleared dark logo")
            else:
                fd = request.files.get("logo_dark")
                if fd and fd.filename:
                    new_dk = _save_brand_logo(fd, "logo-dark")
                    if new_dk:
                        _remove_brand_file(ex.get("logo_dark"))
                        ex["logo_dark"] = new_dk
                        branding_note.append("updated dark logo")
                    else:
                        flash(
                            "Dark logo was not saved — use PNG, JPG, JPEG, GIF, or WebP.",
                            "warning",
                        )
            if branding_note:
                log_audit("settings_branding_updated", "AppSettings", 1, "; ".join(branding_note))

            s.set_extra(ex)
            db.session.commit()
            flash("Settings saved.", "success")
            return redirect(url_for("settings_page"))
        return render_template(
            "settings/index.html",
            settings=s,
            extra=_extra_for_settings_template(s.get_extra()),
            pool=_funds_pool_summary(),
        )

    @app.route("/setting")
    @app.route("/setting/")
    @admin_required
    def settings_page_alias():
        return redirect(url_for("settings_page"))

    @app.route("/settings/notifications", methods=["GET", "POST"])
    @admin_required
    def settings_notifications():
        s = get_or_create_settings()
        ex = s.get_extra()
        if request.method == "POST":
            action = (request.form.get("action") or "").strip()
            if action == "test_email":
                test_to = (request.form.get("test_email_to") or "").strip()
                if not test_to or "@" not in test_to:
                    flash("Enter a valid email address for the test.", "danger")
                elif not mail_configured():
                    flash("SMTP is not fully configured (host and sender required).", "warning")
                else:
                    ok, err = send_email(
                        test_to,
                        "Estithmar — SMTP test",
                        "This is a test message from your Estithmar notification settings.\n\nIf you received this, SMTP is configured correctly.",
                    )
                    if ok:
                        flash("Test email sent.", "success")
                    else:
                        flash(f"Test email failed: {err}", "danger")
                return redirect(url_for("settings_notifications"))

            ex["smtp_host"] = (request.form.get("smtp_host") or "").strip()
            pport = (request.form.get("smtp_port") or "").strip()
            ex["smtp_port"] = pport if pport else "587"
            ex["smtp_use_tls"] = request.form.get("smtp_use_tls") == "1"
            ex["smtp_username"] = (request.form.get("smtp_username") or "").strip()
            if request.form.get("smtp_password_clear") == "1":
                ex.pop("smtp_password", None)
            else:
                pw = (request.form.get("smtp_password") or "").strip()
                if pw:
                    ex["smtp_password"] = pw
            ex["smtp_sender"] = (request.form.get("smtp_sender") or "").strip()

            if request.form.get("twilio_token_clear") == "1":
                ex.pop("twilio_auth_token", None)
            else:
                twt = (request.form.get("twilio_auth_token") or "").strip()
                if twt:
                    ex["twilio_auth_token"] = twt
            ex["twilio_account_sid"] = (request.form.get("twilio_account_sid") or "").strip()
            ex["twilio_whatsapp_from"] = (request.form.get("twilio_whatsapp_from") or "").strip()
            ex["whatsapp_default_cc"] = (request.form.get("whatsapp_default_cc") or "").strip().lstrip("+")

            ex["notify_members_enabled"] = request.form.get("notify_members_enabled") == "1"
            ex["notify_member_payment"] = request.form.get("notify_member_payment") == "1"
            ex["notify_member_subscription"] = request.form.get("notify_member_subscription") == "1"
            ex["notify_member_profit"] = request.form.get("notify_member_profit") == "1"
            ex["notify_member_certificate"] = request.form.get("notify_member_certificate") == "1"
            ex["notify_members_whatsapp"] = request.form.get("notify_members_whatsapp") == "1"

            s.set_extra(ex)
            db.session.commit()
            log_audit("settings_notifications_updated", "AppSettings", s.id, "")
            flash("Notification settings saved.", "success")
            return redirect(url_for("settings_notifications"))

        return render_template(
            "settings/notifications.html",
            settings=s,
            extra=ex,
            mail_ok=mail_configured(),
        )

    @app.route("/settings/payment-methods", methods=["GET", "POST"])
    @admin_required
    def settings_payment_methods():
        if request.method == "POST":
            act = (request.form.get("action") or "").strip()
            if act == "add_bank":
                name = (request.form.get("bank_name") or "").strip()[:120]
                if name:
                    mx = db.session.query(func.coalesce(func.max(PaymentBank.sort_order), 0)).scalar() or 0
                    db.session.add(PaymentBank(name=name, sort_order=int(mx) + 1))
                    db.session.commit()
                    flash("Bank added.", "success")
                else:
                    flash("Enter a bank name.", "danger")
            elif act == "delete_bank":
                bid = request.form.get("bank_id", type=int)
                b = db.session.get(PaymentBank, bid) if bid else None
                if b:
                    db.session.delete(b)
                    db.session.commit()
                    flash("Bank and its accounts were removed.", "info")
            elif act == "add_account":
                bid = request.form.get("bank_id", type=int)
                b = db.session.get(PaymentBank, bid) if bid else None
                acct_no = (request.form.get("account_number") or "").strip()[:120]
                label = (request.form.get("account_label") or "").strip()[:120] or None
                notes = (request.form.get("account_notes") or "").strip()[:300] or None
                if b and acct_no:
                    mx = (
                        db.session.query(func.coalesce(func.max(PaymentBankAccount.sort_order), 0))
                        .filter(PaymentBankAccount.bank_id == b.id)
                        .scalar()
                        or 0
                    )
                    db.session.add(
                        PaymentBankAccount(
                            bank_id=b.id,
                            label=label,
                            account_number=acct_no,
                            notes=notes,
                            sort_order=int(mx) + 1,
                        )
                    )
                    db.session.commit()
                    flash("Bank account added.", "success")
                else:
                    flash("Select a bank and enter an account number.", "danger")
            elif act == "delete_account":
                aid = request.form.get("account_id", type=int)
                a = db.session.get(PaymentBankAccount, aid) if aid else None
                if a:
                    db.session.delete(a)
                    db.session.commit()
                    flash("Bank account removed.", "info")
            elif act == "add_mobile":
                name = (request.form.get("mobile_name") or "").strip()[:120]
                if name:
                    mx = db.session.query(func.coalesce(func.max(PaymentMobileProvider.sort_order), 0)).scalar() or 0
                    db.session.add(PaymentMobileProvider(name=name, sort_order=int(mx) + 1))
                    db.session.commit()
                    flash("Mobile payment option added.", "success")
                else:
                    flash("Enter a name (e.g. EVC, eDahab).", "danger")
            elif act == "delete_mobile":
                mid = request.form.get("mobile_id", type=int)
                mp = db.session.get(PaymentMobileProvider, mid) if mid else None
                if mp:
                    db.session.delete(mp)
                    db.session.commit()
                    flash("Mobile payment option removed.", "info")
            return redirect(url_for("settings_payment_methods"))

        banks = PaymentBank.query.order_by(PaymentBank.sort_order, PaymentBank.name).all()
        banks_with_accounts = [
            (
                b,
                PaymentBankAccount.query.filter_by(bank_id=b.id)
                .order_by(PaymentBankAccount.sort_order, PaymentBankAccount.id)
                .all(),
            )
            for b in banks
        ]
        mobiles = PaymentMobileProvider.query.order_by(
            PaymentMobileProvider.sort_order, PaymentMobileProvider.name
        ).all()
        return render_template(
            "settings/payment_methods.html", banks_with_accounts=banks_with_accounts, mobiles=mobiles
        )

    def _contribution_payment_form_choices():
        """Active bank accounts (with bank name) and mobile providers for the contribution form."""
        acct_rows = (
            PaymentBankAccount.query.join(PaymentBank, PaymentBankAccount.bank_id == PaymentBank.id)
            .filter(PaymentBankAccount.is_active, PaymentBank.is_active)
            .order_by(PaymentBank.sort_order, PaymentBank.name, PaymentBankAccount.sort_order, PaymentBankAccount.id)
            .all()
        )
        bank_accounts = []
        for a in acct_rows:
            lab = (a.label or "Account").strip()
            bank_accounts.append(
                {
                    "id": a.id,
                    "label": f"{a.bank.name} — {lab} — {a.account_number}",
                }
            )
        mobile_list = (
            PaymentMobileProvider.query.filter_by(is_active=True)
            .order_by(PaymentMobileProvider.sort_order, PaymentMobileProvider.name)
            .all()
        )
        return bank_accounts, mobile_list

    # --- Users (admin) ---
    @app.route("/users")
    @admin_required
    def users_list():
        users = (
            AppUser.query.options(joinedload(AppUser.agent), joinedload(AppUser.member))
            .order_by(AppUser.username)
            .all()
        )
        agents = Agent.query.order_by(Agent.full_name).all()
        counts = {
            "total": len(users),
            "active": sum(1 for u in users if u.is_active),
            "inactive": sum(1 for u in users if not u.is_active),
            "admin": sum(1 for u in users if u.role == "admin"),
            "operator": sum(1 for u in users if u.role == "operator"),
            "agent_role": sum(1 for u in users if u.role == "agent"),
            "member_role": sum(1 for u in users if u.role == "member"),
        }
        return render_template(
            "users/list.html",
            users=users,
            agents=agents,
            roles=USER_ROLES,
            role_labels=dict(USER_ROLES),
            counts=counts,
        )

    @app.route("/users/new", methods=["GET", "POST"])
    @admin_required
    def users_new():
        agents = Agent.query.order_by(Agent.full_name).all()
        members_choices = Member.query.order_by(Member.full_name.asc(), Member.id.asc()).all()

        def _form(**kwargs):
            return render_template(
                "users/form.html",
                agents=agents,
                members=members_choices,
                roles=USER_ROLES,
                role_labels=dict(USER_ROLES),
                **kwargs,
            )

        if request.method == "POST":
            username = request.form.get("username", "").strip()
            pwd = request.form.get("password", "")
            role = request.form.get("role") or "operator"
            if role not in ("admin", "operator", "agent", "member"):
                role = "operator"
            aid = request.form.get("agent_id", type=int) or None
            membid = request.form.get("member_id", type=int) or None
            if role != "agent":
                aid = None
            if role != "member":
                membid = None
            email = (request.form.get("email") or "").strip()
            if not email or "@" not in email or len(email) < 5:
                flash("A valid email address is required.", "danger")
                return _form(user=None)
            if AppUser.query.filter(func.lower(AppUser.email) == email.lower()).first():
                flash("That email is already in use.", "danger")
                return _form(user=None)
            if not username or len(pwd) < 4:
                flash("Username and password (min 4 chars) required.", "danger")
                return _form(user=None)
            if AppUser.query.filter_by(username=username).first():
                flash("Username already exists.", "danger")
                return _form(user=None)
            if role == "member" and not membid:
                flash("Select the member record for this portal login.", "danger")
                return _form(user=None)
            if role == "member" and AppUser.query.filter_by(member_id=membid).first():
                flash("That member already has a login account.", "danger")
                return _form(user=None)
            u = AppUser(
                username=username,
                email=email[:120],
                password_hash=generate_password_hash(pwd),
                full_name=request.form.get("full_name", "").strip(),
                role=role,
                agent_id=aid,
                member_id=membid,
                is_active=True,
            )
            db.session.add(u)
            db.session.commit()
            notify_user_credentials(
                to_email=email,
                username=username,
                role_label=dict(USER_ROLES).get(role, role),
                password_plain=pwd,
            )
            flash("User created.", "success")
            return redirect(url_for("users_list"))
        return _form(user=None)

    @app.route("/users/<int:id>/edit", methods=["GET", "POST"])
    @admin_required
    def users_edit(id):
        u = AppUser.query.options(joinedload(AppUser.agent), joinedload(AppUser.member)).get_or_404(id)
        agents = Agent.query.order_by(Agent.full_name).all()
        members_choices = Member.query.order_by(Member.full_name.asc(), Member.id.asc()).all()

        def _form(**kwargs):
            return render_template(
                "users/form.html",
                agents=agents,
                members=members_choices,
                roles=USER_ROLES,
                role_labels=dict(USER_ROLES),
                **kwargs,
            )

        if request.method == "POST":
            u.full_name = request.form.get("full_name", "").strip()
            email = (request.form.get("email") or "").strip()
            if not email or "@" not in email or len(email) < 5:
                flash("A valid email address is required.", "danger")
                return _form(user=u)
            em_ex = AppUser.query.filter(
                func.lower(AppUser.email) == email.lower(), AppUser.id != u.id
            ).first()
            if em_ex:
                flash("That email is already in use.", "danger")
                return _form(user=u)
            u.email = email[:120]
            role = request.form.get("role") or "operator"
            if role not in ("admin", "operator", "agent", "member"):
                role = "operator"
            u.role = role
            aid = request.form.get("agent_id", type=int) or None
            membid = request.form.get("member_id", type=int) or None
            u.agent_id = aid if role == "agent" else None
            u.member_id = membid if role == "member" else None
            if role == "member" and not membid:
                flash("Select the member record for this portal login.", "danger")
                return _form(user=u)
            if role == "member":
                taken = AppUser.query.filter(AppUser.member_id == membid, AppUser.id != u.id).first()
                if taken:
                    flash("That member already has a login account.", "danger")
                    return _form(user=u)
            pwd = request.form.get("password", "").strip()
            if pwd:
                u.password_hash = generate_password_hash(pwd)
            u.is_active = request.form.get("is_active") == "1"
            db.session.commit()
            flash("User updated.", "success")
            return redirect(url_for("users_list"))
        return _form(user=u)

    # --- Accounting (general ledger) ---
    @app.route("/accounting")
    @role_required("admin", "operator")
    def accounting_hub():
        r = forbid_agent()
        if r:
            return r
        ensure_chart_of_accounts()
        db.session.commit()
        tb = trial_balance_rows()
        tb_summary = [r for r in tb if r["debit"] > 0 or r["credit"] > 0]
        total_dr = sum((r["debit"] for r in tb), Decimal("0"))
        total_cr = sum((r["credit"] for r in tb), Decimal("0"))
        settings = get_or_create_settings()
        return render_template(
            "accounting/hub.html",
            settings=settings,
            currency_code=settings.currency_code or "USD",
            trial_balance=tb_summary,
            total_debits=total_dr,
            total_credits=total_cr,
            balanced=(total_dr - total_cr).quantize(Decimal("0.01")) == Decimal("0"),
            accounting_on=accounting_enabled(),
            accounting_section="hub",
        )

    @app.route("/accounting/settings", methods=["GET", "POST"])
    @role_required("admin", "operator")
    def accounting_settings():
        r = forbid_agent()
        if r:
            return r
        s = get_or_create_settings()
        ex = s.get_extra()
        if request.method == "POST":
            month_raw = request.form.get("fiscal_year_start_month", type=int)
            if month_raw is not None and 1 <= month_raw <= 12:
                ex["accounting_fiscal_year_start_month"] = month_raw
            if current_user.role == "admin":
                ex["accounting_enabled"] = request.form.get("accounting_enabled") == "1"
            s.set_extra(ex)
            db.session.commit()
            log_audit("accounting_settings_updated", "AppSettings", s.id, "")
            flash("Accounting settings saved.", "success")
            return redirect(url_for("accounting_settings"))
        settings = s
        return render_template(
            "accounting/settings.html",
            settings=settings,
            extra=ex,
            currency_code=settings.currency_code or "USD",
            accounting_on=accounting_enabled(),
            accounting_section="settings",
        )

    @app.route("/accounting/chart-of-accounts", methods=["GET", "POST"])
    @role_required("admin", "operator")
    def accounting_chart():
        r = forbid_agent()
        if r:
            return r
        ensure_chart_of_accounts()
        db.session.commit()
        settings = get_or_create_settings()
        if request.method == "POST":
            if request.form.get("action") == "toggle_active":
                tid = request.form.get("account_toggle_id", type=int)
                acc = db.session.get(Account, tid) if tid else None
                if not acc:
                    flash("Account not found.", "danger")
                elif acc.system_key and acc.system_key in SYSTEM_ACCOUNT_KEYS:
                    flash("System accounts cannot be deactivated.", "warning")
                else:
                    acc.is_active = not acc.is_active
                    db.session.commit()
                    log_audit(
                        "accounting_chart_account_toggle",
                        "Account",
                        acc.id,
                        f"active={acc.is_active}",
                    )
                    flash(
                        f"Account {acc.code} is now {'active' if acc.is_active else 'inactive'}.",
                        "success",
                    )
                    return redirect(url_for("accounting_chart"))
            else:
                code = (request.form.get("code") or "").strip()[:32]
                name = (request.form.get("name") or "").strip()[:200]
                atype = (request.form.get("account_type") or "").strip().lower()
                sort_order = request.form.get("sort_order", type=int) or 0
                valid_types = ("asset", "liability", "equity", "revenue", "expense")
                if not code or not name or atype not in valid_types:
                    flash("Enter a unique code, name, and valid account type.", "danger")
                elif Account.query.filter_by(code=code).first():
                    flash("That account code already exists.", "danger")
                else:
                    db.session.add(
                        Account(
                            code=code,
                            name=name,
                            account_type=atype,
                            sort_order=sort_order,
                            is_active=True,
                        )
                    )
                    db.session.commit()
                    log_audit("accounting_chart_account_added", "Account", None, f"code={code}")
                    flash(f"Account {code} added.", "success")
                    return redirect(url_for("accounting_chart"))
        rows = Account.query.order_by(Account.sort_order, Account.code).all()
        return render_template(
            "accounting/chart.html",
            accounts=rows,
            settings=settings,
            currency_code=settings.currency_code or "USD",
            accounting_section="chart",
            system_account_keys=SYSTEM_ACCOUNT_KEYS,
        )

    @app.route("/accounting/ledger")
    @role_required("admin", "operator")
    def accounting_ledger():
        r = forbid_agent()
        if r:
            return r
        ensure_chart_of_accounts()
        db.session.commit()
        account_id = request.args.get("account_id", type=int)
        date_from = _parse_date(request.args.get("date_from"))
        date_to = _parse_date(request.args.get("date_to"))
        if date_from and date_to and date_from > date_to:
            flash("From date was after To date — dates were swapped.", "warning")
            date_from, date_to = date_to, date_from
        settings = get_or_create_settings()
        aq = Account.query
        if account_id:
            aq = aq.filter(or_(Account.is_active, Account.id == account_id))
        else:
            aq = aq.filter_by(is_active=True)
        account_list = aq.order_by(Account.sort_order, Account.code).all()
        account = db.session.get(Account, account_id) if account_id else None
        opening_balance = Decimal("0")
        lines: list = []
        if account:
            opening_balance, lines = ledger_lines_for_account(
                account.id,
                date_from=date_from,
                date_to=date_to,
                limit=2000,
            )
        return render_template(
            "accounting/ledger.html",
            settings=settings,
            currency_code=settings.currency_code or "USD",
            accounts=account_list,
            account=account,
            lines=lines,
            opening_balance=opening_balance,
            date_from=date_from,
            date_to=date_to,
            accounting_section="ledger",
        )

    @app.route("/accounting/journal")
    @role_required("admin", "operator")
    def accounting_journal():
        r = forbid_agent()
        if r:
            return r
        date_from = _parse_date(request.args.get("date_from"))
        date_to = _parse_date(request.args.get("date_to"))
        st_raw = request.args.get("source_type") or ""
        source_type = st_raw if st_raw else None
        entries = journal_entries_filtered(
            limit=400,
            date_from=date_from,
            date_to=date_to,
            source_type=source_type,
        )
        settings = get_or_create_settings()
        return render_template(
            "accounting/journal_list.html",
            entries=entries,
            settings=settings,
            currency_code=settings.currency_code or "USD",
            date_from=date_from,
            date_to=date_to,
            source_type=source_type or "",
            journal_source_types=JOURNAL_SOURCE_TYPES,
            accounting_section="journal",
        )

    @app.route("/accounting/journal/<int:entry_id>")
    @role_required("admin", "operator")
    def accounting_journal_detail(entry_id):
        r = forbid_agent()
        if r:
            return r
        je = JournalEntry.query.get_or_404(entry_id)
        lines = lines_for_entry(entry_id)
        settings = get_or_create_settings()
        ids = [ln.account_id for ln in lines]
        acct = (
            {a.id: a for a in Account.query.filter(Account.id.in_(ids)).all()}
            if ids
            else {}
        )
        can_void = je.status == "posted" and je.source_type == "manual"
        return render_template(
            "accounting/journal_detail.html",
            entry=je,
            lines=lines,
            account_map=acct,
            settings=settings,
            currency_code=settings.currency_code or "USD",
            accounting_section="journal",
            can_void=can_void,
        )

    @app.route("/accounting/journal/<int:entry_id>/void", methods=["POST"])
    @role_required("admin", "operator")
    def accounting_journal_void(entry_id):
        r = forbid_agent()
        if r:
            return r
        ok, msg = void_manual_journal_entry(entry_id)
        if ok:
            try:
                db.session.commit()
                log_audit("accounting_journal_void", "JournalEntry", entry_id, "")
                flash("Journal entry voided. It no longer affects the trial balance.", "success")
            except Exception:
                db.session.rollback()
                flash("Could not void entry.", "danger")
        else:
            db.session.rollback()
            flash(msg, "danger")
        return redirect(url_for("accounting_journal_detail", entry_id=entry_id))

    @app.route("/accounting/trial-balance")
    @role_required("admin", "operator")
    def accounting_trial_balance():
        r = forbid_agent()
        if r:
            return r
        ensure_chart_of_accounts()
        db.session.commit()
        tb = trial_balance_rows()
        total_dr = sum((r["debit"] for r in tb), Decimal("0"))
        total_cr = sum((r["credit"] for r in tb), Decimal("0"))
        settings = get_or_create_settings()
        return render_template(
            "accounting/trial_balance.html",
            trial_balance=tb,
            total_debits=total_dr,
            total_credits=total_cr,
            settings=settings,
            currency_code=settings.currency_code or "USD",
            accounting_section="trial",
        )

    @app.route("/accounting/trial-balance/export.csv")
    @role_required("admin", "operator")
    def accounting_export_trial_balance():
        r = forbid_agent()
        if r:
            return r
        ensure_chart_of_accounts()
        db.session.commit()
        tb = trial_balance_rows()
        settings = get_or_create_settings()
        sym = settings.currency_symbol or "$"
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(
            ["Code", "System key", "Account", "Type", f"Debit ({sym})", f"Credit ({sym})", f"Balance ({sym})"]
        )
        for row in tb:
            w.writerow(
                [
                    row["code"],
                    row.get("system_key") or "",
                    row["name"],
                    row["type"],
                    row["debit"],
                    row["credit"],
                    row["balance"],
                ]
            )
        total_dr = sum((row["debit"] for row in tb), Decimal("0"))
        total_cr = sum((row["credit"] for row in tb), Decimal("0"))
        w.writerow(["", "", "Totals", total_dr, total_cr, ""])
        return Response(
            buf.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=trial_balance.csv"},
        )

    @app.route("/accounting/journal/export.csv")
    @role_required("admin", "operator")
    def accounting_export_journal():
        r = forbid_agent()
        if r:
            return r
        date_from = _parse_date(request.args.get("date_from"))
        date_to = _parse_date(request.args.get("date_to"))
        st_raw = request.args.get("source_type") or ""
        source_type = st_raw if st_raw else None
        entries = journal_entries_filtered(
            limit=2000,
            date_from=date_from,
            date_to=date_to,
            source_type=source_type,
        )
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["Date", "Reference", "Source type", "Source ID", "Memo"])
        for e in entries:
            w.writerow(
                [
                    e.entry_date,
                    e.reference or "",
                    e.source_type or "",
                    e.source_id if e.source_id is not None else "",
                    (e.memo or "").replace("\n", " "),
                ]
            )
        return Response(
            buf.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=journal_entries.csv"},
        )

    @app.route("/accounting/manual", methods=["GET", "POST"])
    @role_required("admin", "operator")
    def accounting_manual_entry():
        r = forbid_agent()
        if r:
            return r
        ensure_chart_of_accounts()
        db.session.commit()
        accounts = Account.query.filter_by(is_active=True).order_by(Account.sort_order, Account.code).all()
        accounts_grouped = _accounts_grouped_for_journal(accounts)
        settings = get_or_create_settings()
        cur_code = settings.currency_code or "USD"
        cur_sym = settings.currency_symbol or "$"

        def _manual_render(**extra):
            return render_template(
                "accounting/manual.html",
                accounts=accounts,
                accounts_grouped=accounts_grouped,
                settings=settings,
                currency_code=cur_code,
                currency_symbol=cur_sym,
                accounting_section="manual",
                accounting_gl_enabled=accounting_enabled(),
                **extra,
            )

        if request.method == "POST":
            d = _parse_date(request.form.get("entry_date")) or date.today()
            memo = (request.form.get("memo") or "").strip()[:500]
            ref = (request.form.get("reference") or "").strip()[:64]
            a1 = request.form.get("account_debit_id", type=int)
            a2 = request.form.get("account_credit_id", type=int)
            dr = _parse_decimal(request.form.get("amount_debit"))
            cr = _parse_decimal(request.form.get("amount_credit"))
            form_date = (request.form.get("entry_date") or "").strip() or date.today().isoformat()
            if not a1 or not a2 or a1 == a2:
                flash("Select two different accounts for debit and credit.", "danger")
                return _manual_render(default_entry_date=form_date)
            amt = dr or cr
            if amt is None or amt <= 0:
                flash("Enter a positive amount.", "danger")
                return _manual_render(default_entry_date=form_date)
            amt = amt.quantize(Decimal("0.01"))
            acc1 = db.session.get(Account, a1)
            acc2 = db.session.get(Account, a2)
            if not acc1 or not acc2:
                flash("Invalid account.", "danger")
                return _manual_render(default_entry_date=form_date)
            je = JournalEntry(
                entry_date=d,
                reference=ref or None,
                memo=memo or "Manual adjustment",
                source_type="manual",
                source_id=None,
                status="posted",
                created_by_user_id=current_user.id if current_user.is_authenticated else None,
            )
            db.session.add(je)
            db.session.flush()
            db.session.add(
                JournalLine(
                    journal_entry_id=je.id,
                    account_id=acc1.id,
                    debit=amt,
                    credit=Decimal("0"),
                    description="Manual",
                    line_no=1,
                )
            )
            db.session.add(
                JournalLine(
                    journal_entry_id=je.id,
                    account_id=acc2.id,
                    debit=Decimal("0"),
                    credit=amt,
                    description="Manual",
                    line_no=2,
                )
            )
            log_audit("accounting_manual_entry", "JournalEntry", je.id, ref or "")
            db.session.commit()
            flash("Journal entry posted.", "success")
            return redirect(url_for("accounting_journal_detail", entry_id=je.id))
        return _manual_render(default_entry_date=date.today().isoformat())

    # --- Reports hub ---
    @app.route("/reports")
    def reports_hub():
        return render_template("reports/index.html")

    @app.route("/reports/monthly")
    def reports_monthly():
        y = request.args.get("year", type=int) or datetime.utcnow().year
        m = request.args.get("month", type=int) or datetime.utcnow().month
        agent_id = request.args.get("agent_id", type=int)
        start = date(y, m, 1)
        end = start + relativedelta(months=1) - relativedelta(days=1)
        q = Contribution.query.filter(Contribution.date >= start, Contribution.date <= end).join(Member)
        if agent_id:
            q = q.filter(Member.agent_id == agent_id)
        if current_user.role == "agent" and current_user.agent_id:
            q = q.filter(Member.agent_id == current_user.agent_id)
        rows = (
            q.options(
                joinedload(Contribution.payment_bank_account).joinedload(PaymentBankAccount.bank),
                joinedload(Contribution.payment_mobile_provider),
            )
            .order_by(Contribution.date.desc())
            .all()
        )
        total = sum((r.amount for r in rows), Decimal("0"))
        agents = Agent.query.filter_by(status="Active").order_by(Agent.full_name).all()
        return render_template(
            "reports/monthly.html",
            rows=rows,
            total=total,
            year=y,
            month=m,
            start=start,
            end=end,
            agent_id=agent_id,
            agents=agents,
        )

    @app.route("/reports/member/<int:member_id>")
    def reports_member_contrib(member_id):
        m = members_scope().filter_by(id=member_id).first_or_404()
        rows = (
            m.contributions.order_by(Contribution.date.desc())
            .options(
                joinedload(Contribution.payment_bank_account).joinedload(PaymentBankAccount.bank),
                joinedload(Contribution.payment_mobile_provider),
            )
            .all()
        )
        return render_template("reports/member_contrib.html", member=m, rows=rows)

    @app.route("/reports/agents")
    def reports_agents():
        r = forbid_agent()
        if r:
            return r
        rows = []
        for ag in Agent.query.order_by(Agent.full_name).all():
            rows.append(
                {
                    "agent": ag,
                    "members": ag.members_count(),
                    "volume": ag.contributions_managed_total(),
                    "share_value": ag.total_subscribed_share_value(),
                }
            )
        rows.sort(key=lambda x: x["volume"], reverse=True)
        settings = get_or_create_settings()
        total_members = sum(r["members"] for r in rows)
        total_volume = sum((r["volume"] for r in rows), Decimal("0"))
        total_share = sum((r["share_value"] for r in rows), Decimal("0"))
        return render_template(
            "reports/agents.html",
            rows=rows,
            settings=settings,
            currency_code=settings.currency_code or "USD",
            totals={
                "agents": len(rows),
                "members": total_members,
                "volume": total_volume,
                "share_value": total_share,
            },
        )

    @app.route("/reports/agents/geography")
    def reports_agents_geography():
        """Roll up agents by country and region — collections & member counts (§4.3)."""
        r = forbid_agent()
        if r:
            return r
        settings = get_or_create_settings()
        cur = settings.currency_code or "USD"
        agents = Agent.query.order_by(Agent.country, Agent.region, Agent.full_name).all()

        by_country: dict[str, list[Agent]] = defaultdict(list)
        for ag in agents:
            key = (ag.country or "").strip() or "—"
            by_country[key].append(ag)

        country_rows = []
        for country in sorted(by_country.keys(), key=lambda x: (x == "—", x.lower())):
            group = by_country[country]
            m_ct = sum(ag.members_count() for ag in group)
            vol = sum((ag.contributions_managed_total() for ag in group), Decimal("0"))
            sh = sum((ag.total_subscribed_share_value() for ag in group), Decimal("0"))
            by_region: dict[str, list[Agent]] = defaultdict(list)
            for ag in group:
                rk = (ag.region or "").strip() or "—"
                by_region[rk].append(ag)
            region_detail = []
            for reg in sorted(by_region.keys(), key=lambda x: (x == "—", x.lower())):
                rlist = by_region[reg]
                region_detail.append(
                    {
                        "region": reg,
                        "agents": rlist,
                        "members": sum(ag.members_count() for ag in rlist),
                        "volume": sum((ag.contributions_managed_total() for ag in rlist), Decimal("0")),
                        "share_value": sum((ag.total_subscribed_share_value() for ag in rlist), Decimal("0")),
                    }
                )
            country_rows.append(
                {
                    "country": country,
                    "agents": group,
                    "members": m_ct,
                    "volume": vol,
                    "share_value": sh,
                    "regions": region_detail,
                }
            )

        total_agents = len(agents)
        total_members = sum(ag.members_count() for ag in agents)
        total_volume = sum((ag.contributions_managed_total() for ag in agents), Decimal("0"))
        total_share = sum((ag.total_subscribed_share_value() for ag in agents), Decimal("0"))

        return render_template(
            "reports/agents_geography.html",
            settings=settings,
            currency_code=cur,
            country_rows=country_rows,
            totals={
                "agents": total_agents,
                "members": total_members,
                "volume": total_volume,
                "share_value": total_share,
            },
        )

    @app.route("/reports/installments")
    def reports_installments():
        query = (
            InstallmentPlan.query.join(ShareSubscription)
            .join(Member, ShareSubscription.member_id == Member.id)
            .options(
                joinedload(InstallmentPlan.subscription).joinedload(ShareSubscription.member),
                joinedload(InstallmentPlan.subscription).joinedload(ShareSubscription.investment),
            )
            .order_by(InstallmentPlan.due_date.asc())
        )
        if current_user.role == "agent" and current_user.agent_id:
            query = query.filter(Member.agent_id == current_user.agent_id)
        rows = query.all()
        today = date.today()
        overdue_rows = []
        unpaid_rows = []
        for r in rows:
            if r.status == "Cancelled":
                continue
            bal = (r.due_amount or Decimal("0")) - (r.paid_amount or Decimal("0"))
            if bal > 0:
                unpaid_rows.append((r, bal))
                if r.status == "Overdue" or (
                    r.due_date is not None and r.due_date < today and r.status in {"Pending", "Partially Paid"}
                ):
                    overdue_rows.append((r, bal))
        unpaid_total = sum((b for _, b in unpaid_rows), Decimal("0"))
        overdue_total = sum((b for _, b in overdue_rows), Decimal("0"))
        settings = get_or_create_settings()
        return render_template(
            "reports/installments.html",
            rows=rows,
            unpaid_rows=unpaid_rows,
            overdue_rows=overdue_rows,
            settings=settings,
            currency_code=settings.currency_code or "USD",
            report_date=today,
            totals={
                "unpaid_count": len(unpaid_rows),
                "unpaid_balance": unpaid_total,
                "overdue_count": len(overdue_rows),
                "overdue_balance": overdue_total,
            },
        )

    @app.route("/reports/members-financial")
    def reports_members_financial():
        """Per-member subscribed, paid, outstanding, confirmed value/units, profit — doc §15.1 summary columns."""
        st = request.args.get("status", "").strip()
        only_outstanding = request.args.get("outstanding", "").strip().lower() in {"1", "true", "yes"}
        q = members_scope().options(joinedload(Member.agent))
        if st in ("Active", "Inactive"):
            q = q.filter(Member.status == st)
        members = q.order_by(Member.member_id.asc()).all()
        rows: list[dict] = []
        agg = {
            "subscribed": Decimal("0"),
            "paid": Decimal("0"),
            "outstanding": Decimal("0"),
            "confirmed_value": Decimal("0"),
            "confirmed_units": Decimal("0"),
            "profit": Decimal("0"),
        }
        for m in members:
            subs = m.subscriptions.filter(ShareSubscription.status != "Cancelled").all()
            t_sub = sum((s.subscribed_amount or Decimal("0") for s in subs), Decimal("0"))
            t_paid = sum((s.paid_total() for s in subs), Decimal("0"))
            out = sum((s.outstanding_balance() for s in subs), Decimal("0"))
            conf_v = sum(
                (s.subscribed_amount or Decimal("0") for s in subs if s.status == "Fully Paid"),
                Decimal("0"),
            )
            conf_u = sum(
                (s.share_units_subscribed or Decimal("0") for s in subs if s.status == "Fully Paid"),
                Decimal("0"),
            )
            if only_outstanding and out <= 0:
                continue
            pt = m.lifetime_profit_received()
            rows.append(
                {
                    "member": m,
                    "subscribed": t_sub,
                    "paid": t_paid,
                    "outstanding": out,
                    "confirmed_value": conf_v,
                    "confirmed_units": conf_u,
                    "profit": pt,
                }
            )
            agg["subscribed"] += t_sub
            agg["paid"] += t_paid
            agg["outstanding"] += out
            agg["confirmed_value"] += conf_v
            agg["confirmed_units"] += conf_u
            agg["profit"] += pt
        settings = get_or_create_settings()
        return render_template(
            "reports/members_financial.html",
            rows=rows,
            totals=agg,
            settings=settings,
            currency_code=settings.currency_code or "USD",
            status_filter=st,
            only_outstanding=only_outstanding,
        )

    @app.route("/reports/profit-calculation")
    def reports_profit_calculation():
        """Narrative + hypothetical distribution preview — doc §11 / §15.5 profit calculation summary."""
        r = forbid_agent()
        if r:
            return r
        settings = get_or_create_settings()
        investments = Investment.query.order_by(Investment.name).all()
        inv_id = request.args.get("investment_id", type=int)
        sample_raw = _parse_decimal(request.args.get("sample_amount"))
        sample_amt = sample_raw if sample_raw is not None and sample_raw > 0 else Decimal("20000")
        selected = db.session.get(Investment, inv_id) if inv_id else None
        preview_rows: list = []
        total_eligible = Decimal("0")
        if selected:
            preview_rows, total_eligible = build_profit_distribution_preview(selected, sample_amt)
        policy_key = policy_label_for_batch()
        return render_template(
            "reports/profit_calculation.html",
            settings=settings,
            currency_code=settings.currency_code or "USD",
            investments=investments,
            selected=selected,
            sample_amount=sample_amt,
            preview_rows=preview_rows,
            total_eligible=total_eligible,
            policy_key=policy_key,
            profit_basis_verified=profit_basis_verified_only(),
            profit_investment_scoped=settings.get_flag("profit_use_investment_scope"),
        )

    @app.route("/reports/profit-summary")
    def reports_profit_summary():
        r = forbid_agent()
        if r:
            return r
        inv_rows = Investment.query.order_by(Investment.name).all()
        out = []
        for inv in inv_rows:
            dist_total = (
                db.session.query(func.coalesce(func.sum(ProfitDistribution.amount), 0))
                .filter(ProfitDistribution.investment_id == inv.id)
                .scalar()
                or 0
            )
            generated = inv.profit_generated or Decimal("0")
            distributed = Decimal(str(dist_total))
            undistributed = generated - distributed
            if undistributed < 0:
                undistributed = Decimal("0")
            out.append((inv, distributed, undistributed))
        return render_template("reports/profit_summary.html", rows=out)

    @app.route("/reports/investments/summary")
    def reports_investment_summary():
        r = forbid_agent()
        if r:
            return r
        rows = (
            Investment.query.options(joinedload(Investment.project))
            .order_by(Investment.name)
            .all()
        )
        ids = [x.id for x in rows]
        dist_totals: dict[int, Decimal] = {}
        if ids:
            sums = (
                db.session.query(
                    ProfitDistribution.investment_id,
                    func.coalesce(func.sum(ProfitDistribution.amount), 0),
                )
                .filter(ProfitDistribution.investment_id.in_(ids))
                .group_by(ProfitDistribution.investment_id)
                .all()
            )
            dist_totals = {iid: Decimal(str(s)) for iid, s in sums}
        sum_invested = Decimal("0")
        sum_profit = Decimal("0")
        sum_distributed = Decimal("0")
        sum_remaining = Decimal("0")
        for inv in rows:
            dist = dist_totals.get(inv.id, Decimal("0"))
            gen = inv.profit_generated or Decimal("0")
            rem = gen - dist
            if rem < 0:
                rem = Decimal("0")
            sum_invested += inv.total_amount_invested or Decimal("0")
            sum_profit += gen
            sum_distributed += dist
            sum_remaining += rem
        settings = get_or_create_settings()
        return render_template(
            "reports/investment_summary.html",
            rows=rows,
            dist_totals=dist_totals,
            settings=settings,
            currency_code=settings.currency_code or "USD",
            totals={
                "lines": len(rows),
                "invested": sum_invested,
                "profit": sum_profit,
                "distributed": sum_distributed,
                "remaining": sum_remaining,
            },
        )

    @app.route("/reports/daily")
    def reports_daily():
        d = _parse_date(request.args.get("date")) or date.today()
        q = Contribution.query.filter(Contribution.date == d).join(Member)
        if current_user.role == "agent" and current_user.agent_id:
            q = q.filter(Member.agent_id == current_user.agent_id)
        rows = (
            q.options(
                joinedload(Contribution.payment_bank_account).joinedload(PaymentBankAccount.bank),
                joinedload(Contribution.payment_mobile_provider),
            )
            .order_by(Contribution.id.desc())
            .all()
        )
        total = sum((r.amount for r in rows), Decimal("0"))
        return render_template("reports/daily.html", rows=rows, total=total, report_date=d)

    @app.route("/reports/projects/profitability")
    def reports_projects_profitability():
        r = forbid_agent()
        if r:
            return r
        cat_map = dict(PROJECT_CATEGORIES)
        rows_out = []
        for p in Project.query.order_by(Project.name).all():
            invs = p.investments.all()
            invested = sum((inv.total_amount_invested or Decimal("0") for inv in invs), Decimal("0"))
            profit = sum((inv.profit_generated or Decimal("0") for inv in invs), Decimal("0"))
            cap_ret = sum((inv.capital_returned or Decimal("0") for inv in invs), Decimal("0"))
            budget_raw = p.total_budget
            has_budget = budget_raw is not None
            headroom = None
            util_pct = None
            if has_budget:
                cap = budget_raw or Decimal("0")
                headroom = max(cap - invested, Decimal("0"))
                if cap > 0:
                    util_pct = min(Decimal("100"), invested / cap * Decimal("100"))
            rows_out.append(
                {
                    "project": p,
                    "invested": invested,
                    "profit": profit,
                    "capital_returned": cap_ret,
                    "budget": budget_raw,
                    "has_budget": has_budget,
                    "headroom": headroom,
                    "util_pct": util_pct,
                    "inv_count": len(invs),
                }
            )
        sum_invested = sum((r["invested"] for r in rows_out), Decimal("0"))
        sum_profit = sum((r["profit"] for r in rows_out), Decimal("0"))
        sum_cap_ret = sum((r["capital_returned"] for r in rows_out), Decimal("0"))
        sum_budget = sum(
            (r["project"].total_budget for r in rows_out if r["project"].total_budget is not None),
            Decimal("0"),
        )
        projects_with_budget = sum(1 for r in rows_out if r["has_budget"])
        settings = get_or_create_settings()
        return render_template(
            "reports/project_profitability.html",
            rows=rows_out,
            cat_map=cat_map,
            settings=settings,
            currency_code=settings.currency_code or "USD",
            totals={
                "projects": len(rows_out),
                "invested": sum_invested,
                "profit": sum_profit,
                "capital_returned": sum_cap_ret,
                "budget": sum_budget,
                "projects_with_budget": projects_with_budget,
            },
        )

    @app.route("/reports/community-model")
    def reports_community_model():
        """Community investment model: narrative + pooled collections, share payments, deployment by category, profit basis."""
        mids = _dashboard_scoped_member_ids()

        def _collected_total() -> Decimal:
            q = db.session.query(func.coalesce(func.sum(Contribution.amount), 0))
            if mids is not None:
                if not mids:
                    return Decimal("0")
                q = q.filter(Contribution.member_id.in_(mids))
            return Decimal(str(q.scalar() or 0))

        total_collected = _collected_total()

        sub_q = ShareSubscription.query.filter(ShareSubscription.status != "Cancelled")
        if mids is not None:
            if mids:
                sub_q = sub_q.filter(ShareSubscription.member_id.in_(mids))
            else:
                sub_q = sub_q.filter(ShareSubscription.id == -1)
        scoped_subs = sub_q.all()

        pool_paid = sum((s.paid_total() for s in scoped_subs), Decimal("0"))
        pool_subscribed = sum((s.subscribed_amount or Decimal("0") for s in scoped_subs), Decimal("0"))
        fully_paid_subs = sum(1 for s in scoped_subs if s.status == "Fully Paid")
        installment_plans = sum(1 for s in scoped_subs if s.payment_plan == "installment")

        by_investment_rows = []
        inv_amounts: dict[int, Decimal] = {}
        for s in scoped_subs:
            if s.investment_id:
                iid = s.investment_id
                inv_amounts[iid] = inv_amounts.get(iid, Decimal("0")) + s.paid_total()
        for iid in sorted(inv_amounts.keys(), key=lambda x: inv_amounts[x], reverse=True):
            inv = db.session.get(Investment, iid)
            if inv:
                by_investment_rows.append((inv, inv_amounts[iid]))

        by_category: dict[str, Decimal] = {k: Decimal("0") for k, _ in PROJECT_CATEGORIES}
        by_category["_unlinked"] = Decimal("0")
        for inv in Investment.query.all():
            amt = inv.total_amount_invested or Decimal("0")
            if inv.project_id:
                p = db.session.get(Project, inv.project_id)
                cat = p.category if p else "_unlinked"
                if cat not in by_category:
                    by_category[cat] = Decimal("0")
                by_category[cat] += amt
            else:
                by_category["_unlinked"] += amt

        total_deployed = sum((inv.total_amount_invested or Decimal("0") for inv in Investment.query.all()), Decimal("0"))
        total_profit_org = sum((inv.profit_generated or Decimal("0") for inv in Investment.query.all()), Decimal("0"))

        return render_template(
            "reports/community_model.html",
            total_collected=total_collected,
            pool_paid=pool_paid,
            pool_subscribed=pool_subscribed,
            fully_paid_subs=fully_paid_subs,
            installment_plans=installment_plans,
            subscription_count=len(scoped_subs),
            by_investment_rows=by_investment_rows,
            by_category=by_category,
            total_deployed=total_deployed,
            total_profit_org=total_profit_org,
            project_categories=PROJECT_CATEGORIES,
        )

    # --- Excel exports ---
    def _xlsx_response(data: bytes, filename: str):
        return Response(
            data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    @app.route("/export/members.xlsx")
    def export_members_xlsx():
        wb = Workbook()
        ws = wb.active
        ws.title = "Members"
        kind_labels = dict(MEMBER_KINDS)
        gender_labels = dict(MEMBER_GENDER_CHOICES)
        headers = [
            "Member ID",
            "Full name",
            "Type",
            "Phone",
            "Address",
            "National ID",
            "Date of birth",
            "Gender",
            "Occupation/employer",
            "Next of kin name",
            "Next of kin relationship",
            "Next of kin phone",
            "Next of kin address",
            "Join date",
            "Status",
            "Agent ID",
            "Contributions total",
            "Profit received",
        ]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True)
        sort = request.args.get("sort", "join_desc").strip()
        q = _members_filtered_query().options(joinedload(Member.agent))
        if sort == "name_asc":
            q = q.order_by(Member.full_name.asc(), Member.id.asc())
        elif sort == "name_desc":
            q = q.order_by(Member.full_name.desc(), Member.id.desc())
        elif sort == "member_id":
            q = q.order_by(Member.member_id.asc())
        else:
            q = q.order_by(Member.join_date.desc(), Member.id.desc())
        for row, mem in enumerate(q.all(), 2):
            ws.cell(row=row, column=1, value=mem.member_id)
            ws.cell(row=row, column=2, value=mem.full_name)
            ws.cell(row=row, column=3, value=kind_labels.get(getattr(mem, "member_kind", None) or "member", "Member"))
            ws.cell(row=row, column=4, value=mem.phone)
            ws.cell(row=row, column=5, value=mem.address)
            ws.cell(row=row, column=6, value=mem.national_id)
            ws.cell(row=row, column=7, value=mem.date_of_birth.isoformat() if mem.date_of_birth else "")
            ws.cell(row=row, column=8, value=gender_labels.get(mem.gender, mem.gender or ""))
            ws.cell(row=row, column=9, value=mem.occupation_employer or "")
            ws.cell(row=row, column=10, value=mem.next_of_kin_name or "")
            ws.cell(row=row, column=11, value=mem.next_of_kin_relationship or "")
            ws.cell(row=row, column=12, value=mem.next_of_kin_phone or "")
            ws.cell(row=row, column=13, value=mem.next_of_kin_address or "")
            ws.cell(row=row, column=14, value=mem.join_date.isoformat() if mem.join_date else "")
            ws.cell(row=row, column=15, value=mem.status)
            ws.cell(row=row, column=16, value=mem.agent.agent_id if mem.agent else "")
            ws.cell(row=row, column=17, value=float(mem.contribution_total()))
            ws.cell(row=row, column=18, value=float(mem.profit_received_total()))
        bio = io.BytesIO()
        wb.save(bio)
        return _xlsx_response(bio.getvalue(), "members.xlsx")

    @app.route("/export/agents.xlsx")
    def export_agents_xlsx():
        r = forbid_agent()
        if r:
            return r
        wb = Workbook()
        ws = wb.active
        ws.title = "Agents"
        headers = [
            "Agent ID",
            "Name",
            "Phone",
            "Email",
            "Region",
            "Territory",
            "Country",
            "Status",
            "Members",
            "Contributions collected",
            "Subscribed share value",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
            ws.cell(row=1, column=col).font = Font(bold=True)
        sort = request.args.get("sort", "created_desc").strip()
        q = _agents_filtered_query()
        if sort == "name_asc":
            q = q.order_by(Agent.full_name.asc(), Agent.id.asc())
        elif sort == "name_desc":
            q = q.order_by(Agent.full_name.desc(), Agent.id.desc())
        elif sort == "agent_id":
            q = q.order_by(Agent.agent_id.asc())
        elif sort == "created_asc":
            q = q.order_by(Agent.created_at.asc(), Agent.id.asc())
        else:
            q = q.order_by(Agent.created_at.desc(), Agent.id.desc())
        rows_ag = q.all()
        batch = _batch_agent_metrics([x.id for x in rows_ag])
        for row, ag in enumerate(rows_ag, 2):
            m = batch.get(ag.id, {"members": 0, "contrib": Decimal("0"), "share": Decimal("0")})
            ws.cell(row=row, column=1, value=ag.agent_id)
            ws.cell(row=row, column=2, value=ag.full_name)
            ws.cell(row=row, column=3, value=ag.phone)
            ws.cell(row=row, column=4, value=ag.email)
            ws.cell(row=row, column=5, value=ag.region)
            ws.cell(row=row, column=6, value=ag.territory)
            ws.cell(row=row, column=7, value=ag.country)
            ws.cell(row=row, column=8, value=ag.status)
            ws.cell(row=row, column=9, value=m["members"])
            ws.cell(row=row, column=10, value=float(m["contrib"]))
            ws.cell(row=row, column=11, value=float(m["share"]))
        bio = io.BytesIO()
        wb.save(bio)
        return _xlsx_response(bio.getvalue(), "agents.xlsx")

    @app.route("/export/contributions.xlsx")
    def export_contributions_xlsx():
        wb = Workbook()
        ws = wb.active
        ws.title = "Contributions"
        headers = [
            "Transaction ID",
            "Receipt",
            "Member ID",
            "Member name",
            "Subscription",
            "Agent",
            "Amount",
            "Date",
            "Payment method",
            "Verified",
            "Notes",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
            ws.cell(row=1, column=col).font = Font(bold=True)
        q = (
            _contributions_filtered_query()
            .options(
                joinedload(Contribution.member).joinedload(Member.agent),
                joinedload(Contribution.subscription),
                joinedload(Contribution.payment_bank_account).joinedload(PaymentBankAccount.bank),
                joinedload(Contribution.payment_mobile_provider),
            )
            .order_by(Contribution.date.desc(), Contribution.id.desc())
        )
        for row, c in enumerate(q.all(), 2):
            agent_label = ""
            if c.member and c.member.agent:
                agent_label = c.member.agent.agent_id
            sub_no = ""
            if c.subscription_id and c.subscription:
                sub_no = c.subscription.subscription_no
            ws.cell(row=row, column=1, value=c.id)
            ws.cell(row=row, column=2, value=c.receipt_no)
            ws.cell(row=row, column=3, value=c.member.member_id)
            ws.cell(row=row, column=4, value=c.member.full_name)
            ws.cell(row=row, column=5, value=sub_no)
            ws.cell(row=row, column=6, value=agent_label)
            ws.cell(row=row, column=7, value=float(c.amount))
            ws.cell(row=row, column=8, value=c.date.isoformat() if c.date else "")
            ws.cell(row=row, column=9, value=c.payment_display_label())
            ws.cell(row=row, column=10, value="Yes" if c.verified else "No")
            ws.cell(row=row, column=11, value=(c.notes or "")[:500])
        bio = io.BytesIO()
        wb.save(bio)
        return _xlsx_response(bio.getvalue(), "contributions.xlsx")

    @app.route("/export/investments.xlsx")
    def export_investments_xlsx():
        r = forbid_agent()
        if r:
            return r
        wb = Workbook()
        ws = wb.active
        ws.title = "Investments"
        headers = [
            "Investment ID",
            "Code",
            "Name",
            "Project",
            "Type",
            "Amount invested",
            "Profit generated",
            "Start",
            "End",
            "Manager",
            "Status",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
            ws.cell(row=1, column=col).font = Font(bold=True)
        for row, inv in enumerate(
            _investments_filtered_query()
            .options(joinedload(Investment.project))
            .order_by(Investment.name)
            .all(),
            2,
        ):
            pname = inv.project.name if inv.project_id and inv.project else ""
            ws.cell(row=row, column=1, value=inv.id)
            ws.cell(row=row, column=2, value=inv.investment_code or "")
            ws.cell(row=row, column=3, value=inv.name)
            ws.cell(row=row, column=4, value=pname)
            ws.cell(row=row, column=5, value=inv.investment_type)
            ws.cell(row=row, column=6, value=float(inv.total_amount_invested or 0))
            ws.cell(row=row, column=7, value=float(inv.profit_generated or 0))
            ws.cell(row=row, column=8, value=inv.start_date.isoformat() if inv.start_date else "")
            ws.cell(row=row, column=9, value=inv.end_date.isoformat() if inv.end_date else "")
            ws.cell(row=row, column=10, value=inv.project_manager)
            ws.cell(row=row, column=11, value=inv.status)
        bio = io.BytesIO()
        wb.save(bio)
        return _xlsx_response(bio.getvalue(), "investments.xlsx")

    @app.route("/export/profit_distributions.xlsx")
    def export_profit_xlsx():
        wb = Workbook()
        ws = wb.active
        ws.title = "Profit distributions"
        headers = [
            "Date",
            "Investment",
            "Member ID",
            "Member",
            "Paid basis",
            "Share %",
            "Amount",
            "Pool",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
            ws.cell(row=1, column=col).font = Font(bold=True)
        rows = profit_rows_scope().order_by(ProfitDistribution.distribution_date.desc()).all()
        for row, r in enumerate(rows, 2):
            ws.cell(row=row, column=1, value=r.distribution_date.isoformat() if r.distribution_date else "")
            ws.cell(row=row, column=2, value=r.investment.name)
            ws.cell(row=row, column=3, value=r.member.member_id)
            ws.cell(row=row, column=4, value=r.member.full_name)
            ws.cell(row=row, column=5, value=float(r.eligible_amount_basis or 0))
            ws.cell(row=row, column=6, value=float(r.share_percentage or 0))
            ws.cell(row=row, column=7, value=float(r.amount))
            ws.cell(row=row, column=8, value=float(r.profit_pool_amount or 0))
        bio = io.BytesIO()
        wb.save(bio)
        return _xlsx_response(bio.getvalue(), "profit_distributions.xlsx")

    @app.route("/export/projects.xlsx")
    def export_projects_xlsx():
        wb = Workbook()
        ws = wb.active
        ws.title = "Projects"
        cat_map = dict(PROJECT_CATEGORIES)
        headers = [
            "Project code",
            "Name",
            "Category",
            "Start",
            "End",
            "Manager",
            "Status",
            "Budget",
            "Investment received",
            "Description",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
            ws.cell(row=1, column=col).font = Font(bold=True)
        filtered_projects = _projects_filtered_query().order_by(Project.name).all()
        for row, p in enumerate(filtered_projects, 2):
            ws.cell(row=row, column=1, value=p.project_code or "")
            ws.cell(row=row, column=2, value=p.name)
            ws.cell(row=row, column=3, value=cat_map.get(p.category, p.category))
            ws.cell(row=row, column=4, value=p.start_date.isoformat() if p.start_date else "")
            ws.cell(row=row, column=5, value=p.end_date.isoformat() if p.end_date else "")
            ws.cell(row=row, column=6, value=p.project_manager or "")
            ws.cell(row=row, column=7, value=p.status)
            ws.cell(row=row, column=8, value=float(p.total_budget or 0))
            ws.cell(row=row, column=9, value=float(p.total_investment_received()))
            ws.cell(row=row, column=10, value=(p.description or "")[:500])
        bio = io.BytesIO()
        wb.save(bio)
        return _xlsx_response(bio.getvalue(), "projects.xlsx")

    def _safe_pdf_text(s):
        return "".join(c if 32 <= ord(c) < 127 or c in "\t\n" else "?" for c in str(s))

    def _pdf_table(title, headers, rows):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, _safe_pdf_text(title), ln=True)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(0, 5, _safe_pdf_text(" | ".join(headers)), ln=True)
        pdf.set_font("Helvetica", "", 8)
        for line in rows:
            pdf.cell(0, 5, _safe_pdf_text(" | ".join(line)), ln=True)
        return pdf.output()

    @app.route("/export/members.pdf")
    def export_members_pdf():
        kind_labels = dict(MEMBER_KINDS)
        headers = ["Member ID", "Name", "Type", "Phone", "Agent", "Status", "Contributions", "Profit"]
        rows_data = []
        for mem in members_scope().order_by(Member.member_id).all():
            mk = getattr(mem, "member_kind", None) or "member"
            rows_data.append(
                [
                    mem.member_id,
                    mem.full_name[:40],
                    kind_labels.get(mk, mk),
                    mem.phone or "",
                    mem.agent.agent_id if mem.agent else "",
                    mem.status,
                    str(mem.contribution_total()),
                    str(mem.profit_received_total()),
                ]
            )
        pdf_bytes = _pdf_table("Estithmar — Members list", headers, rows_data)
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": "attachment; filename=members.pdf"},
        )

    @app.route("/export/contributions.pdf")
    def export_contributions_pdf():
        rows_data = []
        cq = (
            contributions_scope()
            .options(
                joinedload(Contribution.member).joinedload(Member.agent),
                joinedload(Contribution.payment_bank_account).joinedload(PaymentBankAccount.bank),
                joinedload(Contribution.payment_mobile_provider),
            )
            .order_by(Contribution.date.desc())
            .limit(500)
        )
        for c in cq.all():
            ag = ""
            if c.member and c.member.agent:
                ag = c.member.agent.agent_id
            rows_data.append(
                [
                    str(c.id),
                    c.receipt_no or "",
                    c.member.member_id,
                    str(c.amount),
                    c.date.isoformat() if c.date else "",
                    c.payment_display_label(),
                    ag,
                ]
            )
        pdf_bytes = _pdf_table(
            "Estithmar — Contributions",
            ["Txn ID", "Receipt", "Member", "Amount", "Date", "Payment", "Agent"],
            rows_data,
        )
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": "attachment; filename=contributions.pdf"},
        )

    @app.route("/audit")
    def audit_logs():
        if current_user.role != "admin":
            flash("Audit logs are only available to administrators.", "warning")
            return redirect(url_for("dashboard"))
        page = max(1, request.args.get("page", 1, type=int) or 1)
        per_page = min(max(10, request.args.get("per_page", 50, type=int) or 50), 200)
        total_count = _audit_filtered_query().count()
        total_pages = max(1, (total_count + per_page - 1) // per_page) if total_count else 1
        page = min(max(1, page), total_pages)
        logs = (
            _audit_filtered_query()
            .order_by(AuditLog.created_at.desc())
            .offset((page - 1) * per_page)
            .limit(per_page)
            .all()
        )

        entity_type_rows = (
            db.session.query(AuditLog.entity_type)
            .filter(AuditLog.entity_type.isnot(None), AuditLog.entity_type != "")
            .distinct()
            .order_by(AuditLog.entity_type)
            .all()
        )
        entity_type_choices = [t[0] for t in entity_type_rows if t[0]]

        action_rows = (
            db.session.query(AuditLog.action).distinct().order_by(AuditLog.action.asc()).limit(200).all()
        )
        action_choices = [a[0] for a in action_rows if a[0]]

        def _audit_list_qs(page_num: int) -> str:
            d = request.args.to_dict(flat=True)
            d["page"] = str(page_num)
            return url_for("audit_logs") + "?" + urlencode(d)

        qs_no_page = request.args.to_dict(flat=True)
        qs_no_page.pop("page", None)
        export_url = url_for("audit_export_csv") + "?" + urlencode(qs_no_page)
        next_page_url = _audit_list_qs(page + 1) if page < total_pages else None
        prev_page_url = _audit_list_qs(page - 1) if page > 1 else None
        first_page_url = _audit_list_qs(1) if page > 1 else None
        last_page_url = _audit_list_qs(total_pages) if total_pages > 1 and page < total_pages else None
        row_start = (page - 1) * per_page + 1 if total_count else 0
        row_end = min(page * per_page, total_count) if total_count else 0

        def _audit_merge_date_range(df: date, dt: date) -> str:
            d = request.args.to_dict(flat=True)
            d["date_from"] = df.isoformat()
            d["date_to"] = dt.isoformat()
            d.pop("page", None)
            return url_for("audit_logs") + "?" + urlencode(d)

        def _audit_drop_dates() -> str:
            d = request.args.to_dict(flat=True)
            d.pop("date_from", None)
            d.pop("date_to", None)
            d.pop("page", None)
            return url_for("audit_logs") + "?" + urlencode(d) if d else url_for("audit_logs")

        today_d = date.today()
        audit_preset_urls = {
            "today": _audit_merge_date_range(today_d, today_d),
            "7d": _audit_merge_date_range(today_d - timedelta(days=7), today_d),
            "30d": _audit_merge_date_range(today_d - timedelta(days=30), today_d),
            "clear_dates": _audit_drop_dates(),
        }

        return render_template(
            "audit/list.html",
            logs=logs,
            total_count=total_count,
            page=page,
            per_page=per_page,
            total_pages=total_pages,
            next_page_url=next_page_url,
            prev_page_url=prev_page_url,
            first_page_url=first_page_url,
            last_page_url=last_page_url,
            row_start=row_start,
            row_end=row_end,
            export_url=export_url,
            entity_type_choices=entity_type_choices,
            action_choices=action_choices,
            action_filter=(request.args.get("action") or "").strip(),
            entity_type_filter=(request.args.get("entity_type") or "").strip(),
            date_from=_parse_date(request.args.get("date_from")),
            date_to=_parse_date(request.args.get("date_to")),
            q=(request.args.get("q") or "").strip(),
            audit_href=_audit_href,
            audit_self_href=url_for("audit_logs"),
            audit_preset_urls=audit_preset_urls,
            fmt_time_ago=_fmt_time_ago,
        )

    @app.route("/audit/export.csv")
    def audit_export_csv():
        if current_user.role != "admin":
            flash("Audit export is only available to administrators.", "warning")
            return redirect(url_for("dashboard"))
        rows = _audit_filtered_query().order_by(AuditLog.created_at.desc()).limit(10000).all()
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["created_at_utc", "action", "entity_type", "entity_id", "details"])
        for r in rows:
            w.writerow(
                [
                    r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else "",
                    r.action or "",
                    r.entity_type or "",
                    r.entity_id if r.entity_id is not None else "",
                    (r.details or "").replace("\r\n", " ").replace("\n", " "),
                ]
            )
        return Response(
            "\ufeff" + buf.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=audit-log.csv"},
        )

    @app.route("/transactions")
    def transactions_redirect():
        return redirect(url_for("contributions_list"))

    @app.route("/invoices")
    @app.route("/invoices.html")
    def invoices_list():
        """Payment receipts as invoices: same underlying data as contributions, scoped and filterable."""
        member_id = request.args.get("member_id", type=int)
        agent_id = request.args.get("agent_id", type=int)
        subscription_id = request.args.get("subscription_id", type=int)
        date_from = _parse_date(request.args.get("date_from"))
        date_to = _parse_date(request.args.get("date_to"))
        verified = (request.args.get("verified") or "").strip().lower()
        qtext = (request.args.get("q") or "").strip()
        page = max(1, request.args.get("page", 1, type=int) or 1)
        per_page = min(max(10, request.args.get("per_page", 50, type=int) or 50), 200)

        total = _contributions_filtered_query().count()
        sum_scalar = (
            _contributions_filtered_query()
            .with_entities(func.coalesce(func.sum(Contribution.amount), 0))
            .scalar()
        )
        tot_dec = Decimal(str(sum_scalar or 0))
        rows_q = _contributions_filtered_query()
        total_pages = max(1, (total + per_page - 1) // per_page) if total else 1
        page = min(max(1, page), total_pages)
        rows = (
            rows_q.options(
                joinedload(Contribution.member).joinedload(Member.agent),
                joinedload(Contribution.subscription),
                joinedload(Contribution.payment_bank_account).joinedload(PaymentBankAccount.bank),
                joinedload(Contribution.payment_mobile_provider),
            )
            .order_by(Contribution.date.desc(), Contribution.id.desc())
            .offset((page - 1) * per_page)
            .limit(per_page)
            .all()
        )

        sub_balance_cache: dict[int, Decimal] = {}
        for c in rows:
            if c.subscription_id and c.subscription_id not in sub_balance_cache:
                sub = c.subscription
                sub_balance_cache[c.subscription_id] = (
                    sub.outstanding_balance() if sub else Decimal("0")
                )
        invoice_labels: dict[int, str] = {}
        for c in rows:
            if not c.subscription_id:
                invoice_labels[c.id] = "Paid"
            else:
                bal = sub_balance_cache.get(c.subscription_id, Decimal("0"))
                invoice_labels[c.id] = "Partially Paid" if bal > 0 else "Paid"

        members = members_scope().filter_by(status="Active").order_by(Member.full_name).all()
        agents = []
        if current_user.role != "agent":
            agents = Agent.query.filter_by(status="Active").order_by(Agent.full_name).all()
        sub_filter_q = ShareSubscription.query.join(Member, ShareSubscription.member_id == Member.id)
        if current_user.role == "agent" and current_user.agent_id:
            sub_filter_q = sub_filter_q.filter(Member.agent_id == current_user.agent_id)
        if current_user.role == "member" and current_user.member_id:
            sub_filter_q = sub_filter_q.filter(ShareSubscription.member_id == current_user.member_id)
        if member_id:
            sub_filter_q = sub_filter_q.filter(ShareSubscription.member_id == member_id)
        filter_subscriptions = sub_filter_q.order_by(ShareSubscription.subscription_date.desc()).limit(400).all()

        def _contrib_list_qs(page_num: int) -> str:
            d = request.args.to_dict(flat=True)
            d["page"] = str(page_num)
            return url_for("invoices_list") + "?" + urlencode(d)

        qs_no_page = request.args.to_dict(flat=True)
        qs_no_page.pop("page", None)
        export_url = url_for("export_contributions_xlsx") + "?" + urlencode(qs_no_page)
        contributions_list_url = url_for("contributions_list") + "?" + urlencode(qs_no_page)
        next_page_url = _contrib_list_qs(page + 1) if page < total_pages else None
        prev_page_url = _contrib_list_qs(page - 1) if page > 1 else None
        first_page_url = _contrib_list_qs(1) if page > 1 else None
        last_page_url = _contrib_list_qs(total_pages) if total_pages > 1 and page < total_pages else None
        row_start = (page - 1) * per_page + 1 if total else 0
        row_end = min(page * per_page, total) if total else 0

        settings = get_or_create_settings()
        extra = settings.get_extra()
        return render_template(
            "invoices.html",
            rows=rows,
            invoice_labels=invoice_labels,
            total_collected=tot_dec,
            total_count=total,
            members=members,
            member_id=member_id,
            agent_id=agent_id,
            subscription_id=subscription_id,
            filter_subscriptions=filter_subscriptions,
            agents=agents,
            date_from=date_from,
            date_to=date_to,
            verified_filter=verified,
            q=qtext,
            page=page,
            per_page=per_page,
            total_pages=total_pages,
            export_url=export_url,
            next_page_url=next_page_url,
            prev_page_url=prev_page_url,
            first_page_url=first_page_url,
            last_page_url=last_page_url,
            row_start=row_start,
            row_end=row_end,
            currency_code=settings.currency_code or "USD",
            settings=settings,
            extra=extra,
            contributions_list_url=contributions_list_url,
        )

    @app.route("/invoice-detail.html")
    def invoice_detail_legacy():
        invoice_id = request.args.get("id", type=int)
        if not invoice_id:
            flash("Invoice id is required.", "warning")
            return redirect(url_for("invoices_list"))
        return redirect(url_for("invoice_detail", id=invoice_id))

    @app.route("/invoice-detail/<int:id>")
    def invoice_detail(id):
        c = (
            Contribution.query.options(
                joinedload(Contribution.member).joinedload(Member.agent),
                joinedload(Contribution.subscription),
                joinedload(Contribution.verified_by),
                joinedload(Contribution.payment_bank_account).joinedload(PaymentBankAccount.bank),
                joinedload(Contribution.payment_mobile_provider),
            )
            .get_or_404(id)
        )
        if current_user.role == "agent" and current_user.agent_id and c.member.agent_id != current_user.agent_id:
            flash("Access denied.", "danger")
            return redirect(url_for("invoices_list"))
        if current_user.role == "member" and current_user.member_id and c.member_id != current_user.member_id:
            flash("Access denied.", "danger")
            return redirect(url_for("invoices_list"))
        invoice_no = c.receipt_no or f"INV-{c.id:05d}"
        status = "Paid"
        if c.subscription_id and c.subscription and c.subscription.outstanding_balance() > 0:
            status = "Partially Paid"
        settings = get_or_create_settings()
        extra = settings.get_extra()
        return render_template(
            "invoice-detail.html",
            c=c,
            invoice_no=invoice_no,
            status=status,
            settings=settings,
            extra=extra,
        )
